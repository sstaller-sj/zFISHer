import cv2
import nd2reader
from nd2reader import ND2Reader
import os
import numpy as np
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from PIL import Image, ImageTk
from PIL import ImageEnhance
from PIL import ImageGrab
import math
import sys
import shutil
from itertools import combinations
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import cell
from openpyxl.cell import _writer
from openpyxl import writer
import platform

#Second Newest IMPORTS SINCE LAST VERSION
import threading
import queue
import pystackreg
from pystackreg import StackReg
from skimage import io, exposure, color
from matplotlib import pyplot as plt
from tkinter import colorchooser

#Newest IMPORTS SINCE LAST VERSION
from PIL import Image, ImageDraw, ImageFont

turbomode = True

class MainApplication():
    def __init__(self, master):
        self.master = master
        self.newWindow = None
        self.toWelcomeGUI()

    def destroy_current_window(self):
        if self.newWindow is not None:
            self.newWindow.destroy()
            self.newWindow = None  # Reset newWindow to None
            self.app = None  # Reset app to None

    def toWelcomeGUI(self):
        if turbomode: 
            self.toOverlay()
            return
        self.newWindow = tk.Toplevel(self.master)
        self.app = WelcomeWindowGUI(self.newWindow)

    def toFileSelect(self):
        self.newWindow = tk.Toplevel(self.master)
        self.app = FileInputGUI(self.newWindow)

    def toOutputCombiner(self):
        self.newWindow = tk.Toplevel(self.master)
        self.app = OutputCombinerGUI(self.newWindow)

    def toInputProcessing(self):
        self.newWindow = tk.Toplevel(self.master)
        self.app = InputProcessingGUI(self.newWindow)

    def toOverlay(self):
        self.destroy_current_window()
        self.newWindow = tk.Toplevel(self.master)
        self.app = OverlayGUI(self.newWindow)

    def toNucpick(self):
        self.destroy_current_window()
        self.newWindow = tk.Toplevel(self.master)
        self.app = NucPickGUI(self.newWindow)

    def toROIpick(self):
        self.destroy_current_window()
        self.newWindow = tk.Toplevel(self.master)
        self.app = Picking_MAIN(self.newWindow)
    
    def toCalculations(self):
        self.destroy_current_window()
        self.newWindow = tk.Toplevel(self.master)
        self.app = Calculations(self.newWindow)      

    def toOutputExport(self):
       # self.destroy_current_window()
        self.newWindow = tk.Toplevel(self.master)
        self.app = WriteExcel(self.newWindow)    
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class WelcomeWindowGUI():
    def __init__(self, master):
        self.master = master
        self.master.title("ZFISHER --- Welcome")
        self.master.geometry("500x500")
        self.master.config(bg="lightblue")

        
        # Configure the grid to center the buttons
        master.grid_rowconfigure(0, weight=1)
        master.grid_rowconfigure(1, weight=1)
        master.grid_rowconfigure(2, weight=1)
        master.grid_columnconfigure(0, weight=1)
        master.grid_columnconfigure(1, weight=1)
        master.grid_columnconfigure(2, weight=1)

        self.titlelabel = tk.Label(master, text="Welcome to ZFISHER!", background="black", foreground="white", font=("Helvetica", 35))
        self.titlelabel.grid(row=0, column=1, sticky="nsew", padx=0, pady=20)
        # Create buttons to open file dialogs for File 1 and File 2
        self.imageprocessbutton = tk.Button(master, text="Image Process 2 nd2's", command=self.tofileselect, background="black", foreground="green", font=("Helvetica", 20))
        self.imageprocessbutton.grid(row=1, column=1, sticky="nsew", padx=30, pady=30)

        #Combine 2+ outputs button
        self.ocombbut = tk.Button(master, text="Combine ZFISHER Outputs", command=self.outcombiner, background="black", foreground='red', font=("Helvetica", 20))
        self.ocombbut.grid(row=2, column=1, sticky="nsew", padx=30, pady=50)
        #Create Finish Button
        #self.finishbutton = tk.Button(master, text="Leave", command=self.finish_file_select)
        #self.finishbutton.grid(row=2, column=2)

    def tofileselect(self):
        mainapp.toFileSelect()
        self.master.destroy()
    
    def outcombiner(self):
        mainapp.toOutputCombiner()
        self.master.destroy()
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class FileInputGUI():
    def __init__(self, master):

        inputfile1path = ''
        inputfile2path = ''

        # Create the main application window
        #root = tk.Tk()
        #root.title("nd2 File Selection")

        self.master = master
        self.master.title("ZFISHER -- nd2 File Selection")
        

        # Create labels for File 1 and File 2
        self.headerlabel = tk.Label(master, text="Select nd2 Files to be Analyzed")
        self.headerlabel.grid(row=0, column=0, columnspan=2, sticky="nsew")

        self.f1label = tk.Label(master, text="File 1:")
        self.f1label.grid(row=1, column=0, sticky="w")

        self.f2label = tk.Label(master, text="File 2:")
        self.f2label.grid(row=2, column=0, sticky="w")

        # Create entry widgets to display the file paths
        self.f1entry = tk.Entry(master, width=50)
        self.f1entry.grid(row=1, column=1, sticky="e")

        self.f2entry = tk.Entry(master, width=50)
        self.f2entry.grid(row=2, column=1, sticky="e")

        # Create buttons to open file dialogs for File 1 and File 2
        self.f1button = tk.Button(master, text="Browse", command=self.open_file1)
        self.f1button.grid(row=1, column=2)

        self.f2button = tk.Button(master, text="Browse", command=self.open_file2)
        self.f2button.grid(row=2, column=2)

        #Create F1 and F2 nametags
        self.f1nametaglabel = tk.Label(master, text="File 1 Nametag:")
        self.f1nametaglabel.grid(row=3, column=0, sticky="w")     
        self.f1nametagentry = tk.Entry(master, width=50)
        self.f1nametagentry.grid(row=3, column=1, sticky="e")
        self.f1nametagentry.delete(0,tk.END)
        self.f1nametagentry.insert(0,"F1")

        self.f2nametaglabel = tk.Label(master, text="File 2 Nametag:")
        self.f2nametaglabel.grid(row=4, column=0, sticky="w")     
        self.f2nametagentry = tk.Entry(master, width=50)
        self.f2nametagentry.grid(row=4, column=1, sticky="e")
        self.f2nametagentry.delete(0,tk.END)
        self.f2nametagentry.insert(0,"F2")

        #Create Finish Button
        self.finishbutton = tk.Button(master, text="Finish", command=self.finish_file_select)
        self.finishbutton.grid(row=5, columnspan=3)

    def open_file1(self):
        filepath = filedialog.askopenfilename()
        self.f1entry.delete(0, tk.END)
        self.f1entry.insert(0, filepath)

    def open_file2(self):
        filepath = filedialog.askopenfilename()
        self.f2entry.delete(0, tk.END)
        self.f2entry.insert(0, filepath)

    def finish_file_select(self):
        #Send File paths to DynData
        self.inputfile1path = self.f1entry.get()
        self.inputfile2path = self.f2entry.get()
        dyn_data.f1inputpath = self.inputfile1path
        dyn_data.f2inputpath = self.inputfile2path

        #Check Nametags and Send to DynData
        self.f1nametag = self.f1nametagentry.get()
        self.f2nametag = self.f2nametagentry.get()
        if not self.f1nametag:
            print("F1 nametag = empty")
            self.f1nametag = "F1"
        if not self.f2nametag:
            print("F2 nametag = empty")
            self.f1nametag = "F2"
        dyn_data.f1inputpath = self.inputfile1path
        dyn_data.f2inputpath = self.inputfile2path

        dyn_data.f1nametag = self.f1nametag
        dyn_data.f2nametag = self.f2nametag

        print(f"FILE 1: {self.inputfile1path}")
        print(f"FILE 2: {self.inputfile2path}")
        print(f"FILE 1 nametag: {self.f1nametag}")
        print(f"FILE 2 nametag: {self.f2nametag}")

        mainapp.toInputProcessing()
        self.master.destroy()
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class OutputCombinerGUI():  # NOT WORKING
    def __init__(self, master):
        self.master = master
        self.master.title("ZFISHER --- Combine 2+ Output Files")
        # self.master.geometry("500x500")
        self.master.config(bg="violet")

        style = ttk.Style()
        style.theme_use('clam')
        
        # Configure the grid to center the buttons
        master.grid_rowconfigure(0, weight=1)
        master.grid_rowconfigure(1, weight=1)
        master.grid_rowconfigure(2, weight=1)
        master.grid_rowconfigure(3, weight=1)
        master.grid_columnconfigure(0, weight=1)
        master.grid_columnconfigure(1, weight=1)
        master.grid_columnconfigure(2, weight=1)
        master.grid_columnconfigure(3, weight=1)

        self.titlelabel = tk.Label(master, text="Combine Multiple Outputs into a Single File (NOT FUNCTIONAL)", background="darkblue", foreground="white", font=("Helvetica", 15))
        self.titlelabel.grid(row=0, column=0, sticky="nsew", padx=10, pady=20, columnspan=3)

        # Add a label to the left of the entry box
        self.fcountlabel = tk.Label(master, text="# of Files to Combine:", background="violet", font=("Helvetica", 16))
        self.fcountlabel.grid(row=1, column=0, sticky="e", padx=10, pady=5)

        self.fcountentry = tk.Entry(master, width=10)
        self.fcountentry.insert(0, "2")
        self.fcountentry.grid(row=1, column=1)

        self.fcountconfirmbutton = tk.Button(master, text="Confirm", command=self.fcountconfirm, background="black", foreground="red", font=("Helvetica", 20))
        self.fcountconfirmbutton.grid(row=1, column=2, sticky="nsew",padx=10, pady=5)

        self.diag = tk.Label(master,text="Define # of files to combine", background="white", foreground="black", font=("Helvetica", 12))
        self.diag.grid(row=2, column=0, sticky="nsew", padx=10, pady=5, columnspan=3)

        #Dictionary to store file widgets
        self.file_widgets = {}
    def fcountconfirm(self):
        fcount_inp = self.fcountentry.get()
        if fcount_inp.isdigit(): 
            self.diag.config(text=f"Select paths for the {fcount_inp} output files",foreground="green" )
            self.fselectframe()
        else: self.diag.config(text=f"Invalid, must input integer",foreground="red")

    def fselectframe(self):
        fselframe = tk.Frame(self.master)
        fselframe.grid(row=3, column=0, sticky="nsew", padx=10, pady=5, columnspan=3)

        fcountstr = float(self.fcountentry.get())
        fcount = round(fcountstr)

        self.file_widgets.clear()
        for i in range(0, fcount):
            label = tk.Label(fselframe, text=f"#{i+1}")
            label.grid(row=i, column=0)
            
            entry = tk.Entry(fselframe, width=50)
            entry.grid(row=i, column=1, sticky="e")
            
            self.button = tk.Button(fselframe, text="Browse", command= lambda i=i: self.open_file(i))  # Define the command later
            self.button.grid(row=i, column=2)
            
            # Store widgets in the dictionary with a unique key
            self.file_widgets[i] = {'label': label, 'entry': entry, 'button': self.button}

        #Create Finish Button
        self.finishbutton = tk.Button(fselframe, text="Combine", command=self.finish_file_select)
        self.finishbutton.grid(row=fcount+1, column=0, columnspan= 3)
    
        print(self.file_widgets)

        print(self.file_widgets[0]['label'])
        print(self.file_widgets[0]['entry'])
        print(self.file_widgets[0]['button'])

    def open_file(self,id):
        print(self.file_widgets[id]['label'])
        print(self.file_widgets[id]['entry'])
        print(self.file_widgets[id]['button'])   

        filepath = filedialog.askopenfilename()
        self.file_widgets[id]['entry'].delete(0, tk.END)
        self.file_widgets[id]['entry'].insert(0, filepath)    
        

    def outcombiner(self):
        #wincontroller.start_outputcombiner()
        pass

    def finish_file_select(self):
        for key, widget_dict in self.file_widgets.items():
            entry_widget = widget_dict.get('entry')
            if entry_widget:
                print(f"{key} - {entry_widget.get()}")


        #wincontroller.start_outputcombiner()
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class InputProcessingGUI():
    def __init__(self, master):
        self.master = master
        self.master.title("ZFISHER --- Input Processing")
    
        # Header Label
        self.headerlabel = tk.Label(master, text="Nd2 Files are Being Processed...")
        self.headerlabel.grid(row=0, column=0, columnspan=2, sticky="nsew")

        # Progress Bars
        self.progress1 = tk.IntVar()
        self.progress_label1 = tk.Label(master, text="File 1: 0%")
        self.progress_label1.grid(row=1, column=0, sticky="nsew")
        self.progressbar1 = ttk.Progressbar(master, orient="horizontal", length=200, mode="determinate", variable=self.progress1)
        self.progressbar1.grid(row=2, column=0, sticky="nsew")

        self.progress2 = tk.IntVar()
        self.progress_label2 = tk.Label(master, text="File 2: 0%")
        self.progress_label2.grid(row=1, column=1, sticky="nsew")
        self.progressbar2 = ttk.Progressbar(master, orient="horizontal", length=200, mode="determinate", variable=self.progress2)
        self.progressbar2.grid(row=2, column=1, sticky="nsew")

        self.progressMIP = tk.IntVar()
        self.progressMIP_label = tk.Label(master, text="MIP Generation Progress: 0%")
        self.progressMIP_label.grid(row=10, column=0, sticky="e")
        self.progressbarMIP = ttk.Progressbar(master, orient="horizontal", length=200, mode="determinate", variable=self.progressMIP)
        self.progressbarMIP.grid(row=10, column=1, sticky="w")

        # Modifiable Directory Names
        self.inputs_directory_folder = "Inputs"
        self.inputs_directory_F1 = "FILE_1"
        self.inputs_directory_F2 = "FILE_2"
        self.processing_directory_folder = "Processing"
        self.p_MIP_directory = "MIP"
        self.p_MIP_raw_directory = "RAW_MIP"
        self.p_MIP_masked_directory = "Masked"
        self.p_zslices_directory = "zslices"
        self.outputs_directory_folder = "Outputs"

        # File Metadata
        self.f1numchan = "?"
        self.f2numchan = "?"
        self.f1ntag = "name"
        self.f2ntag = "name"
        self.f1filepath = dyn_data.f1inputpath  # replace with actual path
        self.f2filepath = dyn_data.f2inputpath  # replace with actual path
        self.f1numslices = 10  # replace with actual number
        self.f2numslices = 10  # replace with actual number
        self.f1channels = ["ch1", "ch2"]  # replace with actual channels
        self.f2channels = ["ch1", "ch2"]  # replace with actual channels

        self.f1perc = 0
        self.f2perc = 0

        self.f1pmax = 0
        self.f2pmax = 0
        self.f1pdone = 0
        self.f2pdone = 0

        self.chanframemaker()
        
        # Create a queue for progress updates
        self.progress_queue = queue.Queue()

        # Start metadata_processor and create_directories in separate threads
        self.thread1 = threading.Thread(target=self.metadata_processor)
        self.thread2 = threading.Thread(target=self.create_directories)

        self.thread1.start()
        self.thread2.start()

        # Check if threads are done
        self.check_threads()

    def check_threads(self):
        if self.thread1.is_alive() or self.thread2.is_alive():
            self.master.after(100, self.check_threads)  # Check again after 100ms
        else:
            print("Starting Z slicing")
            self.zsliceloader()

    def zslicer(self, file, cindex, cname, numslices, fIND):
        with ND2Reader(file) as nd2_file:
            nd2_file.iter_axes = 'z'
            nd2_file.default_coords['c'] = cindex
            export_dir = os.path.join(self.processing_directory_folder, f"FILE_{fIND}", self.p_zslices_directory, f"C{cindex}_{cname}")
            os.makedirs(export_dir, exist_ok=True)  # Ensure the export directory exists
            for frame_index in range(numslices):
                z_level_data = nd2_file[frame_index]
                z_slice_img_output = z_level_data
                output_name = f"F{fIND}_C{cindex}_{cname}_z{frame_index+1}_.tif"
                output_path = os.path.join(export_dir, output_name)
                cv2.imwrite(output_path, z_slice_img_output)
                
                # Calculate and send progress update
                progress = int(((frame_index + 1) / numslices) * 100)
                self.progress_queue.put((fIND, progress))
        
        
    def zsliceloader(self):
        self.headerlabel.config(text="Slicing Zstacks of Each nd2...")
        self.threads = []
        # FILE1
        for cindex, chan in enumerate(self.f1channels):
            thread = threading.Thread(target=self.zslicer, args=(self.f1filepath, cindex, chan, self.f1numslices, 1))
            self.threads.append(thread)
        # FILE2
        for cindex, chan in enumerate(self.f2channels):
            thread = threading.Thread(target=self.zslicer, args=(self.f2filepath, cindex, chan, self.f2numslices, 2))
            self.threads.append(thread)

        # Start all threads
        for thread in self.threads:
            thread.start()

        # Start a thread to monitor progress
        self.progress_thread = threading.Thread(target=self.update_progress)
        self.progress_thread.start()

    def miploader(self):
        self.headerlabel.config(text="Generating MIP of Each nd2...")
        self.mipthreads = []
        self.mipchancount = len(self.f1channels) + len(self.f2channels)
        self.mipmakecounter = 0

        for cindex, chan in enumerate(self.f1channels):
            mipthread = threading.Thread(target=self.mipmaker, args=(self.f1filepath, cindex, chan, self.f1numslices, 1))
            self.mipthreads.append(mipthread)


        for cindex, chan in enumerate(self.f2channels):
            mipthread = threading.Thread(target=self.mipmaker, args=(self.f2filepath, cindex, chan, self.f1numslices, 2))
            self.mipthreads.append(mipthread)



        for mipthread in self.mipthreads:
            mipthread.start()

        
        # Start a thread to monitor progress
        self.progress_thread = threading.Thread(target=self.update_mipprogress)
        self.progress_thread.start()

    def mipmaker(self, file, cindex, cname, numslices, fIND):
        with ND2Reader(file) as nd2_file:
            nd2_file.iter_axes = 'z'
            nd2_file.default_coords['c'] = cindex
            frames = [frame for frame in nd2_file]
            frame_stack = np.stack(frames, axis=0)
            MIP_z_output = np.max(frame_stack, axis=0)
            
            #ADJUSTED MIP
            MIP_z_cooked_output = MIP_z_output*255
            output_img = MIP_z_cooked_output
            export_dir = os.path.join(self.processing_directory_folder, f"FILE_{fIND}", self.p_MIP_directory, f"C{cindex}_{cname}")
            output_name = os.path.join(f"F{fIND}_C{cindex}_{cname}_MIP_.tif")          
            output_path = os.path.join(export_dir,output_name)
            cv2.imwrite(output_path,output_img)
            #RAW MIP
            MIP_z_raw_output = MIP_z_output
            output_img = MIP_z_raw_output
            export_dir = os.path.join(self.processing_directory_folder, f"FILE_{fIND}", self.p_MIP_raw_directory, f"C{cindex}_{cname}")
            output_name = os.path.join(f"F{fIND}_C{cindex}_{cname}_MIP_.tif")          
            output_path = os.path.join(export_dir,output_name)
            cv2.imwrite(output_path,output_img)

        self.mipmakecounter += 1
        progress_percent = (self.mipmakecounter / self.mipchancount) * 100
        self.progressMIP.set(progress_percent)
        self.progressMIP_label.config(text=f"MIP Generation Progress:  {int(progress_percent)}%")
        self.master.update_idletasks()

            
    def update_progress(self):
        while any(thread.is_alive() for thread in self.threads):
            try:
                file_index, progress = self.progress_queue.get(timeout=0.1)
                if file_index == 1:
                    self.progress1.set(progress)
                    self.progress_label1.config(text=f"File 1: {progress}%")
                elif file_index == 2:
                    self.progress2.set(progress)
                    self.progress_label2.config(text=f"File 2: {progress}%")
            except queue.Empty:
                continue

        print("All threads have finished")
        self.headerlabel.config(text="Input nd2(s) Processing COMPLETE.")
        #self.continue_button()
        #self.finishprocessing()
        self.miploader()

    def update_mipprogress(self):
        #track when mip files are completed
        while any(mipthread.is_alive() for mipthread in self.mipthreads):
            try:
                file_index, progress = self.progress_queue.get(timeout=0.1)
                if file_index == 1:
                    self.progress1.set(progress)
                    self.progress_label1.config(text=f"File 1: {progress}%")
                elif file_index == 2:
                    self.progress2.set(progress)
                    self.progress_label2.config(text=f"File 2: {progress}%")
            except queue.Empty:
                continue

        print("All MIP threads have finished")
        self.headerlabel.config(text="Input All MIP threads have finished(s) Processing COMPLETE.")
        self.progressMIP_label.config(text=f"MIP Generation Progress: 100%")
        self.master.update_idletasks()
        self.continue_button()
        #self.finishprocessing()
        #self.finishprocessing()

    def continue_button(self):
        #continuebutton = tk.Button(self.master, background="white", text="CONTINUE", font=("Helvetica", 20), command=self.finishprocessing)
        #continuebutton.grid(row=5,column=0,padx=20,pady=20, columnspan=2)    

        #auto move forward

        self.finishprocessing()



    def metadata_processor(self):
        print("METADATA PROCESSOR")
        #FILE1
        with ND2Reader(self.f1filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f1numchan = len(nd2_file.metadata['channels'])
            self.f1numslices = len(nd2_file)
            self.f1channels = nd2_file.metadata['channels']

            self.f1pmax = float(self.f1numchan*self.f1numslices)

        #FILE2
        with ND2Reader(self.f2filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f2numchan = len(nd2_file.metadata['channels'])
            self.f2numslices = len(nd2_file)
            self.f2channels = nd2_file.metadata['channels']

            self.f2pmax = float(self.f2numchan*self.f2numslices)

        dyn_data.f1channels = self.f1channels
        dyn_data.f2channels = self.f2channels
        dyn_data.f1numslices = self.f1numslices
        dyn_data.f2numslices = self.f2numslices
        self.f1ntag = dyn_data.f1nametag
        self.f2ntag = dyn_data.f2nametag

        self.updatechanframe()
    def create_directories(self):
        print("MAKE DIRECTORIES")
        self.headerlabel.config(text="Creating File Directories...")
        #----------------------------------------------------------------------------------------------------#
        #INPUTS
        #----------------------------------------------------------------------------------------------------#
        if not os.path.exists(self.inputs_directory_folder):  os.makedirs(self.inputs_directory_folder)  #Create Outputs directory
        self.empty_directory(self.inputs_directory_folder)     #Empty outputs directory
        i_f1_dir = os.path.join(self.inputs_directory_folder, self.inputs_directory_F1)
        i_f2_dir = os.path.join(self.inputs_directory_folder,self.inputs_directory_F2)
        os.makedirs(i_f1_dir)
        os.makedirs(i_f2_dir)
        shutil.copyfile(self.f1filepath, os.path.join(i_f1_dir, os.path.basename(self.f1filepath)))
        shutil.copyfile(self.f2filepath, os.path.join(i_f2_dir, os.path.basename(self.f2filepath)))
        i_f1 = os.listdir(i_f1_dir)[0]
        i_f2 = os.listdir(i_f2_dir)[0]
        self.input_nd2_files = [i_f1,i_f2]
        self.input_nd2_files_count = len(self.input_nd2_files)
        #----------------------------------------------------------------------------------------------------#
        #OUTPUTS
        #----------------------------------------------------------------------------------------------------#
        if not os.path.exists(self.outputs_directory_folder):  os.makedirs(self.outputs_directory_folder)  #Create Outputs directory
        self.empty_directory(self.outputs_directory_folder)     #Empty outputs directory
        file_ind_dirname = f"FILE_1"
        p_file_dir = os.path.join(self.outputs_directory_folder, file_ind_dirname)
        os.makedirs(p_file_dir)
        file_ind_dirname = f"FILE_2"
        p_file_dir = os.path.join(self.outputs_directory_folder, file_ind_dirname)
        os.makedirs(p_file_dir)
        file_ind_dirname = f"IMAGES"
        p_file_dir = os.path.join(self.outputs_directory_folder, file_ind_dirname)
        os.makedirs(p_file_dir)
        #----------------------------------------------------------------------------------------------------#
        #PROCESSING
        #----------------------------------------------------------------------------------------------------#
        if not os.path.exists(self.processing_directory_folder):  os.makedirs(self.processing_directory_folder)    #Create Processing directory
        self.empty_directory(self.processing_directory_folder) 

        #FILE1
        file_ind_dirname = f"FILE_1"
        p_file_dir = os.path.join(self.processing_directory_folder, file_ind_dirname)
        os.makedirs(p_file_dir)
        #F1MIP
        p_file_MIP_dir = os.path.join(p_file_dir, self.p_MIP_directory) 
        os.makedirs(p_file_MIP_dir)
        for index,chan in enumerate(self.f1channels):
            os.makedirs(os.path.join(p_file_MIP_dir, f"C{index}_{chan}")  )
        #F1RAWMIP
        p_file_MIP_raw_dir = os.path.join(p_file_dir, self.p_MIP_raw_directory) 
        os.makedirs(p_file_MIP_raw_dir)
        for index,chan in enumerate(self.f1channels):
            os.makedirs(os.path.join(p_file_MIP_raw_dir, f"C{index}_{chan}")  )
        #F1MASKED
        p_file_MIP_masked_dir = os.path.join(p_file_dir, self.p_MIP_masked_directory) 
        os.makedirs(p_file_MIP_masked_dir)
        for index,chan in enumerate(self.f1channels):
            os.makedirs(os.path.join(p_file_MIP_masked_dir, f"C{index}_{chan}")  )
        #F1ZSLICES
        p_file_ZSLICE_dir = os.path.join(p_file_dir, self.p_zslices_directory) 
        os.makedirs(p_file_ZSLICE_dir)
        for index,chan in enumerate(self.f1channels):
            os.makedirs(os.path.join(p_file_ZSLICE_dir, f"C{index}_{chan}")  )
        #FILE2
        file_ind_dirname = f"FILE_2"
        p_file_dir = os.path.join(self.processing_directory_folder, file_ind_dirname)
        os.makedirs(p_file_dir)
        #F1MIP
        p_file_MIP_dir = os.path.join(p_file_dir, self.p_MIP_directory) 
        os.makedirs(p_file_MIP_dir)
        for index,chan in enumerate(self.f2channels):
            os.makedirs(os.path.join(p_file_MIP_dir, f"C{index}_{chan}")  )
        #F2RAWMIP
        p_file_MIP_raw_dir = os.path.join(p_file_dir, self.p_MIP_raw_directory) 
        os.makedirs(p_file_MIP_raw_dir)
        for index,chan in enumerate(self.f1channels):
            os.makedirs(os.path.join(p_file_MIP_raw_dir, f"C{index}_{chan}")  )
        #F2MASKED
        p_file_MIP_masked_dir = os.path.join(p_file_dir, self.p_MIP_masked_directory) 
        os.makedirs(p_file_MIP_masked_dir)
        for index,chan in enumerate(self.f1channels):
            os.makedirs(os.path.join(p_file_MIP_masked_dir, f"C{index}_{chan}")  )
        #F2ZSLICES
        p_file_ZSLICE_dir = os.path.join(p_file_dir, self.p_zslices_directory) 
        os.makedirs(p_file_ZSLICE_dir)
        for index,chan in enumerate(self.f1channels):
            os.makedirs(os.path.join(p_file_ZSLICE_dir, f"C{index}_{chan}")  )
    def empty_directory(self,directory):
        for item in os.listdir(directory):
            item_path = os.path.join(directory, item)
            if os.path.isfile(item_path):
                os.remove(item_path)
            elif os.path.isdir(item_path):
                self.empty_directory(item_path)
                os.rmdir(item_path)
    def chanframemaker(self):
        #Make F1/F2 frames
        f1frame = tk.Frame(self.master, background="black")
        f1frame.grid(row=3,column=0,padx=10)
        self.f1_header_l = tk.Label(f1frame, text="File 1", background="black",foreground="white")
        self.f1_header_l.grid(row=2, column=0, padx=5, pady=5)
        self.f1_ntag_l = tk.Label(f1frame, text=f"Nametag: {self.f1ntag}", background="black",foreground="white")
        self.f1_ntag_l.grid(row=4, column=0, padx=5, pady=5)
        self.f1_filepath_l = tk.Label(f1frame, text=f"Path: {self.f1filepath}", background="black",foreground="white")
        self.f1_filepath_l.grid(row=6, column=0, padx=5, pady=5)
        self.f1_numchan_l = tk.Label(f1frame, text=f"# Channels: {self.f1numchan}", background="black",foreground="white")
        self.f1_numchan_l.grid(row=8, column=0, padx=5, pady=5)
        self.f1_numslices_l = tk.Label(f1frame, text=f"# Zslices: {self.f1numslices}", background="black",foreground="white")
        self.f1_numslices_l.grid(row=10, column=0, padx=5, pady=5)
        self.f1_channels_l = tk.Label(f1frame, text=f"Channels: {self.f1channels}", background="black",foreground="white")
        self.f1_channels_l.grid(row=12, column=0, padx=5, pady=5)
        self.f1_perc_l = tk.Label(f1frame, text=f"%: {self.f1perc}", background="black",foreground="white")
        self.f1_perc_l.grid(row=14, column=0, padx=5, pady=5)

        f2frame = tk.Frame(self.master, background="white")
        f2frame.grid(row=3,column=1,padx=10)
        self.f2_header = tk.Label(f2frame, text="File 2", background="white",foreground="black")
        self.f2_header.grid(row=2, column=0, padx=5, pady=5)
        self.f2_ntag_l = tk.Label(f2frame, text=f"Nametag: {self.f2ntag}", background="black",foreground="white")
        self.f2_ntag_l.grid(row=4, column=0, padx=5, pady=5)
        self.f2_filepath_l = tk.Label(f2frame, text=f"Path: {self.f2filepath}", background="black",foreground="white")
        self.f2_filepath_l.grid(row=6, column=0, padx=5, pady=5)
        self.f2_numchan_l = tk.Label(f2frame, text=f"# Channels: {self.f2numchan}", background="black",foreground="white")
        self.f2_numchan_l.grid(row=8, column=0, padx=5, pady=5)
        self.f2_numslices_l = tk.Label(f2frame, text=f"# Zslices: {self.f2numslices}", background="black",foreground="white")
        self.f2_numslices_l.grid(row=10, column=0, padx=5, pady=5)
        self.f2_channels_l = tk.Label(f2frame, text=f"Channels: {self.f2channels}", background="black",foreground="white")
        self.f2_channels_l.grid(row=12, column=0, padx=5, pady=5)
        self.f2_perc_l = tk.Label(f2frame, text=f"%: {self.f2perc}", background="black",foreground="white")
        self.f2_perc_l.grid(row=14, column=0, padx=5, pady=5)

    def updatechanframe(self):
        #F1FRAME
        self.f1_ntag_l.config(text=f"Nametag: {self.f1ntag}")
        self.f1_filepath_l.config(text=f"File: {self.f1filepath}")
        self.f1_numchan_l.config(text=f"# Channels: {self.f1numchan}")
        self.f1_numslices_l.config(text=f"# Zslices: {self.f1numslices}")
        self.f1_channels_l.config(text=f"Channels: {self.f1channels}")

        f1chantext = f"Channels: \n"
        for index,chan in enumerate(self.f1channels):
            f1chantext += f"C{index}: {chan} \n"

        self.f1_channels_l.config(text=f"{f1chantext}")



        #F2FRAME
        self.f2_ntag_l.config(text=f"Nametag: {self.f2ntag}")
        self.f2_filepath_l.config(text=f"File: {self.f2filepath}")
        self.f2_numchan_l.config(text=f"# Channels: {self.f2numchan}")
        self.f2_numslices_l.config(text=f"# Zslices: {self.f2numslices}")
        self.f2_channels_l.config(text=f"Channels: {self.f2channels}")

        f2chantext = f"Channels: \n"
        for index,chan in enumerate(self.f2channels):
            f2chantext += f"C{index}: {chan} \n"

        self.f2_channels_l.config(text=f"{f2chantext}")
    def finishprocessing(self):
        
        mainapp.toOverlay()
        self.master.destroy()
        
       # self.nucmask__init__()

        #mainapp.toOverlay()
        #self.master.destroy()


    def nucmask__init__(self):

        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                self.f1dapichan = index
                self.F1_DAPI_dir = f"C{self.f1dapichan}_DAPI"

        self.arr_nucleus_contours = None    
        
        

        #Get Zslice from C0 of F1 for 
        self.slice_ismask = self.get_masking_slice()

        self.mip_tomask = self.get_MIP_tomask()
        
        self.generate_nucleus_mask(self.slice_ismask, self.mip_tomask)

    def get_MIP_tomask(self):

        #self.mip_tomask_dir = os.path.join(self.processing_directory_folder,"FILE_1",self.p_MIP_directory,self.chanidstring)
        self.mip_tomask_dir = os.path.join(self.processing_directory_folder,"FILE_1","RAW_MIP",self.F1_DAPI_dir)

        mip_tomask_list = os.listdir(self.mip_tomask_dir)
        mip_tomask = os.path.join(self.mip_tomask_dir,mip_tomask_list[0])

        return mip_tomask

    def get_masking_slice(self):


        #Get folder for zslice mask, get number of slices, pick a slice in the middle
        self.nucleus_channel_zslice_path = os.path.join(self.processing_directory_folder,"FILE_1",self.p_zslices_directory,self.F1_DAPI_dir)
        sliceslist = os.listdir(self.nucleus_channel_zslice_path)
        #slice_formask = sliceslist[round((len(sliceslist))/2)]

        entries = os.listdir(self.nucleus_channel_zslice_path)
        # Filter out directories and count the number of files
        num_files = len([entry for entry in entries if os.path.isfile(os.path.join(self.nucleus_channel_zslice_path, entry))])
        half_num_files = round(num_files/2)
        
        slice_formask =f"F1_{self.F1_DAPI_dir}_z{half_num_files}_.tif"
        slice_formask_path = os.path.join(self.nucleus_channel_zslice_path,slice_formask)
        print(f"Slice used to mask: {slice_formask_path}")
        return slice_formask_path

    def generate_nucleus_mask(self,dapi_nucleus_mask_input,dna_mip_input):

        dapi_nucleus_base_img_in = cv2.imread(dapi_nucleus_mask_input, cv2.IMREAD_UNCHANGED)
        dapi_nucleus_base_img = dapi_nucleus_base_img_in * 255

        cv2.imwrite('dapinucleusbaseimg.tif', dapi_nucleus_base_img_in )
        
        dapi_nucleus_base_img_8bit = cv2.convertScaleAbs(dapi_nucleus_base_img_in, alpha=(255.0/65535.0))
        cv2.imwrite('dapinucleusbaseimg2.tif', dapi_nucleus_base_img_8bit )

        dapi_nucleus_base_img = dapi_nucleus_base_img_8bit

        dna_mip_img = cv2.imread(dna_mip_input, cv2.IMREAD_GRAYSCALE)
        gblur_img = cv2.GaussianBlur(dapi_nucleus_base_img,(35,35),0)
        ret3, nucleus_threshold_img = cv2.threshold(gblur_img,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)

        # Convert nucleus_threshold_mask to uint8 if necessary
        nucleus_threshold_mask = nucleus_threshold_img.astype(np.uint8)

        # Ensure the size of the mask matches the size of the source image
        assert dna_mip_img.shape[:2] == nucleus_threshold_mask.shape[:2], "Image and mask must have the same size"

        # Apply the mask
        masked_dna_img = cv2.bitwise_and(dna_mip_img, dna_mip_img, mask=nucleus_threshold_mask)


        output_path = os.path.join(self.processing_directory_folder, f"masked_dna_mip.tif")    
        cv2.imwrite(output_path,masked_dna_img)

        #----count cells--------

        edged_img = cv2.Canny(nucleus_threshold_img, 30,150)

        dna_mip_base = cv2.imread(dna_mip_input, cv2.COLOR_BGR2GRAY)
        contours, hierarchy = cv2.findContours(nucleus_threshold_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        contoured_dna_img = dna_mip_base.copy()
        cv2.drawContours(contoured_dna_img, contours, -1, (0,255,0), 2, cv2.LINE_AA)
        #print(list(contours))

        index = 1
        isolated_count = 0
        cluster_count = 0
        #contoured_dna_img_1 = dna_mip_base.copy()
        #contoured_dna_img_2 = cv2.cvtColor(contoured_dna_img_1, cv2.COLOR_GRAY2BGR) * 255
        contoured_dna_img_1 = cv2.imread(dna_mip_input, cv2.IMREAD_GRAYSCALE)
        contoured_dna_img_2 = cv2.cvtColor(contoured_dna_img_1, cv2.COLOR_GRAY2BGR) *255
        #np -> [index, i , x, y]
        arr_nucleus_contours = np.empty([1,4])
        #print(arr_nucleus_contours)

        for cntr in contours:
            area = cv2.contourArea(cntr)
            convex_hull = cv2.convexHull(cntr)
            convex_hull_area = cv2.contourArea(convex_hull)
            if convex_hull_area > 0:
                ratio = area / convex_hull_area
            else: ratio = 0

            
            print(index, area, convex_hull_area, ratio)
            x,y,w,h = cv2.boundingRect(cntr)
            cv2.putText(contoured_dna_img_2, str(index), (x,y), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (0,0,255), 2)
            if ratio < 0.91:
                # cluster contours in red
                cv2.drawContours(contoured_dna_img_2, [cntr], 0, (0,0,255), 2)
                cluster_count = cluster_count + 1
            else:
                # isolated contours in green
                cv2.drawContours(contoured_dna_img_2, [cntr], 0, (0,0,255), 2)
                isolated_count = isolated_count + 1

            epsilon = 0.001 * cv2.arcLength(cntr, True)
            approx = cv2.approxPolyDP(cntr, epsilon, True)
            n = approx.ravel()  
            i = 0
            font = cv2.FONT_HERSHEY_COMPLEX 
            for j in n : 
                if(i % 2 == 0): 
                    x = n[i] 
                    y = n[i + 1] 
        
                    # String containing the co-ordinates. 
                    string = str(x) + " " + str(y)

                    #Generate contours cooridinates array to pass into tkinter visualization
                    arr_nucleus_contours = np.vstack((arr_nucleus_contours , [index,i,x,y]))
        
                    if(i == 0): 
                        # text on topmost co-ordinate. 
                        cv2.putText(contoured_dna_img_2, "Arrow tip", (x, y), 
                                        font, 0.5, (255, 0, 0))  
                    else: 
                        # text on remaining co-ordinates. 
                        cv2.putText(contoured_dna_img_2, string, (x, y),  
                                font, 0.5, (0, 255, 0))  
                    #print(f"index: {index} i: {i} x: {x} y: {y}")

                i = i + 1
            index = index + 1
        #print('number_clusters:',cluster_count)
        #print('number_isolated:',isolated_count)


        arr_nucleus_contours = np.delete(arr_nucleus_contours,(0),axis=0)

        #Pass to DynData Class
        self.arr_nucleus_contours = arr_nucleus_contours
        dyn_data.arr_nucleus_contours = self.arr_nucleus_contours   

        print(arr_nucleus_contours)
        print(len(arr_nucleus_contours))


        mainapp.toOverlay()
        self.master.destroy()

        #cell_pick_gui_main(arr_nucleus_contours)
        #---------------
        #cv2.imshow('Base Image', dapi_nucleus_base_img)
        #cv2.imshow('Gaussian Blur Image', gblur_img)
        #cv2.imshow('Threshold Nucleus Image', nucleus_threshold_img)
        #cv2.imshow('Masked DNA Image', masked_dna_img)
        #cv2.imshow('Contoured Image', contoured_dna_img)
        #cv2.imshow('Contoured2 Image',contoured_dna_img_2)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class OverlayGUI():
    def __init__(self, master):
        #GET XYZOFFSET
        self.registerXY = RegisterXY()
        #----
        self.f1ntag = dyn_data.f1nametag
        self.f2ntag = dyn_data.f2nametag

        #region Load Info from other classes to perform overlay
        self.inputs_directory_F1 = "FILE_1"
        self.inputs_directory_F2 = "FILE_2"
        self.inputs_directory_folder = "Inputs"
        self.processing_directory_folder = "Processing"
        self.p_MIP_directory = "MIP"
        self.p_MIP_raw_directory = "RAW_MIP"
        self.p_zslices_directory = "zslices"   

        self.F1_base_path = os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F1))[0]
        self.F2_base_path = os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F2))[0]

        self.f1filepath = os.path.join(self.inputs_directory_folder,self.inputs_directory_F1,self.F1_base_path)
        self.f2filepath = os.path.join(self.inputs_directory_folder,self.inputs_directory_F2,self.F2_base_path)


        if turbomode: self.metadata_processor()

        self.f1dapichan = None
        self.f2dapichan = None
        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                self.f1dapichan = index
                self.F1_DAPI_dir = f"C{self.f1dapichan}_DAPI"
                
        
        for index, chan in enumerate(dyn_data.f2channels):
            if chan == "DAPI":
                self.f2dapichan = index
                self.F2_DAPI_dir = f"C{self.f2dapichan}_DAPI"

       # print(f"dapi1 = {self.f1dapichan}")
       # print(f"dapi2 = {self.f2dapichan}")

        self.F1_C0B_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_raw_directory,self.F1_DAPI_dir)
        self.F2_C0B_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_raw_directory,self.F2_DAPI_dir)

        print(f"dapi1 = {self.F1_C0B_MIP_path}")
        print(f"dapi2 = {self.F2_C0B_MIP_path}")

        #self.F1_C0B_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_directory,self.p_C0B_directory)
        #self.F1_C1G_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_directory,self.p_C1G_directory)
        #self.F1_C2R_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_directory,self.p_C2R_directory)
        #self.F1_C3M_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_directory,self.p_C3M_directory)
        #self.F2_C0B_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_directory,self.p_C0B_directory)
        #self.F2_C1G_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_directory,self.p_C1G_directory)
        #self.F2_C2R_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_directory,self.p_C2R_directory)
        #self.F2_C3M_MIP_path = os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_directory,self.p_C3M_directory)

        #IMAGES h=2044 w= 2048
        self.F1_C0B_MIP = self.get_MIP_file(self.F1_C0B_MIP_path)
        #self.F1_C1G_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_directory,self.p_C1G_directory))
        #self.F1_C2R_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_directory,self.p_C2R_directory))
        #self.F1_C3M_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_directory,self.p_C3M_directory))
        self.F2_C0B_MIP = self.get_MIP_file(self.F2_C0B_MIP_path)
        #self.F2_C1G_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_directory,self.p_C1G_directory))
        #self.F2_C2R_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_directory,self.p_C2R_directory))
        #self.F2_C3M_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_directory,self.p_C3M_directory))
        #endregion

        #----------------------------------------------
       
        #-----------------------------------------------------------------------#
        #Offset on drawn canvas

        print(f"translation offset: {dyn_data.f_offset}")

        
        self.f2_offset_x = 0
        self.f2_offset_y = 0

        #Offset on base image --> passed to other classes
        self.f2_offset_x_scaled = 0
        self.f2_offset_y_scaled = 0
        #-----------------------------------------------------------------------#
        #Initial Opacity
        self.opac = 0.5
        self.myscalex = 1
        self.myscaley = 1

        #Define Master Window (Overlay)
        self.master = master
        self.master.title("ZFISHER --- Registration")


        self.hwmultiplier = self.F1_C0B_MIP.height
        self.screen_width = master.winfo_screenwidth() * 0.8
        self.screen_height = master.winfo_screenheight() * 0.8
        self.hwmult = self.F1_C0B_MIP.width/self.F1_C0B_MIP.height



        #Create canvas
        self.canvas = tk.Canvas(master, width=self.screen_height*self.hwmult, height=self.screen_height)
        self.canvas.pack(side=tk.BOTTOM)
        self.master.update_idletasks()



    
        
        #region ---CREATE TOGGLE WINDOW---
        self.control_window = tk.Toplevel(self.master)
        self.control_window.title("Control Panel - Registration")

        #FILENAME FRAME
        self.fnameframe = tk.Frame(self.control_window)
        self.fnameframe.grid(row=1, column=1, columnspan=4)

        f1 = os.path.join(self.inputs_directory_folder,self.inputs_directory_F1,os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F1))[0])
        f2 = os.path.join(self.inputs_directory_folder,self.inputs_directory_F2,os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F2))[0])
        with ND2Reader(f1) as nd2_file:
            self.f1width = nd2_file.metadata['width']
            self.f1height = nd2_file.metadata['height']
        with ND2Reader(f2) as nd2_file:
            self.f2width = nd2_file.metadata['width']
            self.f2height = nd2_file.metadata['height']
        self.f1_filename_lab = tk.Label(self.fnameframe, text=f"{self.f1ntag}(F1): {self.F1_base_path}")
        self.f1_filename_lab.grid(row=1, column=0, columnspan=2)
        self.f1_fdim_lab = tk.Label(self.fnameframe, text =f"{self.f1width}x{self.f1height}")
        self.f1_fdim_lab.grid(row=1, column=2, columnspan=2)

        self.f2_filename_lab = tk.Label(self.fnameframe, text=f"{self.f2ntag}(F2): {self.F2_base_path}")
        self.f2_filename_lab.grid(row=2, column=0, columnspan=2)
        self.f2_fdim_lab = tk.Label(self.fnameframe, text =f"{self.f2width}x{self.f2height}")
        self.f2_fdim_lab.grid(row=2, column=2, columnspan=2)

        self.canvas_size_lab = tk.Label(self.fnameframe, text =f"Canvas Size: {self.canvas.winfo_width()} x {self.canvas.winfo_height()}")
        self.canvas_size_lab.grid(row=3, column=0, columnspan=4)
       #TOGGLES
        self.togglesframe = tk.Frame(self.control_window)
        self.togglesframe.grid(row=2, column=1, columnspan=4)
        self.list_label = tk.Label(self.togglesframe, text="-----FILE_CHANNEL_TOGGLE-----")
        self.list_label.grid(row=1, column=1, columnspan=4)
        self.F1_C0B_toggle = tk.BooleanVar(value=True)

        self.F2_C0B_toggle = tk.BooleanVar(value=True)

        #self.F1_C0B_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F1_C0B", variable=self.F1_C0B_toggle, command=self.toggle_channels).grid(row=2, column=1, columnspan=2) if self.F1_C0B_MIP is not None else None
        #self.F1_C1G_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F1_C1G", variable=self.F1_C1G_toggle, command=self.toggle_channels).grid(row=3, column=1, columnspan=2) if self.F1_C1G_MIP is not None else None
        #self.F1_C2R_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F1_C2R", variable=self.F1_C2R_toggle, command=self.toggle_channels).grid(row=4, column=1, columnspan=2) if self.F1_C2R_MIP is not None else None
        #self.F1_C3M_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F1_C3M", variable=self.F1_C3M_toggle, command=self.toggle_channels).grid(row=5, column=1, columnspan=2) if self.F1_C3M_MIP is not None else None
        #self.F2_C0B_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F2_C0B", variable=self.F2_C0B_toggle, command=self.toggle_channels).grid(row=2, column=3, columnspan=2) if self.F2_C0B_MIP is not None else None
        #self.F2_C1G_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F2_C1G", variable=self.F2_C1G_toggle, command=self.toggle_channels).grid(row=3, column=3, columnspan=2) if self.F2_C1G_MIP is not None else None
       # self.F2_C2R_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F2_C2R", variable=self.F2_C2R_toggle, command=self.toggle_channels).grid(row=4, column=3, columnspan=2) if self.F2_C2R_MIP is not None else None
       # self.F2_C3M_toggle_checkbox = tk.Checkbutton(self.togglesframe, text="F2_C3M", variable=self.F2_C3M_toggle, command=self.toggle_channels).grid(row=5, column=3, columnspan=2) if self.F2_C3M_MIP is not None else None
        #-spacer1
        self.spacer_label = tk.Label(self.control_window, text="------------------------------------")
        self.spacer_label.grid(row=4, column=1, columnspan=4)
        #F2 OPAC ENTRY
        self.opacframe = tk.Frame(self.control_window)
        self.opacframe.grid(row=6, column=1, columnspan=4)
        self.f2_opac_lab = tk.Label(self.opacframe, text=f"{self.f2ntag} (F2)  Transparency: " )
        self.f2_opac_lab.grid(row=1,column=1)
        self.opac_e_text = tk.StringVar(value= self.opac)
        self.f2_opac_entry = tk.Entry(self.opacframe, width=5, textvariable=self.opac_e_text)
        self.f2_opac_entry.grid(row=1,column=2)
        self.f2_opac_e_butt = tk.Button(self.opacframe, text=f"Set {self.f2ntag} (F2) % ", pady = 2, command=self.setf2opac)
        self.f2_opac_e_butt.grid(row=1,column=3)
        #SCALEBAR TOGGLE
        self.scalebar_toggle_var = tk.BooleanVar(value=True)
        self.scalebar_toggle_checkbox = tk.Checkbutton(self.control_window,text="Show Scalebar", variable=self.scalebar_toggle_var, command=self.toggle_scalebar)
        self.scalebar_toggle_checkbox.grid(row=7, column=1, columnspan=4)
        #-spacer2
        self.spacer2_label = tk.Label(self.control_window, text="------------------------------------")
        self.spacer2_label.grid(row=8, column=1, columnspan=4)
        #OFFSET ENTRY
        self.offentframe = tk.Frame(self.control_window)
        self.offentframe.grid(row=9, column=1,columnspan=4)
        self.c_os_x_l = tk.Label(self.offentframe, text="X:", fg="white")
        self.c_os_x_l.grid(row=9, column=1, sticky=tk.E)
        self.c_os_x_e_text = tk.StringVar(value= self.f2_offset_x)
        self.c_os_x_e = tk.Entry(self.offentframe, width=5, textvariable=self.c_os_x_e_text)
        
        self.c_os_x_e.grid(row=9, column=2, sticky=tk.W)
        self.c_os_x_e.insert(0,self.f2_offset_x)
        
        self.c_os_y_l = tk.Label(self.offentframe, text="Y:", fg="white")
        self.c_os_y_l.grid(row=9, column=3, sticky=tk.E)
        self.c_os_y_e_text = tk.StringVar(value= self.f2_offset_y)
        self.c_os_y_e = tk.Entry(self.offentframe, width=5, textvariable=self.c_os_y_e_text)
        
        self.c_os_y_e.grid(row=9, column=4, sticky=tk.W)
        self.c_os_y_e.insert(0,self.f2_offset_y)

        self.c_os_x_e_text.trace_add("write", self.on_offset_entry)
        self.c_os_y_e_text.trace_add("write", self.on_offset_entry)

        self.os_ent_but = tk.Button(self.offentframe, text="Set Offset", pady = 2, command=self.manualoffset)
        self.os_ent_but.grid(row=10, column=1, columnspan=5)
        #OFFSET LABEL
        self.offsetframe = tk.Frame(self.control_window)
        self.offsetframe.grid(row=10, column=1, columnspan=4)
        self.offset_label = tk.Label(self.offsetframe, text="Canvas Offset: (0,0)", fg="white")
        self.offset_label.grid(row=1, column=1, columnspan=4)
        self.f2offset_label = tk.Label(self.offsetframe, text=f"{self.f2ntag} (F2) Offset: (0,0)", fg="white")
        self.f2offset_label.grid(row=2, column=0, columnspan=4)
        #-spacer3
        self.spacer2_label = tk.Label(self.control_window, text="------------------------------------")
        self.spacer2_label.grid(row=11, column=1, columnspan=4)
        #RESET ALL
        self.reset_button = tk.Button(self.control_window, text="RESET ALL", pady=4, command=self.reset_all)
        self.reset_button.grid(row=12, column=1, columnspan=4,pady=5)
        #-spacer4
        self.spacer2_label = tk.Label(self.control_window, text="------------------------------------")
        self.spacer2_label.grid(row=13, column=1, columnspan=4)

        #Finish Button
        self.finish_button = tk.Button(self.control_window, text=f"Finalize {self.f2ntag} (F2) Offset", pady=12, command=self.finalize_offset)
        self.finish_button.grid(row=14, column=0, columnspan=5, pady=5)
        #endregion


        self.y_scaler = self.f1height/ self.canvas.winfo_height()
        self.x_scaler = self.f1width/ self.canvas.winfo_width()



        self.toggle_channels()

        self.draw_scalebar()
        self.scalebar_visible = True

        self.c_os_x_e_text.set(dyn_data.f_offset[0])
        self.c_os_y_e_text.set(dyn_data.f_offset[1])
        self.manualoffset()

        # Bind mouse events

    def setf2opac(self):
        newopac = float(self.opac_e_text.get())
        print(newopac)
        self.opac = newopac
        self.toggle_channels()

    def manualoffset(self):
        x = float(self.c_os_x_e_text.get())
        y = float(self.c_os_y_e_text.get())
        print(f"{x},{y}")

        #The input x,y is in native image dimensions
        self.f2_offset_x = x / self.x_scaler
        self.f2_offset_y = y / self.y_scaler
        self.f2_offset_x_scaled = x 
        self.f2_offset_y_scaled = y      

        decimal_places = 2
        self.c_os_x_e_text.set(f"{self.f2_offset_x_scaled:.{decimal_places}f}")
        self.c_os_y_e_text.set(f"{self.f2_offset_y_scaled:.{decimal_places}f}")

        self.offset_label.config(text=f"Canvas Offset: ({round(self.f2_offset_x,2)},{round(self.f2_offset_y,2)})")
        self.f2offset_label.config(text=f"{self.f2ntag} (F) Offset: ({round(self.f2_offset_x_scaled,2)},{round(self.f2_offset_y_scaled,2)})")

        self.canvas.coords(self.f2comp,self.f2_offset_x,self.f2_offset_y)

    def reset_all(self):
        self.f2_offset_x = 0
        self.f2_offset_y = 0
        self.f2_offset_x_scaled = 0
        self.f2_offset_y_scaled = 0
        self.c_os_x_e_text.set(self.f2_offset_x)
        self.c_os_y_e_text.set(self.f2_offset_y)
        self.offset_label.config(text=f"Offset: ({round(self.f2_offset_x,2)},{round(self.f2_offset_y,2)})")
        self.f2offset_label.config(text=f"{self.f2ntag} (F2) Offset: ({round(self.f2_offset_x_scaled,2)},{round(self.f2_offset_y_scaled,2)})")
        self.opac = 0.5
        self.opac_e_text.set(self.opac)
        self.myscalex = 1
        self.myscaley = 1
        self.F1_C0B_toggle.set(True)
        self.F2_C0B_toggle.set(True)
        self.scalebar_toggle_var.set(True)
        self.scalebar_visible = True


        self.toggle_channels()
        self.draw_scalebar()



    def on_offset_entry(self, *args):
        pass
       #x_text = self.c_os_x_e_text.get()
       #y_text = self.c_os_y_e_text.get()
       #if x_text and y_text:
        #    self.f2_offset_x = float(self.c_os_x_e_text.get())
         #   self.f2_offset_y = float(self.c_os_y_e_text.get())


    def f1_toggle_imgprocessor(self):
        self.added_image = None
        #F1C0B
        if self.F1_C0B_MIP is not None and self.F1_C0B_toggle.get():
                    self.F1C0B_p = self.F1_C0B_MIP.copy()
                    self.F1C0B_p = np.array(self.F1C0B_p)
                    self.pil_image = Image.fromarray(self.F1C0B_p)
                    self.added_image = self.pil_image.copy()
        #F1C1G
        #if self.F1_C1G_MIP is not None and self.F1_C1G_toggle.get():  
        #    self.F1C1G_p = self.F1_C1G_MIP.copy()  
        #    if self.added_image is None:
        #        self.F1C1G_p = np.array(self.F1C1G_p)
        #        self.pil_image = Image.fromarray(self.F1C1G_p)
        #       self.added_image = self.pil_image.copy()
         #   else:
         #       self.added_image = np.array(self.added_image)
         #       self.F1C1G_p = np.array(self.F1C1G_p)
         #       self.sum_image = cv2.add(self.added_image,self.F1C1G_p)
         #       self.pil_image = Image.fromarray(self.sum_image)
        #        self.added_image = self.pil_image.copy()
        #F1C2R
       # if self.F1_C2R_MIP is not None and self.F1_C2R_toggle.get():  
       #     self.F1C2R_p = self.F1_C2R_MIP.copy()  
       #     if self.added_image is None:
       #         self.F1C2R_p = np.array(self.F1C2R_p)
       #         self.pil_image = Image.fromarray(self.F1C2R_p)
       #         self.added_image = self.pil_image.copy()
       #     else:
        #        self.added_image = np.array(self.added_image)
        #        self.F1C2R_p = np.array(self.F1C2R_p)
        #        self.sum_image = cv2.add(self.added_image,self.F1C2R_p)
        #        self.pil_image = Image.fromarray(self.sum_image)
        #        self.added_image = self.pil_image.copy()
        #F1C3M
       # if self.F1_C3M_MIP is not None and self.F1_C3M_toggle.get():  
       #     self.F1C3M_p = self.F1_C3M_MIP.copy()  
       #     if self.added_image is None:
        #        self.F1C3M_p = np.array(self.F1C3M_p)
        #        self.pil_image = Image.fromarray(self.F1C3M_p)
        #        self.added_image = self.pil_image.copy()
        #    else:
         #       self.added_image = np.array(self.added_image)
         ##       self.F1C3M_p = np.array(self.F1C3M_p)
        #        self.sum_image = cv2.add(self.added_image,self.F1C3M_p)
         #       self.pil_image = Image.fromarray(self.sum_image)
         #       self.added_image = self.pil_image.copy()

        return self.added_image
    def f2_toggle_imgprocessor(self):
        self.added_image = None
        #F2C0B
        if self.F2_C0B_MIP is not None and self.F2_C0B_toggle.get():
                    self.F2C0B_p = self.F2_C0B_MIP.copy()
                    self.F2C0B_p = np.array(self.F2C0B_p)
                    self.pil_image = Image.fromarray(self.F2C0B_p)
                    self.added_image = self.pil_image.copy()
        #F2C1G
      #  if self.F2_C1G_MIP is not None and self.F2_C1G_toggle.get():  
      #      self.F2C1G_p = self.F2_C1G_MIP.copy()  
      #      if self.added_image is None:
       #         self.F2C1G_p = np.array(self.F2C1G_p)
       #         self.pil_image = Image.fromarray(self.F2C1G_p)
        #        self.added_image = self.pil_image.copy()
        #    else:
        #        self.added_image = np.array(self.added_image)
       #         self.F2C1G_p = np.array(self.F2C1G_p)
        #        self.sum_image = cv2.add(self.added_image,self.F2C1G_p)
       #         self.pil_image = Image.fromarray(self.sum_image)
       #         self.added_image = self.pil_image.copy()
        #F2C2R
      #  if self.F2_C2R_MIP is not None and self.F2_C2R_toggle.get():  
      #      self.F2C2R_p = self.F2_C2R_MIP.copy()  
      #      if self.added_image is None:
       #         self.F2C2R_p = np.array(self.F2C2R_p)
       #         self.pil_image = Image.fromarray(self.F2C2R_p)
       #         self.added_image = self.pil_image.copy()
       #     else:
       #         self.added_image = np.array(self.added_image)
       #         self.F2C2R_p = np.array(self.F2C2R_p)
         #       self.sum_image = cv2.add(self.added_image,self.F2C2R_p)
         #       self.pil_image = Image.fromarray(self.sum_image)
       #         self.added_image = self.pil_image.copy()
        #F2C3M
      #  if self.F2_C3M_MIP is not None and self.F2_C3M_toggle.get():  
       #     self.F2C3M_p = self.F2_C3M_MIP.copy()  
       #     if self.added_image is None:
       #         self.F2C3M_p = np.array(self.F2C3M_p)
       #         self.pil_image = Image.fromarray(self.F2C3M_p)
       #         self.added_image = self.pil_image.copy()
       #     else:
       #         self.added_image = np.array(self.added_image)
       #         self.F2C3M_p = np.array(self.F2C3M_p)
       #         self.sum_image = cv2.add(self.added_image,self.F2C3M_p)
       #         self.pil_image = Image.fromarray(self.sum_image)
      #          self.added_image = self.pil_image.copy()
                
        return self.added_image

    def toggle_channels(self):
        # Generate F1 and F2 composites based on toggles
        self.f1img = self.f1_toggle_imgprocessor()
        self.f2img = self.f2_toggle_imgprocessor()

        # Ensure the image is in the correct mode
        #if self.f1img.mode not in ("RGB", "RGBA"):
        print("Converted F1 mode:", self.f1img.mode)
        self.f1img = self.f1img.convert("L")
        print("Converted F1 mode:", self.f1img.mode)
        self.f1img = self.f1img.convert("RGBA")
        print("Converted F1 mode:", self.f1img.mode)

        # Put F1 on canvas
        self.f1img.thumbnail((self.canvas.winfo_width(), self.canvas.winfo_height()))
        self.f1photo = ImageTk.PhotoImage(self.f1img)
        self.f1comp = self.canvas.create_image(0, 0, anchor=tk.NW, image=self.f1photo)

        # Ensure f2img is in a compatible mode before applying alpha
        #if self.f2img.mode not in ("RGBA", "RGB"):
        self.f2img = self.f2img.convert("RGBA")

        # Generate F2 Opacity 
        # Create a mask with the desired opacity
        self.f2mask = Image.new("L", self.f2img.size, int(255 * self.opac))
        self.f2imgwopac = self.f2img.copy()
        self.f2imgwopac.putalpha(self.f2mask)
        
        self.f2imgwopac.thumbnail((self.canvas.winfo_width(),self.canvas.winfo_height()))
        self.f2photo = ImageTk.PhotoImage(self.f2imgwopac)
        
        self.f2comp = self.canvas.create_image(self.f2_offset_x, self.f2_offset_y, anchor=tk.NW, image=self.f2photo)

        # Bind mouse events
        self.canvas.tag_bind(self.f2comp, '<Button-1>', self.on_drag_start)
        self.canvas.tag_bind(self.f2comp, '<B1-Motion>', self.on_drag_motion)
        self.canvas.tag_bind(self.f2comp, '<ButtonRelease-1>', self.on_drag_release)

        self.draw_scalebar()

    #region ---DRAGGING---
    def on_drag_start(self, event):
        self.dragging = True
        self.prev_x = event.x
        self.prev_y = event.y
        self.initial_x = event.x
        self.initial_y = event.y

    def on_drag_motion(self, event):
        if self.dragging:
            dx = event.x - self.prev_x
            dy = event.y - self.prev_y
            self.canvas.move(self.f2comp, dx, dy)
            self.prev_x = event.x
            self.prev_y = event.y

            self.update_offset_label(event.x, event.y)

    def on_drag_release(self, event):
        self.dragging = False
        #self.update_offset_label(event.x, event.y)

    def update_offset_label(self, x, y):
        offset_x = x - self.initial_x
        offset_y = y - self.initial_y

        self.f2_offset_x += offset_x 
        self.f2_offset_y += offset_y  

        decimal_places = 2

        self.f1_coords = self.canvas.coords(self.f1comp)
        self.f2_coords = self.canvas.coords(self.f2comp)
        self.relative_x = self.f2_coords[0] - self.f1_coords[0]
        self.relative_y = self.f2_coords[1] - self.f1_coords[1]

        self.f2_offset_x = self.relative_x
        self.f2_offset_y = self.relative_y

        self.f2_offset_x_scaled = self.relative_x * self.x_scaler
        self.f2_offset_y_scaled = self.relative_y * self.y_scaler       

       # print (f"f1 - {f1_coords} ::: f2 - {f2_coords} ::: relative - {relative_x,relative_y}")

        self.c_os_x_e_text.set(f"{self.f2_offset_x_scaled:.{decimal_places}f}")
        self.c_os_y_e_text.set(f"{self.f2_offset_y_scaled:.{decimal_places}f}")

        self.offset_label.config(text=f"Canvas Offset: ({round(self.relative_x,2)},{round(self.relative_y,2)})")
        self.f2offset_label.config(text=f"F2 Offset: ({round(self.f2_offset_x_scaled,2)},{round(self.f2_offset_y_scaled,2)})")
    #endregion
    #region ---Scalebar---
    def toggle_scalebar(self):
        if self.scalebar_visible:
            # Hide the scale bar
            self.canvas.delete("scalebar")
        else:
            # Show the scale bar
            self.draw_scalebar()  # Adjust the length of the scale bar as needed
        self.scalebar_visible = not self.scalebar_visible
    def draw_scalebar(self):
        #1px =  0.108333333333333 microns
        # Determine the position of the scale bar relative to the image
        # For example, you may want to place it at the bottom-right corner
        if self.scalebar_toggle_var.get():
            px_per_micron = 0.108333333333333
            scalebar_microns = 10
            scale_bar_width = scalebar_microns / px_per_micron
            scale_bar_height = 5  # Height of the scale bar rectangle

            # Calculate the coordinates for the scale bar
            image_width = self.canvas.winfo_width()
            image_height = self.canvas.winfo_height()
            scale_bar_x0 = image_width - scale_bar_width - 10  # Adjust as needed
            scale_bar_y0 = image_height - scale_bar_height - 10  # Adjust as needed
            scale_bar_x1 = image_width - 10  # Adjust as needed
            scale_bar_y1 = image_height - 10  # Adjust as needed

            # Draw the scale bar
            self.canvas.create_rectangle(scale_bar_x0, scale_bar_y0, scale_bar_x1, scale_bar_y1, fill="white", tags="scalebar")

            # Optionally, add text to indicate the length of the scale bar
            scale_text_x = scale_bar_x0 + (scale_bar_width / 2)
            scale_text_y = scale_bar_y0 - 7  # Adjust as needed
            self.canvas.create_text(scale_text_x, scale_text_y, text=f"{scalebar_microns} um ({round(scale_bar_width,1)} px)", fill="white", tags="scalebar")
    #endregion
    #------FINALIZE------------#
    def finalize_offset(self):
        dyn_data.f_offset = [self.f2_offset_x_scaled,self.f2_offset_y_scaled]
        print(f"Offset of F2 to F1: {dyn_data.f_offset}")

        self.cropZstackXY = CropZSTACK_XY()
        self.registerZ = RegisterZ()

        mainapp.toNucpick()
        #self.master.destroy()
    #------INITIALIZATION------#
    def get_MIP_file(self,target_dir):
        files = os.listdir(target_dir)
        if len(files) == 0: return None
        else:
            image_name= os.listdir(target_dir)[0]
            image_path = os.path.join(target_dir,image_name)
            image_in = Image.open(image_path)
            normalized_img = self.normalize_image(image_in)
            return normalized_img

    def normalize_image(self,image):
        # Convert the image to a numpy array
        image_array = np.array(image).astype(np.float32)
        
        # Find the minimum and maximum pixel values
        min_val = image_array.min()
        max_val = image_array.max()
        
        print(f"MIN VAL {min_val}")
        print(f"MAX VAL {max_val}")
        # Normalize the image to the range [0, 1]
        normalized_array = (image_array - min_val) / (max_val - min_val)
        
        # Optionally, scale to [0, 255] for 8-bit representation
        normalized_array = (normalized_array * 255).astype(np.uint8)
        
        # Convert back to PIL Image
        normalized_image = Image.fromarray(normalized_array)
        return normalized_image
    
    def metadata_processor(self):
        print("METADATA PROCESSOR")
        #FILE1
        with ND2Reader(self.f1filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f1numchan = len(nd2_file.metadata['channels'])
            self.f1numslices = len(nd2_file)
            self.f1channels = nd2_file.metadata['channels']

            self.f1pmax = float(self.f1numchan*self.f1numslices)

        #FILE2
        with ND2Reader(self.f2filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f2numchan = len(nd2_file.metadata['channels'])
            self.f2numslices = len(nd2_file)
            self.f2channels = nd2_file.metadata['channels']

            self.f2pmax = float(self.f2numchan*self.f2numslices)

        dyn_data.f1channels = self.f1channels
        dyn_data.f2channels = self.f2channels
        dyn_data.f1numslices = self.f1numslices
        dyn_data.f2numslices = self.f2numslices
        self.f1ntag = dyn_data.f1nametag
        self.f2ntag = dyn_data.f2nametag
class RegisterXY():
    def __init__(self):
        #PATHS
        self.f1MIP_path = 'Processing/FILE_1/RAW_MIP/C0_DAPI/F1_C0_DAPI_MIP_.tif'
        self.f2MIP_path = 'Processing/FILE_2/RAW_MIP/C0_DAPI/F2_C0_DAPI_MIP_.tif'
        self.f1_z_DAPI_path = 'Processing/FILE_1/zslices/C0_DAPI'
        self.f2_z_DAPI_path = 'Processing/FILE_2/zslices/C0_DAPI'

        #region Load Info from other classes to perform overlay
        self.inputs_directory_F1 = "FILE_1"
        self.inputs_directory_F2 = "FILE_2"
        self.inputs_directory_folder = "Inputs"
        self.processing_directory_folder = "Processing"
        self.p_MIP_directory = "MIP"
        self.p_MIP_raw_directory = "RAW_MIP"
        self.p_zslices_directory = "zslices"   

        self.F1_base_path = os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F1))[0]
        self.F2_base_path = os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F2))[0]

        self.f1filepath = os.path.join(self.inputs_directory_folder,self.inputs_directory_F1,self.F1_base_path)
        self.f2filepath = os.path.join(self.inputs_directory_folder,self.inputs_directory_F2,self.F2_base_path)

        #VARIABLES
        self.f1MIP = None
        self.f2MIP = None
        self.f1MIP_n = None
        self.f2MIP_n = None
        self.f1sortedslices = []
        self.f2sortedslices = []
        self.f1slicenum = 0
        self.f2slicenum = 0

        self.f2_offsetXY = []

        self.f1_stack_arr_cropped = []
        self.f2_stack_arr_cropped = []

        #Perform Functions

        if turbomode: self.metadata_processor()

        self.load_MIP()
        self.normalizef1f2MIP()
        self.registerXY()

        #self.load_f1f2slices()


    #______FUNCTIONS___________________________
    #LOAD IMAGES---------------------------------------
    def load_MIP(self):     #get file 1 MIP
        f1dapichan = None
        f2dapichan = None

        print(dyn_data.f1channels)
        print(dyn_data.f2channels)
        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                f1dapichan = index
        
        for index, chan in enumerate(dyn_data.f2channels):
            if chan == "DAPI":
                f2dapichan = index


        f1MIPpath = f"Processing/FILE_1/RAW_MIP/C{f1dapichan}_DAPI/F1_C{f1dapichan}_DAPI_MIP_.tif"
        f2MIPpath = f"Processing/FILE_2/RAW_MIP/C{f2dapichan}_DAPI/F2_C{f2dapichan}_DAPI_MIP_.tif"
        f1MIP = Image.open(f1MIPpath)
        f2MIP = Image.open(f2MIPpath)
        self.f1MIP = f1MIP
        self.f2MIP = f2MIP

        print(f"BASE F1 MIP: {f1MIPpath} ")  
        print(f"BASE F2 MIP: {f2MIPpath} ")  
        print(f"F1 mode: {f1MIP.mode}")
        print(f"F2 mode: {f2MIP.mode}")

        image_array_f1 = np.array(f1MIP).astype(np.float16)
        image_array_f2 = np.array(f2MIP).astype(np.float16)
        
        # Find the minimum and maximum pixel values
        f1mipmin_val = image_array_f1.min()
        f1mipmax_val = image_array_f1.max()
        f2mipmin_val = image_array_f2.min()
        f2mipmax_val = image_array_f2.max()
        
        print(f"f1mip MIN VAL {f1mipmin_val}")
        print(f"f1mip MAX VAL {f1mipmax_val}")
        print(f"f2mip MIN VAL {f2mipmin_val}")
        print(f"f2mip MAX VAL {f2mipmax_val}")
    def slice_sort_key(self, filename):
        # Extract the number between "z" and "_.tif" in the filename
        start_index = filename.find('z') + 1
        end_index = filename.find('_.tif')
        number_str = filename[start_index:end_index]
        # Convert the extracted number string to an integer
        return int(number_str)
    def load_f1f2slices(self):
        f1dapichan = None
        f2dapichan = None
        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                f1dapichan = index
        
        for index, chan in enumerate(dyn_data.f2channels):
            if chan == "DAPI":
                f2dapichan = index       
        
        f1slicesdir = f'Processing/FILE_1/zslices/C{f1dapichan}_DAPI'
        f2slicesdir = f'Processing/FILE_1/zslices/C{f2dapichan}_DAPI'

        slices = os.listdir(f1slicesdir)
        sorted_slices = sorted(slices, key=self.slice_sort_key)
        print(sorted_slices)
        for s in sorted_slices:
            path = os.path.join(f1slicesdir,s)
            sliceimage = Image.open(path)
            f1sortedslices.append(sliceimage)

        
        slices = os.listdir(f2slicesdir)
        sorted_slices = sorted(slices, key=self.slice_sort_key)
        print(sorted_slices)
        for s in sorted_slices:
            path = os.path.join(f2slicesdir,s)
            sliceimage = Image.open(path)
            f2sortedslices.append(sliceimage)

    #NORMALIZATION-------------------------------------
    def normalize_image(self,image):
        # Convert the image to a numpy array
        image_array = np.array(image).astype(np.float16)
        
        # Find the minimum and maximum pixel values
        min_val = image_array.min()
        max_val = image_array.max()
        
        print(f"MIN VAL: {min_val}")
        print(f"MAX VAL: {max_val}")
        
        # Normalize the image to the range [0, 1]
        normalized_array = (image_array - min_val) / (max_val - min_val)
        
        # Scale to [0, 65535] for 16-bit representation
        normalized_array = (normalized_array * 65535).astype(np.uint16)
        
        # Convert back to PIL Image
        normalized_image = Image.fromarray(normalized_array)



        return normalized_image
    def normalizef1f2MIP(self):
        self.f1MIP_n = self.normalize_image(self.f1MIP)
        self.f2MIP_n = self.normalize_image(self.f2MIP)


        f1image_array = np.array(self.f1MIP_n).astype(np.float32)
        f2image_array = np.array(self.f2MIP_n).astype(np.float32)

        f1mipmin_val = f1image_array.min()
        f1mipmax_val = f1image_array.max()
        f2mipmin_val = f2image_array.min()
        f2mipmax_val = f2image_array.max()

        print(f"f1mip_n MIN VAL {f1mipmin_val}")
        print(f"f1mip_n MAX VAL {f1mipmax_val}")
        print(f"f2mip_n MIN VAL {f2mipmin_val}")
        print(f"f2mip_n MAX VAL {f2mipmax_val}")   

    #-------------------------------------------------- 
    #XYREGISTRATION------------------------------------
    def ensure_grayscale(self,image):
    # Convert PIL Image to numpy array
        image_array = np.array(image)
        
        # Check if the image has 3 dimensions (indicating an RGB image)
        if image_array.ndim == 3:
            # Convert RGB to grayscale by averaging the channels
            image_array = np.mean(image_array, axis=2)
        
        return image_array
    def registerXY(self):
        #f, ax = plt.subplots(2, 2, figsize=(16, 18))

        transformations = {
            'TRANSLATION': StackReg.TRANSLATION
        # 'RIGID_BODY': StackReg.RIGID_BODY,
        # 'SCALED_ROTATION': StackReg.SCALED_ROTATION,
        # 'AFFINE': StackReg.AFFINE,
        # 'BILINEAR': StackReg.BILINEAR
        }

        ref = self.f1MIP_n
        mov = self.f2MIP_n 
        ref = self.ensure_grayscale(ref)
        mov = self.ensure_grayscale(mov)

        for i, (name, tf) in enumerate(transformations.items()):
            sr = StackReg(tf)
            reg = sr.register_transform(ref, mov)
            print(reg)
            reg = reg.clip(min=0)


            if name != 'BILINEAR': 
                # Extracting and displaying the offset
                offset = sr.get_matrix()[:2, 2]
                print(f"Offset for {name}: {offset}")
                print(sr.get_matrix())

                # Adding text annotation to display the offset
                #ax[i][1].text(0.5, 0.5, f"Offset: {offset}", fontsize=12, color='red',
                #            ha='center', va='center', transform=ax[i][1].transAxes)
            else:
                pass
                #ax[i][1].axis('off')

        offset[0] = (-offset[0])
        offset[1] = (-offset[1])
        self.f2_offsetXY = offset
        print (f"OFFSET: {self.f2_offsetXY}")
        dyn_data.f_offset = self.f2_offsetXY

    def metadata_processor(self):
        print("METADATA PROCESSOR")
        #FILE1
        with ND2Reader(self.f1filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f1numchan = len(nd2_file.metadata['channels'])
            self.f1numslices = len(nd2_file)
            self.f1channels = nd2_file.metadata['channels']

            self.f1pmax = float(self.f1numchan*self.f1numslices)

        #FILE2
        with ND2Reader(self.f2filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f2numchan = len(nd2_file.metadata['channels'])
            self.f2numslices = len(nd2_file)
            self.f2channels = nd2_file.metadata['channels']

            self.f2pmax = float(self.f2numchan*self.f2numslices)

        dyn_data.f1channels = self.f1channels
        dyn_data.f2channels = self.f2channels
        dyn_data.f1numslices = self.f1numslices
        dyn_data.f2numslices = self.f2numslices
        self.f1ntag = dyn_data.f1nametag
        self.f2ntag = dyn_data.f2nametag
class CropZSTACK_XY():
    def __init__(self):
        #PATHS
        self.f1MIP_path = 'Processing/FILE_1/RAW_MIP/C0_DAPI/F1_C0_DAPI_MIP_.tif'
        self.f2MIP_path = 'Processing/FILE_2/RAW_MIP/C0_DAPI/F2_C0_DAPI_MIP_.tif'
        self.f1_z_DAPI_path = 'Processing/FILE_1/zslices/C0_DAPI'
        self.f2_z_DAPI_path = 'Processing/FILE_2/zslices/C0_DAPI'

        #region Load Info from other classes to perform overlay
        self.inputs_directory_F1 = "FILE_1"
        self.inputs_directory_F2 = "FILE_2"
        self.inputs_directory_folder = "Inputs"
        self.processing_directory_folder = "Processing"
        self.p_MIP_directory = "MIP"
        self.p_MIP_raw_directory = "RAW_MIP"
        self.p_zslices_directory = "zslices"   

        self.F1_base_path = os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F1))[0]
        self.F2_base_path = os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F2))[0]

        self.f1filepath = os.path.join(self.inputs_directory_folder,self.inputs_directory_F1,self.F1_base_path)
        self.f2filepath = os.path.join(self.inputs_directory_folder,self.inputs_directory_F2,self.F2_base_path)

        #VARIABLES
        self.f1MIP = None
        self.f2MIP = None
        self.f1MIP_n = None
        self.f2MIP_n = None
        self.f1sortedslices = []
        self.f2sortedslices = []
        self.f1slicenum = 0
        self.f2slicenum = 0

        self.f2_offsetXY = dyn_data.f_offset

        self.f1sortedslices = []
        self.f2sortedslices = []

        self.f1_stack_arr_cropped = []
        self.f2_stack_arr_cropped = []

        #Perform Functions

        if turbomode: self.metadata_processor()
        self.load_MIP()
        self.load_f1f2slices()
        self.f1_stack_arr_cropped,self.f2_stack_arr_cropped = self.crop_f1zstack_ref(self.f1sortedslices,self.f2sortedslices,self.f2_offsetXY)

        dyn_data.f1_Zstack_cropped = self.f1_stack_arr_cropped
        dyn_data.f2_Zstack_cropped = self.f2_stack_arr_cropped 
    def slice_sort_key(self, filename):
        # Extract the number between "z" and "_.tif" in the filename
        start_index = filename.find('z') + 1
        end_index = filename.find('_.tif')
        number_str = filename[start_index:end_index]
        # Convert the extracted number string to an integer
        return int(number_str)
    def load_f1f2slices(self):
        f1dapichan = None
        f2dapichan = None
        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                f1dapichan = index
        
        for index, chan in enumerate(dyn_data.f2channels):
            if chan == "DAPI":
                f2dapichan = index       
        
        f1slicesdir = f'Processing/FILE_1/zslices/C{f1dapichan}_DAPI'
        f2slicesdir = f'Processing/FILE_2/zslices/C{f2dapichan}_DAPI'

        slices = os.listdir(f1slicesdir)
        sorted_slices = sorted(slices, key=self.slice_sort_key)
        print(sorted_slices)
        for s in sorted_slices:
            path = os.path.join(f1slicesdir,s)
            sliceimage = Image.open(path)
            self.f1sortedslices.append(sliceimage)

        
        slices = os.listdir(f2slicesdir)
        sorted_slices = sorted(slices, key=self.slice_sort_key)
        print(sorted_slices)
        for s in sorted_slices:
            path = os.path.join(f2slicesdir,s)
            sliceimage = Image.open(path)
            self.f2sortedslices.append(sliceimage)

        print(f"load f1 f2 = {len(self.f1sortedslices)} --- {len(self.f2sortedslices)}")
    def metadata_processor(self):
        print("METADATA PROCESSOR")
        #FILE1
        with ND2Reader(self.f1filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f1numchan = len(nd2_file.metadata['channels'])
            self.f1numslices = len(nd2_file)
            self.f1channels = nd2_file.metadata['channels']

            self.f1pmax = float(self.f1numchan*self.f1numslices)

        #FILE2
        with ND2Reader(self.f2filepath) as nd2_file:
            print(nd2_file.metadata)
            self.f2numchan = len(nd2_file.metadata['channels'])
            self.f2numslices = len(nd2_file)
            self.f2channels = nd2_file.metadata['channels']

            self.f2pmax = float(self.f2numchan*self.f2numslices)

        dyn_data.f1channels = self.f1channels
        dyn_data.f2channels = self.f2channels
        dyn_data.f1numslices = self.f1numslices
        dyn_data.f2numslices = self.f2numslices
        self.f1ntag = dyn_data.f1nametag
        self.f2ntag = dyn_data.f2nametag
    def load_MIP(self):     #get file 1 MIP
        f1dapichan = None
        f2dapichan = None

        print(dyn_data.f1channels)
        print(dyn_data.f2channels)
        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                f1dapichan = index
        
        for index, chan in enumerate(dyn_data.f2channels):
            if chan == "DAPI":
                f2dapichan = index


        f1MIPpath = f"Processing/FILE_1/RAW_MIP/C{f1dapichan}_DAPI/F1_C{f1dapichan}_DAPI_MIP_.tif"
        f2MIPpath = f"Processing/FILE_2/RAW_MIP/C{f2dapichan}_DAPI/F2_C{f2dapichan}_DAPI_MIP_.tif"
        f1MIP = Image.open(f1MIPpath)
        f2MIP = Image.open(f2MIPpath)
        self.f1MIP = f1MIP
        self.f2MIP = f2MIP

        print(f"BASE F1 MIP: {f1MIPpath} ")  
        print(f"BASE F2 MIP: {f2MIPpath} ")  
        print(f"F1 mode: {f1MIP.mode}")
        print(f"F2 mode: {f2MIP.mode}")

        image_array_f1 = np.array(f1MIP).astype(np.float16)
        image_array_f2 = np.array(f2MIP).astype(np.float16)
        
        # Find the minimum and maximum pixel values
        f1mipmin_val = image_array_f1.min()
        f1mipmax_val = image_array_f1.max()
        f2mipmin_val = image_array_f2.min()
        f2mipmax_val = image_array_f2.max()
        
        print(f"f1mip MIN VAL {f1mipmin_val}")
        print(f"f1mip MAX VAL {f1mipmax_val}")
        print(f"f2mip MIN VAL {f2mipmin_val}")
        print(f"f2mip MAX VAL {f2mipmax_val}")
   
    def crop_f1zstack_ref(self,f1stack_in,f2stack_in,offset):
        tx = -offset[0]
        ty = -offset[1]

        ref = np.array(self.f1MIP)
        mov = np.array(self.f2MIP)
        
        # Get the dimensions of the images
        ref_height, ref_width = ref.shape[:2]
        mov_height, mov_width = mov.shape[:2]

        # Calculate the bounding box for the overlapping region
        start_x = max(0, int(tx))
        start_y = max(0, int(ty))
        end_x = min(ref_width, int(tx) + mov_width)
        end_y = min(ref_height, int(ty) + mov_height)

        # Ensure start and end coordinates are within bounds
        start_x_mov = max(0, -int(tx))
        start_y_mov = max(0, -int(ty))
        end_x_mov = end_x - int(tx)
        end_y_mov = end_y - int(ty)


        f1stack_out = []
        f2stack_out = []

        for slice in f1stack_in:
            slice = np.array(slice)
            cropped_slice = slice[start_y:end_y, start_x:end_x]
            f1stack_out.append(cropped_slice)

        for slice in f2stack_in:
            slice = np.array(slice)
            cropped_slice = slice[start_y_mov:end_y_mov, start_x_mov:end_x_mov]
            f2stack_out.append(cropped_slice) 

        print(f"OFFSET: {self.f2_offsetXY}")
        print(f"INPUT f1stack length {len(f1stack_in)} OUTPUT: {len(f1stack_out)}")
        print(f"INPUT f2stack length {len(f2stack_in)} OUTPUT: {len(f2stack_out)}")

        print(f"IN SIZE f1 {f1stack_in[0].size} (width,height) OUT {f1stack_out[0].shape} (height,width)")
        print(f"IN SIZE f2 {f2stack_in[0].size} (width,height) OUT {f2stack_out[0].shape} (height,width)")

        return f1stack_out,f2stack_out
class RegisterZ():
    def __init__(self):
        #VARIABLES
        self.f1_stack_arr_cropped = dyn_data.f1_Zstack_cropped
        self.f2_stack_arr_cropped = dyn_data.f2_Zstack_cropped

        self.f1slicenum = len(self.f1_stack_arr_cropped)
        self.f2slicenum = len(self.f2_stack_arr_cropped)
        print(f"F1 z:{self.f1slicenum} - F2 z:{self.f2slicenum}")

        
        #Determine shortest stack (either "1" or "2")
        self.shortstack = self.getshortstack()


        #CORRELATE Z STACKS TO DETERMINES SLICE OFFSET
        #!!!!!!-----self.correlate_zstacks()

    #----------------------------------##----------------------------------#
    #----------------------------------##----------------------------------#  
    def correlate_zstacks(self):
        #Get 5 points to judge the  slice stack



        

        #self.getf1f2_s_pos()

        #F1 SLICES --> 71
        #F2 SLICES --> 60

        
        t1 = 0
        t2 = 15
        t3 = 30
        t4 = 45
        t5 = 59
    

        pwinslice1,pcorrelation1, ssslice1, sscor1, mseslice1, mseval1 = self.slicerunner(t1)
        pwinslice2,pcorrelation2, ssslice2, sscor2, mseslice2, mseval2  = self.slicerunner(t2)
        pwinslice3,pcorrelation3, ssslice3, sscor3, mseslice3, mseval3  = self.slicerunner(t3)
        pwinslice4,pcorrelation4, ssslice4, sscor4, mseslice4, mseval4  = self.slicerunner(t4)
        pwinslice5,pcorrelation5, ssslice5, sscor5, mseslice5, mseval5  = self.slicerunner(t5)


        #PEARSON
        print("PEARSON CORRELATION")
        print(f"F1SLICE {t1} - F2 MAX SLICE {pwinslice1} - PCORR {pcorrelation1}")
        print(f"F1SLICE {t2} - F2 MAX SLICE {pwinslice2} - PCORR {pcorrelation2}")
        print(f"F1SLICE {t3} - F2 MAX SLICE {pwinslice3} - PCORR {pcorrelation3}")
        print(f"F1SLICE {t4} - F2 MAX SLICE {pwinslice4} - PCORR {pcorrelation4}")
        print(f"F1SLICE {t5} - F2 MAX SLICE {pwinslice5} - PCORR {pcorrelation5}")

        #STRUCTURAL SIMILARITY
        print("STRUCTURAL SIMILARITY")
        print(f"F1SLICE {t1} - F2 MAX SLICE {ssslice1} - SSCORR {sscor1}")
        print(f"F1SLICE {t2} - F2 MAX SLICE {ssslice2} - SSCORR {sscor2}")
        print(f"F1SLICE {t3} - F2 MAX SLICE {ssslice3} - SSCORR {sscor3}")
        print(f"F1SLICE {t4} - F2 MAX SLICE {ssslice4} - SSCORR {sscor4}")
        print(f"F1SLICE {t5} - F2 MAX SLICE {ssslice5} - SSCORR {sscor5}")

        #MEAN SQUARED ERROR
        print("MEAN SQUARED ERROR")
        print(f"F1SLICE {t1} - F2 MAX SLICE {mseslice1} - MSEVAL {mseval1}")
        print(f"F1SLICE {t2} - F2 MAX SLICE {mseslice2} - MSEVAL {mseval2}")
        print(f"F1SLICE {t3} - F2 MAX SLICE {mseslice3} - MSEVAL {mseval3}")
        print(f"F1SLICE {t4} - F2 MAX SLICE {mseslice4} - MSEVAL {mseval4}")
        print(f"F1SLICE {t5} - F2 MAX SLICE {mseslice5} - MSEVAL {mseval5}")    

    def getshortstack(self):
        if self.f1slicenum <= self.f2slicenum:
            shortstack = 1
            print("f1 stack longer")
        if self.f2slicenum < self.f1slicenum:
            shortstack = 2
            print("f2 stack shorter")
        print(f"shortstack {shortstack}")
        return shortstack
    def slicerunner(self,tchoice):
        pglad = None
        pmax = -1

        ssglad = None
        ssmax = -1

        mseglad = None
        msemin = float('inf')

        #tchoice = t1
        for i in range (0,70):
        # correlation, _ = pearsonr(f1_stack_arr_cropped[35],f2_stack_arr_cropped[i])
            f1_slice = f1_stack_arr_cropped[i].flatten()
            f2_slice = f2_stack_arr_cropped[tchoice].flatten()
            # Calculate Pearson correlation
            pcorrelation, _ = pearsonr(f1_slice, f2_slice)
            # Calculate Structural Similarity
            ssimilarity, _ = ssim(f1_slice, f2_slice, full=True, data_range=65535)
            # Calculate MSE
            mse = mean_squared_error(f1_slice, f2_slice)
            #print(f"t={tchoice} - F1 SLICE {i}, CORRELATION = {pcorrelation}")
            if pcorrelation > pmax: 
                pglad = i
                pmax = pcorrelation
            if ssimilarity > ssmax: 
                ssglad = i
                ssmax = ssimilarity
            if mse < msemin:
                mseglad = i
                msemin = mse
            #print(f"{tglad} ----- {tmax}")


        return pglad, pmax, ssglad, ssmax, mseglad, msemin
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class AutoScrollbar(tk.Scrollbar):
    ''' A scrollbar that hides itself if it's not needed.
        Works only if you use the grid geometry manager '''
    def set(self, lo, hi):
        if float(lo) <= 0.0 and float(hi) >= 1.0:
            self.grid_remove()
        else:
            self.grid()
            tk.Scrollbar.set(self, lo, hi)

    def pack(self, **kw):
        raise tk.TclError('Cannot use pack with this widget')

    def place(self, **kw):
        raise tk.TclError('Cannot use place with this widget')
class NucPickGUI(tk.Frame):
    def __init__(self,mainframe):

        self.inputs_directory_F1 = "FILE_1"
        self.inputs_directory_F2 = "FILE_2"
        self.inputs_directory_folder = "Inputs"
        self.processing_directory_folder = "Processing"
        self.p_MIP_directory = "MIP"
        self.p_MIP_raw_directory = "RAW_MIP"
        self.p_zslices_directory = "zslices"   

        #--
        #GENERATE CONTOURS
        self.nucmask__init__()
        #DEFINE DAPI DIRECTORY 
        self.f1dapichan = None
        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                self.f1dapichan = index
                self.F1_DAPI_dir = f"C{self.f1dapichan}_DAPI"


        #PULL INPUT DATA
        self.arr_nucleus_contours_coordinates = dyn_data.arr_nucleus_contours       #GET NUCLEUS CONTOUR DATA from first blob detector
        self.input_img_path = os.path.join(self.processing_directory_folder,"FILE_1",self.p_MIP_raw_directory, self.F1_DAPI_dir,f"F1_{self.F1_DAPI_dir}_MIP_.tif") #Path to pull drawn canvas image from
        self.masked_mip = os.path.join(self.processing_directory_folder,"masked_dna_mip.tif") #masked MIP reference file to generate new polygon in add_polygon

        #Define polygons container to pass to DynData
        self.polygons = []      #polgons container

        self.manpolypoints = [] #x,y points container for manually drawn polygon

        #Define nucleus counter
        self.nuccount = 0


        #Initialize TKINTER frame
        tk.Frame.__init__(self, master=mainframe)

        self.master.title('ZFISHER --- Nucleus Segmentation')

        # Set the initial size of the main frame
        mainframe.geometry("1024x1022")  # Example size: width=800, height=600

        # Vertical and horizontal scrollbars for canvas
        vbar = AutoScrollbar(self.master, orient='vertical')
        hbar = AutoScrollbar(self.master, orient='horizontal')
        vbar.grid(row=0, column=1, rowspan=2, sticky='ns')
        hbar.grid(row=2, column=0, sticky='we')

        # Create canvas 
        self.canvas = tk.Canvas(self.master, highlightthickness=0,
                                xscrollcommand=hbar.set, yscrollcommand=vbar.set)
        self.canvas.grid(row=0, column=0, sticky='nswe')
        self.canvas.update()  # wait till canvas is created
        vbar.configure(command=self.scroll_y)  # bind scrollbars to the canvas
        hbar.configure(command=self.scroll_x)
        # Make the canvas expandable
        self.master.rowconfigure(0, weight=1)
        self.master.columnconfigure(0, weight=1)

        self.image = Image.open(self.input_img_path)  # open image
        self.image = self.normalize_image(self.image)

        self.width, self.height = self.image.size
        print(self.width)
        print(self.height)
        self.baseHeight = self.height
        self.baseWidth = self.width
        self.imscale = 1.0  # scale for the canvaas image
        self.scalefactor = 1.0
        self.delta = 1.3  # zoom magnitude
        # Put image into container rectangle and use it to set proper coordinates to the image
        self.container = self.canvas.create_rectangle(0, 0, self.width, self.height, width=0)

        #---------------------------------------------------------------------------------------------------------------------------
        #MAKE CONTROL PANEL
        self.control_window = tk.Toplevel(self.master)
        self.control_window.title("Control Panel - Nucleus Segmentation")

        self.list_label = tk.Label(self.control_window, text="-----Number of Nuclei:-----")
        self.list_label.grid(row=1, column=1, columnspan=2)

        self.count_label = tk.Label(self.control_window, text="0")
        self.count_label.grid(row=2, column=1, columnspan=2)

        self.spacer_label = tk.Label(self.control_window, text="------------------------------------")
        self.spacer_label.grid(row=6, column=1, columnspan=2)  

        self.mousepos_label = tk.Label(self.control_window, text="Mouse Position: (0000.00,0000.00)", font=("Courier"))
        self.mousepos_label.grid(row=7, column=1, columnspan=2)
        x_lab = 0
        y_lab = 0
        self.mousepos_label.config(text=f"Mouse Position: ({float(x_lab):07.2f}, {float(y_lab):07.2f})")
        
        self.finish_button = tk.Button(self.control_window, text="Finalize Nuclei Picking", command=self.finalize_nucpicking)
        self.finish_button.grid(row=9, column=1, columnspan=2)     

        self.ManPoly_toggle = tk.BooleanVar(value=False)
        self.polydraw_toggle_checkbox = tk.Checkbutton(self.control_window, text="Manual Polygon", variable=self.ManPoly_toggle)
        self.polydraw_toggle_checkbox.grid(row=10, column=1, columnspan=2) 
        #---------------------------------------------------------------------------------------------------------------------------

        #BINDINGS
        self.canvas.bind('<Configure>', self.show_image)  # canvas is resized
        self.canvas.bind("<Button-1>", self.remove_polygon) #click polygon to remove
        self.canvas.bind("<Button-1>", self.man_poly_point, add='+') #add Vertex for manual polygon 

        #ADD POLYGON (RIGHT MOUSE CLICK)
        #OLD --> self.canvas.bind("<Button-2>", self.add_polygon) #click polygon to remove
        if platform.system() == "Darwin":  # macOS
            self.canvas.bind('<Button-2>', self.add_polygon) # right-click for macOS
            self.canvas.bind('<Button-2>', self.man_poly_point_complete, add='+') # close manual polygon
        elif platform.system() == "Linux":  # Linux
            self.canvas.bind('<Button-3>', self.add_polygon) # right-click for Linux
            self.canvas.bind('<Button-3>', self.man_poly_point_complete, add='+') # close manual polygon

        #DRAG (WHEEL CLICK)
        #OLD --> 
        # self.canvas.bind('<ButtonPress-3>', self.move_from)
        # self.canvas.bind('<B3-Motion>',     self.move_to)
        if platform.system() == "Darwin":  # macOS
             self.canvas.bind('<ButtonPress-3>', self.move_from)
             self.canvas.bind('<B3-Motion>',     self.move_to)
        elif platform.system() == "Linux":  # Linux
             self.canvas.bind('<ButtonPress-2>', self.move_from)
             self.canvas.bind('<B2-Motion>',     self.move_to)

        #ZOOM (MOUSEWHEEL)
        #OLD --> self.canvas.bind("<MouseWheel>", self.zoom)
        if platform.system() == "Darwin":  # macOS
            self.canvas.bind("<MouseWheel>", self.zoom)
        elif platform.system() == "Linux":  # Linux
            self.canvas.bind("<Button-4>", self.zoom)
            self.canvas.bind("<Button-5>", self.zoom)

        self.canvas.bind("<Motion>", self.motion)
        self.add_p_event_processed = False

        #Draw Nucleus POLYGONS --------------------------------------------------------------------------------------------#
        self.draw_init_polygons()
        self.show_image()

    #--FUNCTIONS---------------------------------------------------
        
        self.zoom_level = 1.0  # Initial zoom level

        self.scrollview = [0,0]


    def normalize_image(self,image):
        # Convert the image to a numpy array
        image_array = np.array(image).astype(np.float32)
        
        # Find the minimum and maximum pixel values
        min_val = image_array.min()
        max_val = image_array.max()
        
        print(f"MIN VAL {min_val}")
        print(f"MAX VAL {max_val}")
        # Normalize the image to the range [0, 1]
        normalized_array = (image_array - min_val) / (max_val - min_val)
        
        # Optionally, scale to [0, 255] for 8-bit representation
        normalized_array = (normalized_array * 255).astype(np.uint8)
        
        # Convert back to PIL Image
        normalized_image = Image.fromarray(normalized_array)
        return normalized_image
    
    def motion(self, event):
        # Calculate the mouse position in the original image
        mouse_x = self.canvas.canvasx(event.x)
        mouse_y = self.canvas.canvasy(event.y)

        #mouse_x = self.canvas.canvasx(event.x)/self.imscale
        #mouse_y = self.canvas.canvasy(event.y)/self.imscale

        #x_lab = "{:.2f}".format(mouse_x)
        #y_lab = "{:.2f}".format(mouse_y)
        x_lab = float(mouse_x)
        y_lab = float(mouse_y)
        #self.mousepos_label.config(text=f"Mouse Position: ({x_lab}, {y_lab})")
        #self.mousepos_label.config(text=f"Mouse Position: ({x_lab:03}, {y_lab:03})")
        #self.mousepos_label.config(text=f"Mouse Position: ({x_lab:6.2f}, {y_lab:6.2f})")
        self.mousepos_label.config(text=f"Mouse Position: ({float(x_lab):07.2f}, {float(y_lab):07.2f})")

    def finalize_nucpicking(self):
        self.update_nuc_count
        dyn_data.nucleuscount = self.nuccount

        
        #Set Coordinates back to base
        print(self.canvas.coords(self.container))
        print(self.bbox)


        dx = -self.canvas.coords(self.container)[0]
        dy = -self.canvas.coords(self.container)[1]
        self.canvas.move('all', dx,dy)

        #sbbox = self.canvas.bbox(self.container)
        #swidth = sbbox[2] - sbbox[0]  # Calculate width
        #sheight = sbbox[3] - sbbox[1]  # Calculate height

        self.canvas.scale('all', 0,0, 1/self.imscale, 1/self.imscale)
        #self.show_image()

        #updates polygons
        polygon_ids = self.canvas.find_withtag('polygon')

        self.polygons = []

        nucnum = 1
        for poly_id in polygon_ids:
            coords = self.canvas.coords(poly_id)
            self.polygons.append([nucnum]+ [coords])
            print(coords)
            nucnum += 1


        #nucpoly_output_array = [[row[0], self.canvas.coords(row[1])] for row in self.polygons]
        #dyn_data.nucpolygons = nucpoly_output_array
        dyn_data.nucpolygons = self.polygons
        print("ROW 1")
        print(dyn_data.nucpolygons[0])

        print("FINISHED NUC PICK")
        #self.master.destroy()
        mainapp.toROIpick()
        

    def update_nuc_count(self):
        self.nuccount = len(self.polygons)
        self.count_label.config(text=self.nuccount)
    
   
    #DRAG (CLICK AND DRAG SCREEN)---------------------------------------------------------------------#
    def move_from(self, event):
        ''' Remember previous coordinates for scrolling with the mouse '''
        self.canvas.scan_mark(event.x, event.y)
    def move_to(self, event):
        ''' Drag (move) canvas to the new position '''
        self.canvas.scan_dragto(event.x, event.y, gain=1)
        self.show_image()  # redraw the image
    
    #ZOOM---------------------------------------------------------------------------------------------#
    def zoom(self, event):
        print("scrolling")
        ''' Zoom with mouse wheel '''
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        bbox = self.canvas.bbox(self.container)  # get image area
        if bbox[0] < x < bbox[2] and bbox[1] < y < bbox[3]: pass  # Ok! Inside the image
        else: return  # zoom only inside image area
        scale = 1.0
        # Respond to Linux (event.num) or Windows (event.delta) wheel event
        if event.num == 5 or event.delta > 0:  # scroll down
            print("down)")
            i = min(self.width, self.height)
            if int(i * self.imscale) < 30: return  # image is less than 30 pixels
            self.imscale /= self.delta
            scale        /= self.delta
        if event.num == 4 or event.delta < 0:  # scroll up
            print("up")
            i = min(self.canvas.winfo_width(), self.canvas.winfo_height())
            if i < self.imscale: return  # 1 pixel is bigger than the visible area
            self.imscale *= self.delta
            scale        *= self.delta

        print(self.imscale)
        self.canvas.scale('all', x, y, scale, scale)  # rescale all canvas objects
        self.show_image()
    #-------------------------------------------------------------------------------------------------#
    def show_image(self, event=None):
        ''' Show image on the Canvas '''
        bbox1 = self.canvas.bbox(self.container)  # get image area
        # Remove 1 pixel shift at the sides of the bbox1
        bbox1 = (bbox1[0] + 1, bbox1[1] + 1, bbox1[2] - 1, bbox1[3] - 1)
        bbox2 = (self.canvas.canvasx(0),  # get visible area of the canvas
                self.canvas.canvasy(0),
                self.canvas.canvasx(self.canvas.winfo_width()),
                self.canvas.canvasy(self.canvas.winfo_height()))
        bbox = [min(bbox1[0], bbox2[0]), min(bbox1[1], bbox2[1]),  # get scroll region box
                max(bbox1[2], bbox2[2]), max(bbox1[3], bbox2[3])]
        if bbox[0] == bbox2[0] and bbox[2] == bbox2[2]:  # whole image in the visible area
            bbox[0] = bbox1[0]
            bbox[2] = bbox1[2]
        if bbox[1] == bbox2[1] and bbox[3] == bbox2[3]:  # whole image in the visible area
            bbox[1] = bbox1[1]
            bbox[3] = bbox1[3]
        self.canvas.configure(scrollregion=bbox)  # set scroll region
        self.update_nuc_count()

        print(f"bbox1 ({bbox1[0]},{bbox1[1]},{bbox1[2]},{bbox1[3]})")
        print(f"bbox2 ({bbox2[0]},{bbox2[1]},{bbox2[2]},{bbox2[3]})")
        print(f"bbox ({bbox[0]},{bbox[1]},{bbox[2]},{bbox[3]})")
        # Store bbox1 for later use
        self.bbox1 = bbox1

        x1 = max(bbox2[0] - bbox1[0], 0)  # get coordinates (x1,y1,x2,y2) of the image tile
        y1 = max(bbox2[1] - bbox1[1], 0)
        x2 = min(bbox2[2], bbox1[2]) - bbox1[0]
        y2 = min(bbox2[3], bbox1[3]) - bbox1[1]

        #Scaled positions for adding polygons from cropchunk add_polygon func
        self.si_x1 = x1
        self.si_y1 = y1
        self.si_x2 = x2
        self.si_y2 = y2

        self.grey_image_offset_x = bbox1[0] - bbox2[0]
        self.grey_image_offset_y = bbox1[1] - bbox2[1]       

        print(f"x1,y1,x2,y2 {x1},{y1},{x2},{y2}")
        print(f"grey offset {self.grey_image_offset_x}, {self.grey_image_offset_y}")
        if int(x2 - x1) > 0 and int(y2 - y1) > 0:  # show image if it in the visible area
            x = min(int(x2 / self.imscale), self.width)   # sometimes it is larger on 1 pixel...
            y = min(int(y2 / self.imscale), self.height)  # ...and sometimes not
            image = self.image.crop((int(x1 / self.imscale), int(y1 / self.imscale), x, y))
            imagetk = ImageTk.PhotoImage(image.resize((int(x2 - x1), int(y2 - y1))))
            self.imageid = self.canvas.create_image(max(bbox2[0], bbox1[0]), max(bbox2[1], bbox1[1]),
                                            anchor='nw', image=imagetk)
            self.bgimage= self.canvas.lower(self.imageid)  # set image into background
            self.canvas.imagetk = imagetk  # keep an extra reference to prevent garbage-collection
        print(self.canvas.coords(self.container))
    def draw_init_polygons(self):
        prev_column_value = None
        coordinates = []
        cell_index_table = [0, 0]
        for row in self.arr_nucleus_contours_coordinates:
            #print(self.arr_nucleus_contours_coordinates)
            if prev_column_value is None or row[0] == prev_column_value:
                temp_array = row[2:]
                coordinates = np.hstack((coordinates, temp_array))
            else:
                new_row = [row[0], str(coordinates)]
                cell_index_table = np.vstack((cell_index_table, new_row))
                scaled_coordinates = [coord  if i % 2 == 0 else coord  for i, coord in
                                    enumerate(coordinates)]
                self.draw_polygon(scaled_coordinates, prev_column_value)  # Pass index to draw_polygon function
                coordinates = row[2:]
            prev_column_value = row[0]
        
        #----Remove Edge of FOV Nucs from autopick-------
        temp_polygons = []
        print(f"len temp poly {len(temp_polygons)}")
        counter = 0
        for p_row in self.polygons:
            index, polygon, indx_label = p_row

            coords = self.canvas.coords(polygon)


            heightcutoff = self.height * 0.99
            if 0 in coords:
                print("FOUND A ZERO")
                counter += 1

                self.canvas.delete(indx_label)
                self.canvas.delete(polygon)
            elif any(x >= heightcutoff for x in coords):
                print("FOUND A MAX")
                counter += 1

                self.canvas.delete(indx_label)
                self.canvas.delete(polygon)                
            else:
                temp_polygons.append(p_row)

        self.polygons = temp_polygons
        self.update_nuc_count()


    def draw_polygon(self, coordinates, index):


        index = int(index)
        self.polygons = sorted(self.polygons, key=lambda x: x[0])





        scaled_coordinates = [coord * 1 for coord in coordinates]
        #print(scaled_coordinates)
       #print(len(scaled_coordinates))
        polygon = self.canvas.create_polygon(scaled_coordinates, outline='red', fill="", tags=('polygon'))

 
        newpolycoords = self.canvas.coords(polygon)
            
        #REMOVE polygon if OVERLAPPING
        n_num_points = len(newpolycoords) // 2
        n_avg_x = sum(newpolycoords[i] for i in range(0, len(newpolycoords), 2)) / n_num_points
        n_avg_y = sum(newpolycoords[i] for i in range(1, len(newpolycoords), 2)) / n_num_points

        bboxwidth = 20
        bboxheight = 20

        n_bbox = self.canvas.create_rectangle(n_avg_x - bboxwidth/2, n_avg_y - bboxheight/2, n_avg_x + bboxwidth/2, n_avg_y + bboxheight/2)

        overlapping_items = self.canvas.find_overlapping(n_avg_x - bboxwidth/2, n_avg_y - bboxheight/2, n_avg_x + bboxwidth/2, n_avg_y + bboxheight/2)
        for item_id in overlapping_items:
            tag = self.canvas.gettags(item_id)
            if "polygon" in tag:
                if item_id != polygon: 
                    print(f"OLD POLY OVERLAP {item_id}")
                    self.canvas.delete(polygon)
                    self.canvas.delete(n_bbox)
                    return
        self.canvas.delete(n_bbox)


        #Re-index the polygon if there is a gap in the self.polygons ID list
        missing_numbers = []
        expected_number = 1
        for row in self.polygons:
            while row[0] > expected_number:
                missing_numbers.append(expected_number)
                expected_number += 1
            expected_number = row[0] + 1
        if len(missing_numbers) > 0: index = missing_numbers[0]

        print(f"polygon index {index}")
        for rowinfo in self.polygons:   # Storing index, polygon, and text index
            p_index,p_poly,p_tindex = rowinfo
            if index == p_index:
                print("dupe")


        indx_label = self.display_index_on_polygon(polygon, int(index))  # Display index on polygon

        self.polygons.append((index, polygon, indx_label))  # Storing index, polygon, and text index
        self.update_nuc_count()
        
        # Disable mouse events on the polygon
        self.canvas.tag_bind(polygon, '<Button-1>', lambda event: "break")
        self.canvas.tag_bind(polygon, '<ButtonRelease-1>', lambda event: "break")
        self.canvas.tag_bind(polygon, '<B1-Motion>', lambda event: "break")
        self.canvas.tag_bind(polygon, '<Enter>', lambda event, p=polygon: self.highlight_polygon(p))
        self.canvas.tag_bind(polygon, '<Leave>', lambda event, p=polygon: self.unhighlight_polygon(p))
    def display_index_on_polygon(self, polygon, index):
        coords = self.canvas.coords(polygon)
        x_coords = [int(coords[i]) for i in range(0, len(coords), 2)]
        y_coords = [int(coords[i]) for i in range(1, len(coords), 2)]
        x_center = sum(x_coords) // len(x_coords)  # Using integer division
        y_center = sum(y_coords) // len(y_coords)  # Using integer division
        text_item = self.canvas.create_text(x_center, y_center, text=str(index), font=('Arial', 12), fill='red')
        return text_item

    def remove_polygon(self,event):
        if self.ManPoly_toggle.get() == True: return
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y) 
        item = self.canvas.find_closest(x, y)
        if item and item[0] in (polygon[1] for polygon in self.polygons):
            for polygon_index, (index, p, ilabel) in enumerate(self.polygons):
                if p == item[0]:
                    print(f"Deleted polygon {index}")
                    self.canvas.delete(ilabel)
                    self.canvas.delete(p)  
                    print(f"polylenb {len(self.polygons)}")
                    del self.polygons[polygon_index]
                    print(f"polylena {len(self.polygons)}")
                    
                    for row in self.polygons:
                        print(row)
                    break  # Break the loop after finding the polygon to delete
        if item and item[0] in (polygon[2] for polygon in self.polygons):
            for polygon_index, (index, p, ilabel) in enumerate(self.polygons):
                if ilabel == item[0]:
                    print(f"Deleted polygon {index}")
                    self.canvas.delete(ilabel)
                    self.canvas.delete(p)  
                    print(f"polylenb {len(self.polygons)}")
                    del self.polygons[polygon_index]
                    print(f"polylena {len(self.polygons)}")
                    for row in self.polygons:
                        print(row)
                    break  # Break the loop after finding the polygon to delete
        self.update_nuc_count()

    def add_polygon(self, event):
        if self.ManPoly_toggle.get() == True: return
        mousex = self.canvas.canvasx(event.x)
        mousey = self.canvas.canvasy(event.y) 
        maskedmip = cv2.imread(self.masked_mip, cv2.IMREAD_GRAYSCALE)
        crop_width, crop_height = 300, 300  # Crop pic size

        scropx1 = max(0, int(mousex - crop_width * self.imscale // 2)) 
        scropy1 = max(0, int(mousey - crop_height * self.imscale // 2))
        scropx2 = min(maskedmip.shape[1], int(mousex + crop_width * self.imscale // 2))
        scropy2 = min(maskedmip.shape[0], int(mousey + crop_height * self.imscale // 2))

        print(f"six1_ {self.si_x1},{self.si_y1}")
        smousex = event.x_root - self.canvas.winfo_rootx()
        smousey = event.y_root - self.canvas.winfo_rooty()



        bbcoords = self.canvas.coords(self.container)

        if  self.imscale == 1.0:
            print("NO ZOOM")
            smousex = smousex*1/self.imscale + (max(bbcoords[0],self.si_x1))
            smousey = smousey*1/self.imscale + (max(self.grey_image_offset_y,self.si_y1))
        elif self.imscale > 1.0:
            print("ZOOMED IN")
            smousex = smousex*1/self.imscale + (max(self.grey_image_offset_x,self.si_x1*1/self.imscale))
            smousey = smousey*1/self.imscale + (max(self.grey_image_offset_y,self.si_y1*1/self.imscale))
        elif self.imscale < 1.0:
            print("ZOOMED OUT")
            smousex = ((smousex - max(0,self.grey_image_offset_x))*1/self.imscale) + self.si_x1*1/self.imscale
            smousey = ((smousey - max(0,self.grey_image_offset_y))*1/self.imscale) + self.si_y1*1/self.imscale


        socropx1 = max(0, int(smousex - (crop_width // 2))) 
        socropy1 = max(0, int(smousey - (crop_height // 2)))
        socropx2 = min(maskedmip.shape[0], int(smousex  + (crop_width // 2)))
        socropy2 = min(maskedmip.shape[1], int(smousey  + (crop_height // 2)))

        print(f"imscale {self.imscale}")
        print(f"grey offset {self.grey_image_offset_x}, {self.grey_image_offset_y}")
        print(f"mousepos {mousex},{mousey}")
        print(f"smousepos {smousex},{smousey}")
        print(f"scrop_ {scropx1},{scropy1},{scropx2},{scropy2}")
        print(f"socrop {socropx1},{socropy1},{socropx2},{socropy2}")
        print(f"Canvas position: x={event.x}, y={event.y}")
        print(f"Screen position: x={event.x_root}, y={event.y_root}")
        print(f"six1_{self.si_x1},{self.si_y1}")
        print(f"bbox {self.canvas.coords(self.container)}")
        canvas_position= [event.x,event.y]
        screen_position= [event.x_root, event.y_root]

       # self.canvas.create_rectangle(scropx1,scropy1,scropx2,scropy2, fill="", outline="white")

        cropped_image = maskedmip[socropy1:socropy2, socropx1:socropx2]

            #Save the cropped image
        print("cropped")
        cropfile_path = os.path.join(setup.processing_directory_folder,"cropchunk.tif")
        cv2.imwrite(cropfile_path, cropped_image) 



        #Trace the cropped image
        cropfile= cv2.imread(cropfile_path, cv2.IMREAD_GRAYSCALE)
        contours, hierarchy = cv2.findContours(cropfile, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        #print(contours)
        contoured_cropfile = cropfile.copy()
        cv2.drawContours(contoured_cropfile, contours, -1, (0,255,0), 2, cv2.LINE_AA)

        cont_cropfile_path = os.path.join(setup.processing_directory_folder,"contour_cropchunk.tif")
        cv2.imwrite(cont_cropfile_path, contoured_cropfile)

        index = 1
        isolated_count = 0
        cluster_count = 0

        arr_nucleus_contours = np.empty([1,4])

        for cntr in contours:
            epsilon = 0.001 * cv2.arcLength(cntr, True)
            approx = cv2.approxPolyDP(cntr, epsilon, True)
            n = approx.ravel() 
            i = 0
            for j in n : 
                if(i % 2 == 0): 
                    x = n[i]  
                    y = n[i + 1] 
                    #print(f"---{x},{y}")
                    #Generate contours cooridinates array to pass into tkinter visualization
                    arr_nucleus_contours = np.vstack((arr_nucleus_contours , [index,i,x,y]))
                i += 1
            index += 1
        arr_nucleus_contours = np.delete(arr_nucleus_contours,(0),axis=0)

        #print(len(arr_nucleus_contours))
        print("_________________________")
        if len(arr_nucleus_contours) <=0: 
            print("no contours found")
            self.add_p_event_processed = False
            return
        else: 

            if  self.imscale == 1.0:
                print("NO ZOOM")
                xo = scropx1
                yo = scropy1
                ofx = 0
                ofy = 0
            elif self.imscale > 1.0:
                print("ZOOMED IN")

                cmousex = self.canvas.canvasx(event.x)
                cmousey = self.canvas.canvasy(event.y)
                xo = int(cmousex - crop_width * self.imscale // 2)
                yo = int(cmousey - crop_height * self.imscale // 2)

                #self.canvas.create_oval(xo-10,yo-10,xo+10,yo+10, fill="", outline="white")

                
                ofx = 0
                ofy = 0

                bbcoords = self.canvas.coords(self.container)
                print (f"cmouse {cmousex},  -->{cmousex - crop_width*self.imscale // 2 }")

                if bbcoords[0]<0 : 
                    if (cmousex - crop_width*self.imscale // 2) < bbcoords[0]:
                        print("lilx")
                        ofx = (bbcoords[0])- (cmousex - crop_width*self.imscale // 2) 
                if bbcoords[1]<0 : 
                    print((cmousey - crop_height*self.imscale // 2))
                    if (cmousey - crop_height*self.imscale // 2) < bbcoords[1]:
                        print("lily")
                        ofy = (bbcoords[1]) - (cmousey - crop_height*self.imscale // 2) 


            elif self.imscale < 1.0:
                print("ZOOMED OUT")
                cmousex = self.canvas.canvasx(event.x)
                cmousey = self.canvas.canvasy(event.y)
                xo = int(cmousex - crop_width * self.imscale // 2)
                yo = int(cmousey - crop_height * self.imscale // 2)

              #  self.canvas.create_oval(xo-10,yo-10,xo+10,yo+10, fill="", outline="white")

                
                ofx = 0
                ofy = 0

                bbcoords = self.canvas.coords(self.container)
                print (f"cmouse {cmousex},  -->{cmousex - crop_width*self.imscale // 2 }")
                if bbcoords[0]>0 : 
                    if cmousex - crop_width*self.imscale // 2 < bbcoords[0]:
                        print("bigginx")
                        ofx = (bbcoords[0]) - (cmousex - crop_width*self.imscale // 2) 
                if bbcoords[1]>0 : 
                    if cmousey - crop_height*self.imscale // 2 < bbcoords[1]:
                        print("bigginy")
                        ofy = (bbcoords[1]) - (cmousey - crop_height*self.imscale // 2) 



                
            for row in arr_nucleus_contours:
                row[2] = row[2]*self.imscale + xo + ofx
                row[3] = row[3]*self.imscale + yo + ofy



            if len(arr_nucleus_contours) <= 0: 
                print("No contours")
                self.add_p_event_processed = False
                return
            else:
                self.draw_added_polygons(arr_nucleus_contours)

    def old_add_polygon(self, event):
        mousex = self.canvas.canvasx(event.x)
        mousey = self.canvas.canvasy(event.y) 
        maskedmip = cv2.imread(self.masked_mip, cv2.IMREAD_GRAYSCALE)
        crop_width, crop_height = 200, 200  # Crop pic size

        scropx1 = max(0, int(mousex - crop_width * self.imscale // 2)) 
        scropy1 = max(0, int(mousey - crop_height * self.imscale // 2))
        scropx2 = min(maskedmip.shape[1], int(mousex + crop_width * self.imscale // 2))
        scropy2 = min(maskedmip.shape[0], int(mousey + crop_height * self.imscale // 2))

        print(f"six1_ {self.si_x1},{self.si_y1}")
        smousex = event.x_root - self.canvas.winfo_rootx()
        smousey = event.y_root - self.canvas.winfo_rooty()





        if  self.imscale == 1.0:
            print("NO ZOOM")
            smousex = smousex*1/self.imscale + (max(self.grey_image_offset_x,self.si_x1))
            smousey = smousey*1/self.imscale + (max(self.grey_image_offset_y,self.si_y1))
        elif self.imscale > 1.0:
            print("ZOOMED IN")
            smousex = smousex*1/self.imscale + (max(self.grey_image_offset_x,self.si_x1*1/self.imscale))
            smousey = smousey*1/self.imscale + (max(self.grey_image_offset_y,self.si_y1*1/self.imscale))
        elif self.imscale < 1.0:
            print("ZOOMED OUT")
            smousex = ((smousex - max(0,self.grey_image_offset_x))*1/self.imscale) + self.si_x1*1/self.imscale
            smousey = ((smousey - max(0,self.grey_image_offset_y))*1/self.imscale) + self.si_y1*1/self.imscale


        socropx1 = max(0, int(smousex - (crop_width // 2))) 
        socropy1 = max(0, int(smousey - (crop_height // 2)))
        socropx2 = min(maskedmip.shape[0], int(smousex  + (crop_width // 2)))
        socropy2 = min(maskedmip.shape[1], int(smousey  + (crop_height // 2)))

        print(f"imscale {self.imscale}")
        print(f"grey offset {self.grey_image_offset_x}, {self.grey_image_offset_y}")
        print(f"mousepos {mousex},{mousey}")
        print(f"smousepos {smousex},{smousey}")
        print(f"scrop_ {scropx1},{scropy1},{scropx2},{scropy2}")
        print(f"socrop {socropx1},{socropy1},{socropx2},{socropy2}")
        print(f"Canvas position: x={event.x}, y={event.y}")
        print(f"Screen position: x={event.x_root}, y={event.y_root}")
        canvas_position= [event.x,event.y]
        screen_position= [event.x_root, event.y_root]
        self.canvas.create_rectangle(scropx1,scropy1,scropx2,scropy2, fill="", outline="white")

        cropped_image = maskedmip[socropy1:socropy2, socropx1:socropx2]

            #Save the cropped image
        print("cropped")
        cropfile_path = os.path.join(setup.processing_directory_folder,"cropchunk.tif")
        cv2.imwrite(cropfile_path, cropped_image) 



        #Trace the cropped image
        cropfile= cv2.imread(cropfile_path, cv2.IMREAD_GRAYSCALE)
        contours, hierarchy = cv2.findContours(cropfile, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        #print(contours)
        contoured_cropfile = cropfile.copy()
        cv2.drawContours(contoured_cropfile, contours, -1, (0,255,0), 2, cv2.LINE_AA)

        cont_cropfile_path = os.path.join(setup.processing_directory_folder,"contour_cropchunk.tif")
        cv2.imwrite(cont_cropfile_path, contoured_cropfile)

        index = 1
        isolated_count = 0
        cluster_count = 0

        arr_nucleus_contours = np.empty([1,4])

        for cntr in contours:
            epsilon = 0.001 * cv2.arcLength(cntr, True)
            approx = cv2.approxPolyDP(cntr, epsilon, True)
            n = approx.ravel() 
            i = 0
            for j in n : 
                if(i % 2 == 0): 
                    x = n[i]  + scropx1
                    y = n[i + 1]  + scropy1
                    #print(f"---{x},{y}")
                    #Generate contours cooridinates array to pass into tkinter visualization
                    arr_nucleus_contours = np.vstack((arr_nucleus_contours , [index,i,x,y]))
                i += 1
            index += 1
        arr_nucleus_contours = np.delete(arr_nucleus_contours,(0),axis=0)

        #print(len(arr_nucleus_contours))
        print(f"contours {arr_nucleus_contours}")
        print("_________________________")
        if len(arr_nucleus_contours) <=0: 
            print("no contours found")
            self.add_p_event_processed = False
            return
        else: 

            #Check if polygon is on edge of screen
            print(f"crop {scropx1},{scropy2},{scropx2},{scropy2}")
            temp_anc = arr_nucleus_contours
            keys = []
            for row in temp_anc:
                if row[0] not in keys:
                    keys.append(row[0])
            
            print(f"keys {keys}")
            for row in arr_nucleus_contours:    #[index,i,x,y]
                print(row)

                #Scale values based on scale and offset
                #NO ZOOM
                if  self.imscale == 1.0:
                #Is the crop on the edge of the screen?
                    if scropx1 <= 2 or scropy1 <=2 or scropx2 >= self.baseWidth-2 or scropy2 >= self.baseHeight-2:
                        #LEFT
                        if scropx1<= 2:
                            if row[2] >= scropx2-2 and row[3] <= scropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]                                
                                print("UPRIGHTCORNER")
                            if row[2] >= scropx2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")
                            if row[3] >= scropy2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("BOT")    
                            if row[3] <= scropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("UP")                               
                        #BOT
                        if scropy2 >= self.baseHeight-2:
                            if row[3] <= scropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("UPPER")
                            if row[2] <= scropx1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("LEFT")
                            if row[2] >= scropx2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")


                        #TOP
                        if scropy1<= 2:
                            if row[2] <= scropx1+2 and row[3] >= scropy2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]                                
                                print("BOTLEFTCORNER")
                            if row[2] <= scropx1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("LEFT")
                            if row[2] >= scropx2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")
                        #RIGHT
                        if scropx2 >= self.baseWidth-2:
                            if row[2] <= scropx1+2 and row[3] <= scropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]                                
                                print("UPLEFTCORNER")
                            if row[2] <= scropx1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")
                            if row[3] >= scropy2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("BOT")    
                            if row[3] <= scropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("UP")    
                    else:
                        if row[2] <= scropx1 or row[3] <= scropy1 or row[2] >= scropx2-2 or row[3] >= scropy2 - 2:
                            indextodelete = row[0]
                            temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]
                #ZOOMED IN
                elif self.imscale > 1.0:
                    #smousex = smousex*1/self.imscale + (max(self.grey_image_offset_x,self.si_x1*1/self.imscale))
                    #smousey = smousey*1/self.imscale + (max(self.grey_image_offset_y,self.si_y1*1/self.imscale))
                    if socropx1 <= 2 or socropy1 <=2 or socropx2 >= self.baseWidth-2 or socropy2 >= self.baseHeight-2:
                        #LEFT
                        if socropx1<= 2:
                            if row[2] >= socropx2-2 and row[3] <= socropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]                                
                                print("UPRIGHTCORNER")
                            if row[2] >= socropx2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")
                            if row[3] >= socropy2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("BOT")    
                            if row[3] <= socropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("UP")                               
                        #BOT
                        if socropy2 >= self.baseHeight-2:
                            if row[3] <= socropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("UPPER")
                            if row[2] <= socropx1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("LEFT")
                            if row[2] >= socropx2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")


                        #TOP
                        if socropy1<= 2:
                            if row[2] <= socropx1+2 and row[3] >= socropy2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]                                
                                print("BOTLEFTCORNER")
                            if row[2] <= socropx1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("LEFT")
                            if row[2] >= socropx2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")
                        #RIGHT
                        if socropx2 >= self.baseWidth-2:
                            if row[2] <= socropx1+2 and row[3] <= socropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]                                
                                print("UPLEFTCORNER")
                            if row[2] <= socropx1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("RIGHT")
                            if row[3] >= socropy2-2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("BOT")    
                            if row[3] <= socropy1+2:
                                indextodelete = row[0]
                                temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]   
                                print("UP")    
                    else:
                        if row[2] <= socropx1 or row[3] <= socropy1 or row[2] >= socropx2-2 or row[3] >= socropy2 - 2:
                            indextodelete = row[0]
                            temp_anc = [rowt for rowt in temp_anc if rowt[0] != indextodelete]
                
                #ZOOMED OUT
                elif self.imscale < 1.0:
                    print("ZOOMED OUT")
                   # smousex = ((smousex - max(0,self.grey_image_offset_x))*1/self.imscale) + self.si_x1*1/self.imscale
                   # smousey = ((smousey - max(0,self.grey_image_offset_y))*1/self.imscale) + self.si_y1*1/self.imscale
                    pass


            arr_nucleus_contours = temp_anc
            if len(arr_nucleus_contours) <= 0: 
                print("No contours")
                self.add_p_event_processed = False
                return
            else:
                self.draw_added_polygons(arr_nucleus_contours)

    def man_poly_point(self,event):
        if self.ManPoly_toggle.get() == False: return
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y) 
        #x,y = event.x, event.y
        self.manpolypoints.append((x,y))
        self.draw_mantemppoly()

    def man_poly_point_complete(self,event):
        if self.ManPoly_toggle.get() == False: return
        print("COMPLETE MANUAL POLYGON")
        if len(self.manpolypoints) > 2:
            #self.canvas.create_line(self.manpolypoints, fill='white', width=2, tags='tempmanpolygon')
            self.canvas.delete('tempmanpolygon')
            newindex = len(self.polygons) + 1
            self.draw_polygon(self.manpolypoints, newindex)
            self.manpolypoints = []

    def draw_mantemppoly(self):
        # Redraw polygon based on current points
        self.canvas.delete('tempmanpolygon')
        if len(self.manpolypoints) > 1:
            self.canvas.create_line(self.manpolypoints, fill='green', width=2, tags='tempmanpolygon')


    def draw_added_polygons(self,nuc_cntrs):
        #print(nuc_cntrs)
        prev_column_value = None
        coordinates = []
        cell_index_table = [0, 0]
        for row in nuc_cntrs:
            if prev_column_value is None or row[0] == prev_column_value:
                temp_array = row[2:]
                coordinates = np.hstack((coordinates, temp_array))
            else:
                new_row = [row[0], str(coordinates)]
                cell_index_table = np.vstack((cell_index_table, new_row))
                scaled_coordinates = [coord  if i % 2 == 0 else coord  for i, coord in
                                    enumerate(coordinates)]
                newindex = len(self.polygons) + 1
                    

            
                self.draw_polygon(scaled_coordinates, newindex)  # Pass index to draw_polygon function
                coordinates = row[2:]
            prev_column_value = row[0]
        scaled_coordinates = [coord if i % 2 == 0 else coord for i, coord in enumerate(coordinates)]
        newindex = len(self.polygons) + 1
        self.draw_polygon(scaled_coordinates, newindex)

        self.add_p_event_processed = False
    #------POLYGON HIGHLIGHTING-----------------------------------
    def highlight_polygon(self, polygon):
        self.canvas.itemconfig(polygon, fill='yellow')
    def unhighlight_polygon(self, polygon):
        self.canvas.itemconfig(polygon, fill='')
    #-------------------------------------------------------------  
    #-----SCROLLING CANVAS WITH SCROLLBARS------------------------    
    def scroll_y(self, *args, **kwargs):

        old_view = self.scrollview
        ''' Scroll canvas vertically and redraw the image '''
        self.canvas.yview(*args, **kwargs)  # scroll vertically
        new_view = self.canvas.yview()

        shift = (old_view[0] - new_view[0], old_view[1] - new_view[1])
        
        self.scrollview[0] += shift[0]
        self.scrollview[1] += shift[1]

        print(f"sview {self.scrollview[0]},{self.scrollview[1]}")
        self.show_image()  # redraw the image
    def scroll_x(self, *args, **kwargs):

        old_view = self.scrollview
        ''' Scroll canvas horizontally and redraw the image '''
        self.canvas.xview(*args, **kwargs)  # scroll horizontally
        new_view = self.canvas.xview()
        shift = (old_view[0] - new_view[0], old_view[1] - new_view[1])

        self.scrollview[0] += shift[0]
        self.scrollview[1] += shift[1]

        print(f"sview {self.scrollview[0]},{self.scrollview[1]}")
        self.show_image()  # redraw the image
    #-------------------------------------------------------------

    def nucmask__init__(self):
        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                self.f1dapichan = index
                self.F1_DAPI_dir = f"C{self.f1dapichan}_DAPI"

        self.arr_nucleus_contours = None    
        

        #Get Zslice from C0 of F1 for 
        self.slice_ismask = self.get_masking_slice()
        self.mip_tomask = self.get_MIP_tomask()
        self.generate_nucleus_mask(self.slice_ismask, self.mip_tomask)

    def get_MIP_tomask(self):

        #self.mip_tomask_dir = os.path.join(self.processing_directory_folder,"FILE_1",self.p_MIP_directory,self.chanidstring)
        self.mip_tomask_dir = os.path.join(self.processing_directory_folder,"FILE_1","RAW_MIP",self.F1_DAPI_dir)

        mip_tomask_list = os.listdir(self.mip_tomask_dir)
        mip_tomask = os.path.join(self.mip_tomask_dir,mip_tomask_list[0])

        return mip_tomask
    def get_masking_slice(self):
        #Get folder for zslice mask, get number of slices, pick a slice in the middle
        self.nucleus_channel_zslice_path = os.path.join(self.processing_directory_folder,"FILE_1",self.p_zslices_directory,self.F1_DAPI_dir)
        sliceslist = os.listdir(self.nucleus_channel_zslice_path)
        #slice_formask = sliceslist[round((len(sliceslist))/2)]

        entries = os.listdir(self.nucleus_channel_zslice_path)
        # Filter out directories and count the number of files
        num_files = len([entry for entry in entries if os.path.isfile(os.path.join(self.nucleus_channel_zslice_path, entry))])
        half_num_files = round(num_files/2)
        
        slice_formask =f"F1_{self.F1_DAPI_dir}_z{half_num_files}_.tif"
        slice_formask_path = os.path.join(self.nucleus_channel_zslice_path,slice_formask)
        print(f"Slice used to mask: {slice_formask_path}")
        return slice_formask_path
    def generate_nucleus_mask(self,dapi_nucleus_mask_input,dna_mip_input):

        dapi_nucleus_base_img_in = cv2.imread(dapi_nucleus_mask_input, cv2.IMREAD_UNCHANGED)
        dapi_nucleus_base_img = dapi_nucleus_base_img_in * 255

        cv2.imwrite(os.path.join(self.processing_directory_folder, f"dapinucleusbaseimg.tif"), dapi_nucleus_base_img_in )
        
        dapi_nucleus_base_img_8bit = cv2.convertScaleAbs(dapi_nucleus_base_img_in, alpha=(255.0/65535.0))

        cv2.imwrite(os.path.join(self.processing_directory_folder, f"dapinucleusbaseimg2.tif"), dapi_nucleus_base_img_8bit )
    

        dapi_nucleus_base_img = dapi_nucleus_base_img_8bit

        dna_mip_img = cv2.imread(dna_mip_input, cv2.IMREAD_GRAYSCALE)
        gblur_img = cv2.GaussianBlur(dapi_nucleus_base_img,(35,35),0)
        ret3, nucleus_threshold_img = cv2.threshold(gblur_img,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)

        # Convert nucleus_threshold_mask to uint8 if necessary
        nucleus_threshold_mask = nucleus_threshold_img.astype(np.uint8)

        # Ensure the size of the mask matches the size of the source image
        assert dna_mip_img.shape[:2] == nucleus_threshold_mask.shape[:2], "Image and mask must have the same size"

        # Apply the mask
        masked_dna_img = cv2.bitwise_and(dna_mip_img, dna_mip_img, mask=nucleus_threshold_mask)


        output_path = os.path.join(self.processing_directory_folder, f"masked_dna_mip.tif")    
        cv2.imwrite(output_path,masked_dna_img)


        output_path = os.path.join(self.processing_directory_folder, f"nucleus_threshold_img.tif")    
        cv2.imwrite(output_path,nucleus_threshold_img)

        #NEW -DILATE 
        kernel = np.ones((10, 10), np.uint8)  # Define a kernel for dilation; adjust size for more or less dilation
        dilated_mask = cv2.dilate(nucleus_threshold_img, kernel, iterations=1)  # Dilation
        nucleus_threshold_img = dilated_mask.copy()
        #----count cells--------

        edged_img = cv2.Canny(nucleus_threshold_img, 30,150)

        dna_mip_base = cv2.imread(dna_mip_input, cv2.COLOR_BGR2GRAY)
        contours, hierarchy = cv2.findContours(nucleus_threshold_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        contoured_dna_img = dna_mip_base.copy()
        cv2.drawContours(contoured_dna_img, contours, -1, (0,255,0), 2, cv2.LINE_AA)
        print(list(contours))
        print("CONTOURS LIST)")

        index = 1
        isolated_count = 0
        cluster_count = 0
        #contoured_dna_img_1 = dna_mip_base.copy()
        #contoured_dna_img_2 = cv2.cvtColor(contoured_dna_img_1, cv2.COLOR_GRAY2BGR) * 255
        contoured_dna_img_1 = cv2.imread(dna_mip_input, cv2.IMREAD_GRAYSCALE)
        contoured_dna_img_2 = cv2.cvtColor(contoured_dna_img_1, cv2.COLOR_GRAY2BGR) *255
        #np -> [index, i , x, y]
        arr_nucleus_contours = np.empty([1,4])
        #print(arr_nucleus_contours)

        for cntr in contours:
            area = cv2.contourArea(cntr)
            convex_hull = cv2.convexHull(cntr)
            convex_hull_area = cv2.contourArea(convex_hull)
            if convex_hull_area > 0:
                ratio = area / convex_hull_area
            else: ratio = 0

            
            print(index, area, convex_hull_area, ratio)
            x,y,w,h = cv2.boundingRect(cntr)
            cv2.putText(contoured_dna_img_2, str(index), (x,y), cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (0,0,255), 2)
            if ratio < 0.91:
                # cluster contours in red
                cv2.drawContours(contoured_dna_img_2, [cntr], 0, (0,0,255), 2)
                cluster_count = cluster_count + 1
            else:
                # isolated contours in green
                cv2.drawContours(contoured_dna_img_2, [cntr], 0, (0,0,255), 2)
                isolated_count = isolated_count + 1

            epsilon = 0.001 * cv2.arcLength(cntr, True)
            approx = cv2.approxPolyDP(cntr, epsilon, True)
            n = approx.ravel()  
            i = 0
            font = cv2.FONT_HERSHEY_COMPLEX 
            for j in n : 
                if(i % 2 == 0): 
                    x = n[i] 
                    y = n[i + 1] 
        
                    # String containing the co-ordinates. 
                    string = str(x) + " " + str(y)

                    #Generate contours cooridinates array to pass into tkinter visualization
                    arr_nucleus_contours = np.vstack((arr_nucleus_contours , [index,i,x,y]))
        
                    if(i == 0): 
                        # text on topmost co-ordinate. 
                        cv2.putText(contoured_dna_img_2, "Arrow tip", (x, y), 
                                        font, 0.5, (255, 0, 0))  
                    else: 
                        # text on remaining co-ordinates. 
                        cv2.putText(contoured_dna_img_2, string, (x, y),  
                                font, 0.5, (0, 255, 0))  
                    #print(f"index: {index} i: {i} x: {x} y: {y}")

                i = i + 1
            index = index + 1
        #print('number_clusters:',cluster_count)
        #print('number_isolated:',isolated_count)


        arr_nucleus_contours = np.delete(arr_nucleus_contours,(0),axis=0)

        #Pass to DynData Class
        self.arr_nucleus_contours = arr_nucleus_contours
        dyn_data.arr_nucleus_contours = self.arr_nucleus_contours   

        print(arr_nucleus_contours)
        print(len(arr_nucleus_contours))

        #cell_pick_gui_main(arr_nucleus_contours)
        #---------------
        #cv2.imshow('Base Image', dapi_nucleus_base_img)
        #cv2.imshow('Gaussian Blur Image', gblur_img)
        #cv2.imshow('Threshold Nucleus Image', nucleus_threshold_img)
        #cv2.imshow('Masked DNA Image', masked_dna_img)
        #cv2.imshow('Contoured Image', contoured_dna_img)
        #cv2.imshow('Contoured2 Image',contoured_dna_img_2)
        #cv2.waitKey(0)
        #cv2.destroyAllWindows
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class Picking_MAIN(tk.Frame):
    def __init__(self, mainframe):  # Ensure it takes an argument
        self.root = mainapp.newWindow
        print("INTIALIZE PICKING")
       
        ###------------Pull info to initialize-------------###
        #Get Nucleus and Puncta Arrays DynData
        self.kp_input = dyn_data.keypoints_array
        #Get F2 Offset
        self.f2_offset = dyn_data.f_offset
        #Get Polygon Coords
        self.polygons = dyn_data.nucpolygons #[nucIndex,polyCoords] from Nuc Picking
        #NAMETAGS
        self.f1nametag = dyn_data.f1nametag
        self.f2nametag = dyn_data.f2nametag
        #----------------------------------------------------#     

        self.exportimg = None
        ###-------------Initialization arrays--------------###
        self.polygons_id = []    #[nucInd,polyID,polyTextID]
        self.kp_polygons_arr = [] #[kpID,kp_x,kp_y,nucInd,polyID,polyTextID]

        #Dynamic Info Array
        self.kp_poly_dyn_arr = [] #[kpID,kpOvalID,nucInd,polyID,polyTextID]

        #ARROW ARRAY
        self.arrows_arr = [] #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
        #----------------------------------------------------# 
        #Determine #Channels
        self.f1channels = dyn_data.f1channels
        self.f2channels = dyn_data.f2channels
        #----------------------------------------------------# 
        #ROI ARRAYS that associate ROI with nucleus and GUI ovals/ polygons
        #FILE_1-------------------
        self.kp_poly_F1C0_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F1C1_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F1C2_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F1C3_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F1C4_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F1C5_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        #FILE_2-------------------
        self.kp_poly_F2C0_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F2C1_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F2C2_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F2C3_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F2C4_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        self.kp_poly_F2C5_arr = []   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        #----------------------------------------------------# 
        #LOAD IMAGES
        self.define_img_in_filepaths()
        self.load_input_images()
        #----------------------------------------------------#
        #Initialize MIP and ZSLICE GUIS

        #----------------------------------------------------#
        #CAPTURE IMAGE VARIABLES
        self.capimg_count = 0               #counter for saving image name
        #----------------------------------------------------#
        #Initialize Frames
        tk.Frame.__init__(self, master=mainframe)
        self.master.title('ZFISHER --- ROI Picking')
        mainframe.geometry("1024x1022")
        self.initialize_ROIwindow()

        #----------------------------------------------------#
        #Initialize CONTROL PANEL
        self.initialize_CONTROLwindow()  




        self.canvas.event_generate('<Configure>')
    

    def initialize_ROIwindow(self):

        self.highlighted_keypoints = set()      ###### NOT SURE IF NECESSARY#####

        # Vertical and horizontal scrollbars for canvas
        vbar = AutoScrollbar(self.master, orient='vertical')
        hbar = AutoScrollbar(self.master, orient='horizontal')
        vbar.grid(row=0, column=1, rowspan=2, sticky='ns')
        hbar.grid(row=2, column=0, sticky='we')

        # Create canvas and put image on it
        self.canvas = tk.Canvas(self.master, highlightthickness=0,
                                xscrollcommand=hbar.set, yscrollcommand=vbar.set)
        self.canvas.grid(row=0, column=0, sticky='nswe')
        self.canvas.update()  # wait till canvas is created
        vbar.configure(command=self.scroll_y)  # bind scrollbars to the canvas
        hbar.configure(command=self.scroll_x)

        #Set F2 OPAC
        self.f2opac = 0.7
        # Make the canvas expandable
        self.master.rowconfigure(0, weight=1)
        self.master.columnconfigure(0, weight=1)
        self.scalebar_visible = True

        self.image = Image.open(self.input_img_path)  # open image
        self.modimage = self.image.copy()
        # Put image into container rectangle and use it to set proper coordinates to the image
        self.width, self.height = self.image.size
        #print(f"self.width height {self.width,self.height}")
        self.imscale = 1.0  # scale for the canvaas image
        self.scalefactor = 1.0
        self.delta = 1.3  # zoom magnitude
        

        
        self.container = self.canvas.create_rectangle(0, 0, self.width, self.height, width=0)


        self.padx = int(self.f2_offset[0])
        self.pady = int(self.f2_offset[1])

        print(f"selfcon {self.canvas.coords(self.container)}")
        print(f"padx {self.padx} pady {self.pady}")
        self.containerpadded = self.canvas.create_rectangle(0, 0, self.width+abs(math.ceil((self.padx*2)/2)*2), self.height+abs(math.ceil((self.pady*2)/2)*2), width=0)
        print(f"selfconpadded {self.canvas.coords(self.containerpadded)}")
        bbox_padded = self.canvas.bbox(self.containerpadded)
        width_padcont= int(bbox_padded[2] - bbox_padded[0])
        height_padcont =  int(bbox_padded[3] - bbox_padded[1])
        self.padimgwidth = width_padcont
        self.padimgheight = height_padcont
        print(f"{self.padimgwidth},{self.padimgheight}")

    def load_input_images(self):
        #BASE REFERENCE IMAGE TO DEFINE GUI SIZES
        self.input_img_path = os.path.join(self.processing_directory_folder,"masked_dna_mip.tif")
        #RAW MIP FILES
        self.load_input_MIPs()
        self.load_input_zslices()
   
    def load_input_zslices(self):
        self.f1_zstack = dyn_data.f1_Zstack_cropped
        self.f2_zstack = dyn_data.f2_Zstack_cropped
    def load_input_MIPs(self):
        #FILE 1 RAW MIPS
        self.F1_C0_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_raw_directory),self.f1channels,0)
        self.F1_C1_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_raw_directory),self.f1channels,1)
        self.F1_C2_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_raw_directory),self.f1channels,2)
        self.F1_C3_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_raw_directory),self.f1channels,3)
        self.F1_C4_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_raw_directory),self.f1channels,4)
        self.F1_C5_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F1,self.p_MIP_raw_directory),self.f1channels,5)
        
        #FILE 2 RAW MIPS
        self.F2_C0_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_raw_directory),self.f2channels,0)
        self.F2_C1_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_raw_directory),self.f2channels,1)
        self.F2_C2_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_raw_directory),self.f2channels,2)
        self.F2_C3_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_raw_directory),self.f2channels,3)
        self.F2_C4_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_raw_directory),self.f2channels,4)
        self.F2_C5_MIP = self.get_MIP_file(os.path.join(self.processing_directory_folder,self.inputs_directory_F2,self.p_MIP_raw_directory),self.f2channels,5)   
    def get_MIP_file(self,target_dir,channels,channum):
        print(f"get mip channel length = {len(channels)} --- {channels} ---- {channum}")
        if channum >= len(channels): return None

        filepath = os.path.join(target_dir,f"C{channum}_{channels[channum]}")

        print(filepath)
        target_dir = filepath

        files = os.listdir(target_dir)
        if len(files) == 0: return None
        else:
            image_name= os.listdir(target_dir)[0]
            image_path = os.path.join(target_dir,image_name)
            print(image_path)
            return cv2.imread(image_path, cv2.IMREAD_UNCHANGED) 
    
    
    def define_img_in_filepaths(self):
        #MASKED MIPS for Channel toggle
        self.inputs_directory_F1 = "FILE_1"
        self.inputs_directory_F2 = "FILE_2"
        self.processing_directory_folder = "Processing"
        self.p_MIP_directory = "MIP"
        self.p_MIP_raw_directory = "RAW_MIP"
        self.p_MIP_masked_directory = "Masked"
        self.p_zslices_directory = "zslices"
        self.p_C0_directory = "C0"
        self.p_C1_directory = "C1"
        self.p_C2_directory = "C2"
        self.p_C3_directory = "C3"    
        self.p_C4_directory = "C4"
        self.p_C5_directory = "C5"           

    def initialize_CONTROLwindow(self):
        self.control_window = tk.Toplevel(self.master)
        self.control_window.title("Control Panel - ROI Picking")      
        self.finish_button = tk.Button(self.control_window, text="!!!--Finalize ROIs--!!!", command=self.finalize_kppicking)
        self.finish_button.grid(row=500, column=0, columnspan=3)    
        #---------------------------------------------------------------------------
        #CHANNEL TOGGLES reg

        #CHANNEL FRAME
        self.chanframe = tk.Frame(self.control_window)
        self.chanframe.grid(row=3, column=0)

        #TOGGLES
        self.F1_C0_toggle = tk.BooleanVar(value=True)
        self.F1_C1_toggle = tk.BooleanVar(value=True)
        self.F1_C2_toggle = tk.BooleanVar(value=True)
        self.F1_C3_toggle = tk.BooleanVar(value=True)
        self.F1_C4_toggle = tk.BooleanVar(value=True)
        self.F1_C5_toggle = tk.BooleanVar(value=True)


        self.F2_C0_toggle = tk.BooleanVar(value=True)
        self.F2_C1_toggle = tk.BooleanVar(value=True)
        self.F2_C2_toggle = tk.BooleanVar(value=True)
        self.F2_C3_toggle = tk.BooleanVar(value=True)
        self.F2_C4_toggle = tk.BooleanVar(value=True)
        self.F2_C5_toggle = tk.BooleanVar(value=True)     

        #Color Picker----------------------------------------------
        self.chancolorarr = [
            "#FF0000",  # Red   F1 - 0
            "#00FF00",  # Green F1 - 1
            "#F3F3F3",  # Purple F1 - 2
            "#FFFF00",  # Yellow    F1 - 3
            "#FF00FF",  # Magenta   F1 - 4
            "#00FFFF",  # Cyan      F1 - 5

            "#FFA500",  # Orange    F2 - 0
            "#800080",  # Purple    F2 - 1
            "#008000",  # Dark Green    F2 - 2
            "#FF4500",  # Orange Red    F2 - 3
            "#980000",  # Orange   F2 - 4
            "#8B4513"   # Saddle Brown  F2 - 5
        ]

        for index, chan in enumerate(dyn_data.f1channels):
            if chan == "DAPI":
                self.chancolorarr[index] = "#0000FF"
        for index, chan in enumerate(dyn_data.f2channels):
            if chan == "DAPI":
                self.chancolorarr[index+6] = "#0000FF"

        #--------------------------------------------------------
        #Toggles if we should analyze
        self.F1_C0_out_toggle = tk.BooleanVar(value=True)
        self.F1_C1_out_toggle = tk.BooleanVar(value=True)
        self.F1_C2_out_toggle = tk.BooleanVar(value=True)
        self.F1_C3_out_toggle = tk.BooleanVar(value=True)
        self.F1_C4_out_toggle = tk.BooleanVar(value=True)
        self.F1_C5_out_toggle = tk.BooleanVar(value=True)    

        self.F2_C0_out_toggle = tk.BooleanVar(value=True)
        self.F2_C1_out_toggle = tk.BooleanVar(value=True)
        self.F2_C2_out_toggle = tk.BooleanVar(value=True)
        self.F2_C3_out_toggle = tk.BooleanVar(value=True)
        self.F2_C4_out_toggle = tk.BooleanVar(value=True)
        self.F2_C5_out_toggle = tk.BooleanVar(value=True)
        #CHANNEL KP RB Selectors
        self.KP_rb_selection = tk.StringVar()
        #CHANNEL KP PICK TOGGLES
        self.F1_C0_toggle_KP = tk.BooleanVar(value=True)
        self.F1_C1_toggle_KP = tk.BooleanVar(value=True)
        self.F1_C2_toggle_KP = tk.BooleanVar(value=True)
        self.F1_C3_toggle_KP = tk.BooleanVar(value=True)
        self.F1_C4_toggle_KP = tk.BooleanVar(value=True)
        self.F1_C5_toggle_KP = tk.BooleanVar(value=True)

        self.F2_C0_toggle_KP = tk.BooleanVar(value=True)
        self.F2_C1_toggle_KP = tk.BooleanVar(value=True)
        self.F2_C2_toggle_KP = tk.BooleanVar(value=True)
        self.F2_C3_toggle_KP = tk.BooleanVar(value=True)
        self.F2_C4_toggle_KP = tk.BooleanVar(value=True)
        self.F2_C5_toggle_KP = tk.BooleanVar(value=True)

        #BRIGHTNESS AND CONTRAST FRAME
        # Create a Frame to hold the sliders
        self.bc_frame_maker()
        self.chan_kp_autopicker_framemaker()

        #BINDINGS
        self.canvas.bind('<Configure>', self.toggle_image_channels)  # canvas is resized 
        self.canvas.bind("<Motion>", self.motion)
       # self.canvas.bind("<Button-1>", self.kp_add)
        self.canvas.bind("<Button-1>", self.left_mouseclick_wrapper)

        #KP_REMOVE
        #self.canvas.bind("<Button-2>", self.kp_remove)
        if platform.system() == "Darwin":  # macOS
            self.canvas.bind('<Button-2>', self.right_mouseclick_wrapper) # right-click for macOS
        elif platform.system() == "Linux":  # Linux
            self.canvas.bind('<Button-3>', self.right_mouseclick_wrapper) # right-click for Linux

        #ARROW
        self.arrow_sp = None
        self.arrow_ep = None



        #DRAG
        #self.canvas.bind('<ButtonPress-3>', self.move_from)
        #self.canvas.bind('<B3-Motion>',     self.move_to)
        if platform.system() == "Darwin":  # macOS
             self.canvas.bind('<ButtonPress-3>', self.move_from)
             self.canvas.bind('<B3-Motion>',     self.move_to)
        elif platform.system() == "Linux":  # Linux
             self.canvas.bind('<ButtonPress-2>', self.move_from)
             self.canvas.bind('<B2-Motion>',     self.move_to)

        #ZOOM    
        #self.canvas.bind("<MouseWheel>", self.zoom)
        if platform.system() == "Darwin":  # macOS
            self.canvas.bind("<MouseWheel>", self.zoom)
        elif platform.system() == "Linux":  # Linux
            self.canvas.bind("<Button-4>", self.zoom)
            self.canvas.bind("<Button-5>", self.zoom)      


        self.miscframe = tk.Frame(self.control_window)
        self.miscframe.grid(row=5,column=1)

        self.scalebar_toggle_var = tk.BooleanVar(value=True)
        self.scalebar_toggle_checkbox = tk.Checkbutton(self.miscframe,text="Show Scalebar", variable=self.scalebar_toggle_var, command=self.toggle_scalebar).grid(row=2, column=1, columnspan=1)

        self.polygon_toggle_var = tk.BooleanVar(value=True)
        self.polygon_toggle_checkbox = tk.Checkbutton(self.miscframe,text="Show Nuclei", variable=self.polygon_toggle_var, command=self.polygon_toggle).grid(row=2, column=2, columnspan=1)

        self.arrow_toggle_var = tk.BooleanVar(value=False)
        self.arrow_toggle_checkbox = tk.Checkbutton(self.miscframe, text="Draw Arrows", variable=self.arrow_toggle_var, command=None).grid(row=4, column=1, columnspan=2)

        #CAPTURE IMAGE
        self.imgcap_button = tk.Button(self.miscframe, text="IMG CAPTURE WINDOW", command=lambda: self.capture_image())
        self.imgcap_button.grid(row=5,column=1, columnspan=2)

        #Mouse position label 
        self.mousepos_label = tk.Label(self.miscframe, text="Current Mouse Position: (0000.00,0000.00)", fg="white", font=("Courier"))
        self.mousepos_label.grid(row=3, column=0, columnspan=3)  # Use grid instead of pack
        x_lab = 0
        y_lab = 0
        self.mousepos_label.config(text=f"Mouse Position: ({float(x_lab):07.2f}, {float(y_lab):07.2f})")

        print(f"init2 F1C1G {self.F1_C1_MIP.size}")

        self.toggle_image_channels()

        #print(f"init3 F1C1G {self.F1_C1G_MIP.shape[0]},{self.F1_C1G_MIP[1]}")

        self.draw_scalebar()
        self.scalebar_visible = True

        self.draw_input_polygons()



    #-------------------------------------------------------------------------------
    #CAPTURE IMAGE OF WINDOW
    #-------------------------------------------------------------------------------
    def capture_image(self):

        self.capimg_count += 1 
        #THIS CAPTURES WITH DRAWN WIDGETS -------
        # Get the coordinates of the canvas
        x = self.canvas.winfo_rootx()
        y = self.canvas.winfo_rooty()
        x1 = x + self.canvas.winfo_width()
        y1 = y + self.canvas.winfo_height()

        #self.master.attributes("-topmost", True)
        self.control_window.lower(self.master)
        self.control_window.update() 
        # Capture the canvas content using ImageGrab
        img = ImageGrab.grab((x, y, x1, y1))

        # Define the export directory and output path
        export_dir = "Outputs/IMAGES"
        output_name = f"{self.capimg_count}_imgcapture_widget.tiff"
        output_path = os.path.join(export_dir, output_name)

        # Save the image as PNG
        img.save(output_path, 'TIFF')

        print(f"Image saved at: {output_path}")

        #SAVE the NONWIDGET VERSION ------
        filename = f"{self.capimg_count}_imgcapture_NOwidget.tiff"
        self.exportimg._PhotoImage__photo.write(f"Outputs/IMAGES/{filename}.tiff")
        return
       
        print("Captured Window Image")

        
        return
        # Get the coordinates of the canvas
        x = self.canvas.winfo_rootx()
        y = self.canvas.winfo_rooty()
        x1 = x + self.canvas.winfo_width()
        y1 = y + self.canvas.winfo_height()

        # Capture the canvas content using ImageGrab
        img = ImageGrab.grab((x, y, x1, y1))

        # Define the export directory and output path
        export_dir = "Outputs"
        output_name = "captured_canvas.png"
        output_path = os.path.join(export_dir, output_name)

        # Save the image as PNG
        img.save(output_path, 'PNG')

        print(f"Image saved at: {output_path}")
        return
        output_image = self.exportimg
        export_dir = "Outputs"
        output_name = "test1.tiff"
        output_path = os.path.join(export_dir, output_name)
        cv2.imwrite(output_path, output_image)
    #-------------------------------------------------------------------------------
    #INPUT FUNCTION WRAPPERS
    #-------------------------------------------------------------------------------
    def left_mouseclick_wrapper(self,event):
        if not self.arrow_toggle_var.get(): self.kp_add(event)
        if self.arrow_toggle_var.get(): self.make_arrow(event)
    def right_mouseclick_wrapper(self,event):
        if not self.arrow_toggle_var.get(): self.kp_remove(event)
        if self.arrow_toggle_var.get(): self.remove_arrow(event)
    #-------------------------------------------------------------------------------
    #-----KEYPOINTS-INITIALIZATION--------------------------
    def remove_edge_kps(self):
        #INPUTS-------------------------------------------------------
        ###self.kp_poly_dyn_arr = #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        #OUTPUTS------------------------------------------------------
        # assign to final dynamic array #self.kp_poly_dyn_arr = #[kpID,kpOvalID,nucInd,polyID,polyTextID]

        ovalIDs_to_remove = []
        polygons = []

        for rowinfo in self.kp_poly_dyn_arr:
            _,kpOvalID,_,_,_ = rowinfo
            oval_coords = self.canvas.coords(kpOvalID)
            oval_x1, oval_y1, oval_x2, oval_y2 = oval_coords

            for row_poly_info in self.kp_poly_dyn_arr:
                    # Ensure row_poly_info is a list or a tuple
                    _,_,_,polyID,_ = row_poly_info
                    poly_coords = self.canvas.coords(polyID)
                    num_coords = len(poly_coords)
                    for i in range(0, num_coords, 2):
                        x1, y1 = poly_coords[i], poly_coords[i + 1]
                        x2, y2 = poly_coords[(i + 2) % num_coords], poly_coords[(i + 3) % num_coords]

                        # Check if the oval bounding box intersects with the line segment
                        if (oval_x1 < max(x1, x2) and oval_x2 > min(x1, x2) and
                            oval_y1 < max(y1, y2) and oval_y2 > min(y1, y2)):
                            # Oval is touching the polygon edge, remove it
                            ovalIDs_to_remove.append(kpOvalID)
                            break

        # Remove oval IDs that need to be removed
        print(f"# edge ovalIDs to remove: {len(ovalIDs_to_remove)}______ovalIDs_to_remove")
        print(f"# unique edge ovalIDS to remove: {len(set(ovalIDs_to_remove))}______set(ovalIDs_to_remove)")
        for ovalID in ovalIDs_to_remove:
            for row in self.kp_poly_dyn_arr:    #[kpID,kpOvalID,nucInd,polyID,polyTextID]!Pickler73


                if row[1] == ovalID:
                    self.kp_poly_dyn_arr.remove(row)
                    self.canvas.delete(ovalID)

        #self.kp_poly_dyn_arr = [row for row in self.kp_poly_dyn_arr if row[1] not in ovalIDs_to_remove]

        print(f"# kp remaining after removal: {len(self.kp_poly_dyn_arr)}______kp_poly_dyn_arr ")
        
        print(self.kp_poly_dyn_arr[0])
    def draw_input_keypoints(self):
        #INPUTS-------------------------------------------------------
        #self.kp_polygons_arr  #[kpID,kp_x,kp_y,NucID,PolygonID,PolygonTextID]
        #OUTPUTS------------------------------------------------------
        # assign to final dynamic array #self.kp_poly_dyn_arr = #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        for inforow in self.kp_polygons_arr:
                kpID, kp_x, kp_y, nucInd, polyID, polyTextID = inforow
                # Draw a circle for each keypoint on the canvas and hide it
                oval_id = self.canvas.create_oval(kp_x - 5, kp_y - 5, kp_x + 5, kp_y + 5, outline='yellow', fill="", state="normal", width="0.5")
                self.kp_poly_dyn_arr.append([kpID, oval_id, nucInd, polyID, polyTextID])
        print(f"# ovalIDs assigned to kp: {len(self.kp_poly_dyn_arr)}______self.kp_poly_dyn_arr ")
    def assign_input_keypoints_to_polygons(self):
        #INPUTS-------------------------------------------------------
        #self.polygons_id    #[NucID,PolygonID,PolygonTextID]
        #OUTPUTS------------------------------------------------------
        #self.kp_polygons_arr  #[kpID,kp_x,kp_y,NucID,PolygonID,PolygonTextID]
        print(f"# initialization kps: {len(self.kp_input)}______self.kp_input")
        kp_id = 0  # assign kpID to all input KPs
        for row in self.kp_input:
            kp_id += 1  # [x,y]
            kp_x, kp_y = row[0], row[1]
            # Check if the keypoint lies inside any polygon
            for polyinfo in self.polygons_id:          #[nucInd,polyID,polyTextID]
                    nucInd, polyID, polyTextID = polyinfo
                    polygon_coords = self.canvas.coords(polyID)
                    crossings = 0
                    for i in range(len(polygon_coords) // 2):
                        x1, y1 = polygon_coords[i * 2], polygon_coords[i * 2 + 1]
                        x2, y2 = polygon_coords[(i * 2 + 2) % len(polygon_coords)], polygon_coords[
                            (i * 2 + 3) % len(polygon_coords)]

                        # Check if the ray intersects with the edge
                        if ((y1 > kp_y) != (y2 > kp_y)) and (kp_x < (x2 - x1) * (kp_y - y1) / (y2 - y1) + x1):
                            crossings += 1

                    # If the number of crossings is odd, the point is inside the polygon
                    if crossings % 2 == 1:
                        found_polygon = polyID
                        temparray = [kp_id,kp_x,kp_y,nucInd,found_polygon,polyTextID]#[kpID,kp_x,kp_y,NucID,PolygonID,PolygonTextID]
                        self.kp_polygons_arr.append(temparray)
                        break
        print(f"# kps associated with polygons: {len(self.kp_input)}______self.kp_polygons_arr")
    #-----POLYGONS-INITIALIZATION---------------------------
    def draw_input_polygons(self):
        for nucindex, polygoncoords in self.polygons:
            polygon_id= self.canvas.create_polygon(polygoncoords, outline="red", fill="", tags='polygon')
            self.canvas.move(polygon_id,abs(self.padx),abs(self.pady))
            polygon_textid = self.display_index_on_polygon(polygon_id, int(nucindex)) 
            temparray = [nucindex, polygon_id, polygon_textid]
            self.polygons_id.append(temparray)
    def display_index_on_polygon(self, polygon_id, index):
        coords = self.canvas.coords(polygon_id)
        x_coords = [int(coords[i]) for i in range(0, len(coords), 2)]
        y_coords = [int(coords[i]) for i in range(1, len(coords), 2)]
        
        # Check if either x_coords or y_coords is empty before performing division
        if len(x_coords) == 0 or len(y_coords) == 0:
            # Handle the case where one of the coordinate lists is empty
            print("Error: Empty coordinate list.")
            return
        
        x_center = sum(x_coords) // len(x_coords)  # Using integer division
        y_center = sum(y_coords) // len(y_coords)  # Using integer division
        text_item = self.canvas.create_text(x_center, y_center, text=str(index), font=('Arial', 12), fill='red', tags='polygontext')
        return text_item
    def polygon_toggle(self):
        polygon_items = self.canvas.find_withtag('polygon')
        polygon_text_items = self.canvas.find_withtag('polygontext')
        if self.polygon_toggle_var.get():
            for p in polygon_items:
                self.canvas.itemconfigure(p, state="normal")
            for pt in polygon_text_items:
                self.canvas.itemconfigure(pt, state="normal")
        else: 
            for p in polygon_items:
                self.canvas.itemconfigure(p, state="hidden")
            for pt in polygon_text_items:
                self.canvas.itemconfigure(pt, state="hidden")
    #-------------------------------------------------------
    #---------KP PICKING------------
    def kp_autopick(self,channel,radius):

        chan_file_dict = {
            0: self.F1_C0B_MIP_path,
            1: self.F1_C1G_MIP_path,
            2: self.F1_C2R_MIP_path,
            3: self.F1_C3M_MIP_path,
        }

        chan_arr_dict = {
            0: self.kp_poly_C0B_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            1: self.kp_poly_C1G_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            2: self.kp_poly_C2R_arr,  #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            3: self.kp_poly_C3M_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
        }

        chan_oval_color_dict = {
            0:  "orange",       #Blue 
            1:  "red",          #Green
            2:  "green",        #red
            3:  "yellow",       #magenta
        }

        #--------------------------
        #BLOB DETECTOR TO GET KP x,y
        file = chan_file_dict[channel]
        print(file)
        print(radius)
        new_kps = self.blobdetector.detect_chan_blobs(file,radius)
        print(new_kps)
        print(len(new_kps))
        #---------------------------
        kp_id = 0  # assign kpID to all input KPs
        for row in self.kp_input:
            kp_id += 1  # [x,y]
            kp_x, kp_y = row[0], row[1]
            newoval = self.canvas.create_oval(kp_x - 5, kp_y - 5, kp_x + 5, kp_y + 5, outline="white", fill="", state="normal", )
            # Check if the keypoint lies inside any polygon
            for polyinfo in self.polygons_id:          #[nucInd,polyID,polyTextID]
                    nucInd, polyID, polyTextID = polyinfo
                    polygon_coords = self.canvas.coords(polyID)
                    crossings = 0
                    for i in range(len(polygon_coords) // 2):
                        x1, y1 = polygon_coords[i * 2], polygon_coords[i * 2 + 1]
                        x2, y2 = polygon_coords[(i * 2 + 2) % len(polygon_coords)], polygon_coords[
                            (i * 2 + 3) % len(polygon_coords)]

                        # Check if the ray intersects with the edge
                        if ((y1 > kp_y) != (y2 > kp_y)) and (kp_x < (x2 - x1) * (kp_y - y1) / (y2 - y1) + x1):
                            crossings += 1


                    # If the number of crossings is odd, the point is inside the polygon
                    if crossings % 2 == 1:
                        found_polygon = polyID
                        oval_id = self.canvas.create_oval(kp_x - 5, kp_y - 5, kp_x + 5, kp_y + 5, outline=chan_oval_color_dict[channel], fill="", state="normal", )
                        
                        kp_id_found = False
                        for row in chan_arr_dict[channel]:
                            if row[0] == kp_id:
                                kp_id_found = True
                                break
                        if not kp_id_found:
                            # If kp_id is not found in any row, set it to kp_id
                            kp_id = kp_id
                        else:
                            # If kp_id is found in any row, set it to length of array + 1
                            kp_id = len(chan_arr_dict[channel]) + 1
                                
                        temparray = [kp_id,oval_id,nucInd,found_polygon,polyTextID]#[kpID,ovalID,nucInd, polyID, polyTextID]
                        
                        
                        chan_arr_dict[channel].append(temparray)
                        break
                # Draw a circle for each keypoint on the canvas and hide it
                
                #self.kp_poly_dyn_arr.append([kpID, oval_id, nucInd, polyID, polyTextID])
        #print(f"# ovalIDs assigned to kp: {len(self.kp_poly_dyn_arr)}______self.kp_poly_dyn_arr ")
    def kp_add(self, event):
        ppm = 0.108333333333333 #px_per_micron 
        kpx = self.canvas.canvasx(event.x)
        kpy = self.canvas.canvasy(event.y)       

        print("add kp")
        rb_dict = {
            "F1_C0": 0,
            "F1_C1": 1,
            "F1_C2": 2,
            "F1_C3": 3,
            "F1_C4": 4,
            "F1_C5": 5,
            "F2_C0": 6,
            "F2_C1": 7,
            "F2_C2": 8,
            "F2_C3": 9,
            "F2_C4": 10,
            "F2_C5": 11,

        }

        chan_arr_dict = {
            0: self.kp_poly_F1C0_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            1: self.kp_poly_F1C1_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            2: self.kp_poly_F1C2_arr,  #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            3: self.kp_poly_F1C3_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            4: self.kp_poly_F1C4_arr,
            5: self.kp_poly_F1C5_arr,
            6: self.kp_poly_F2C0_arr, 
            7: self.kp_poly_F2C1_arr, 
            8: self.kp_poly_F2C2_arr, 
            9: self.kp_poly_F2C3_arr, 
            10: self.kp_poly_F2C4_arr, 
            11: self.kp_poly_F2C5_arr
        }

        chan_oval_color_dict = {
            
            0:  "orange",       #F1 - DAPI - COB
            1:  "red",          #F1- FITC - C1G
            2:  "green",        #F1 - RED - C2R
            3:  "yellow",       #F1 - FARRED - C3M
            4:  "magenta",        #F2 - DAPI - C0B
            5:  "indigo",       #F2 - FITC - C1G
            6:  "white",       #F2 - RED - C2R
            7:  "cyan",           #F2 - FARRED = C3M
            8:  "purple",
            9:  "pink",
            10: "gray",
            11: "tan",
        }

        chan_kp_rad_dict = {
            0:  lambda: self.F1_C0_ap_entry.get() if hasattr(self, 'F1_C0_ap_entry') else None,
            1:  lambda: self.F1_C1_ap_entry.get() if hasattr(self, 'F1_C1_ap_entry') else None,
            2:  lambda: self.F1_C2_ap_entry.get() if hasattr(self, 'F1_C2_ap_entry') else None,
            3:  lambda: self.F1_C3_ap_entry.get() if hasattr(self, 'F1_C3_ap_entry') else None,
            4:  lambda: self.F1_C4_ap_entry.get() if hasattr(self, 'F1_C4_ap_entry') else None,
            5:  lambda: self.F1_C5_ap_entry.get() if hasattr(self, 'F1_C5_ap_entry') else None,
            6:  lambda: self.F2_C0_ap_entry.get() if hasattr(self, 'F2_C0_ap_entry') else None,
            7:  lambda: self.F2_C1_ap_entry.get() if hasattr(self, 'F2_C1_ap_entry') else None,
            8:  lambda: self.F2_C2_ap_entry.get() if hasattr(self, 'F2_C2_ap_entry') else None,
            9:  lambda: self.F2_C3_ap_entry.get() if hasattr(self, 'F2_C3_ap_entry') else None,
            10:  lambda: self.F2_C4_ap_entry.get() if hasattr(self, 'F2_C4_ap_entry') else None,
            11:  lambda: self.F2_C5_ap_entry.get() if hasattr(self, 'F2_C5_ap_entry') else None,
        }


        print("add")
        selectedchannel = self.KP_rb_selection.get()
        selchankey = rb_dict[selectedchannel]
        print(f"{selectedchannel} --> KEY: {rb_dict[selectedchannel]}")

        print(chan_kp_rad_dict[selchankey]())
        kpradentry = float(chan_kp_rad_dict[selchankey]())
        kpradius = kpradentry / ppm * self.imscale


        OvalID = self.canvas.create_oval(kpx - kpradius, kpy - kpradius, kpx + kpradius, kpy + kpradius, outline=chan_oval_color_dict[selchankey], fill="", state="normal", tags='kp')


       # kparray = chan_arr_dict[selchankey]

        #Pass KP and KP Channel array to assign polygon and put in array

        self.kp_assign_polygon(OvalID,selchankey)

        #self.canvas.bind("<Button-2>", lambda event: self.kp_remove(event,OvalID))
    def kp_assign_polygon(self,kp,selchankey):
        #INPUTS-------------------------------------------------------
        #self.polygons_id    #[NucID,PolygonID,PolygonTextID]
        #kp -> kpid on canvas and its x,y position
        #OUTPUTS------------------------------------------------------
        #self.kp_polygons_arr  #[kpID,OvalID,kp_x,kp_y,NucID,PolygonID,PolygonTextID]
        #-------------------------------------------------------------
        #-------------------------------------------------------------
        chan_arr_dict = {
            0: self.kp_poly_F1C0_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            1: self.kp_poly_F1C1_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            2: self.kp_poly_F1C2_arr,  #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            3: self.kp_poly_F1C3_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            4: self.kp_poly_F1C4_arr, 
            5: self.kp_poly_F1C5_arr, 
            6: self.kp_poly_F2C0_arr, 
            7: self.kp_poly_F2C1_arr, 
            8: self.kp_poly_F2C2_arr, 
            9: self.kp_poly_F2C3_arr, 
            10: self.kp_poly_F2C4_arr, 
            11: self.kp_poly_F2C5_arr, 
        }

        #Determine KP ID in channel array
        kparr = chan_arr_dict[selchankey]
        kparr = sorted(kparr, key=lambda x: x[0]) #sort them in ascending order of kpID
        if len(kparr) == 0: 
            kpID = 1  #if there are no keypoints make it first KPID = 1
        else:
            missing_numbers = []
            expected_number = 1
            for inforow in kparr:   #[kpID,OvalID,kp_x,kp_y,NucID,PolygonID,PolygonTextID]
                #kpID,OvalID,kp_x,kp_y,NucID,PolyID,PolyTextID = inforow

                while inforow[0] > expected_number:
                    missing_numbers.append(expected_number)
                    expected_number += 1
                expected_number = inforow[0] + 1
            if len(missing_numbers) > 0: kpID = missing_numbers[0]
            else: kpID = len(kparr) + 1
        

        kp_x = (self.canvas.coords(kp)[0] + self.canvas.coords(kp)[2])/2
        kp_y = (self.canvas.coords(kp)[1] + self.canvas.coords(kp)[3])/2

        print(self.canvas.coords(kp))
        NucID = 1
        PolyID = 1
        PolyTextID = 1

        NucID,PolyID,PolyTextID = self.findpolygon(kp_x,kp_y)
        print(f" {NucID} - {PolyID} - {PolyTextID}")

        temparray = [kpID,kp,kp_x,kp_y,NucID,PolyID,PolyTextID]

        chan_arr_dict[selchankey].append(temparray)

        print(chan_arr_dict[selchankey])
    def findpolygon(self, x, y):
        foundNucID = 0
        foundPolyID = 0
        foundPolyTextID = 0
        for row in self.polygons_id:
            NucID, PolyID, PolyTextID = row
            poly_coords = self.canvas.coords(PolyID)
            num_coords = len(poly_coords)
            inside = False
            j = num_coords - 2
            for i in range(0, num_coords, 2):
                x_i, y_i = poly_coords[i], poly_coords[i + 1]
                x_j, y_j = poly_coords[j], poly_coords[j + 1]
                if ((y_i > y) != (y_j > y)) and (x < (x_j - x_i) * (y - y_i) / (y_j - y_i) + x_i):
                    inside = not inside
                j = i
            if inside:
                foundNucID = NucID
                foundPolyID = PolyID
                foundPolyTextID = PolyTextID
                break  # Stop searching after finding the first matching polygon
        return foundNucID, foundPolyID, foundPolyTextID

    def kp_move(self, event):
        print("move")
    def kp_remove(self, event):

        print("REMOVE ROI")
        rb_dict = {
            "F1_C0": 0,
            "F1_C1": 1,
            "F1_C2": 2,
            "F1_C3": 3,
            "F1_C4": 4,
            "F1_C5": 5,
            "F2_C0": 6,
            "F2_C1": 7,
            "F2_C2": 8,
            "F2_C3": 9,
            "F2_C4": 10,
            "F2_C5": 11
        }
        chan_arr_dict = {
            0: self.kp_poly_F1C0_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            1: self.kp_poly_F1C1_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            2: self.kp_poly_F1C2_arr,  #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            3: self.kp_poly_F1C3_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            4: self.kp_poly_F1C4_arr, 
            5: self.kp_poly_F1C5_arr, 
            6: self.kp_poly_F2C0_arr, 
            7: self.kp_poly_F2C1_arr, 
            8: self.kp_poly_F2C2_arr, 
            9: self.kp_poly_F2C3_arr, 
            10: self.kp_poly_F2C4_arr, 
            11: self.kp_poly_F2C5_arr
        }
        chan_kp_rad_dict = {
            0:  lambda: self.F1_C0_ap_entry.get() if hasattr(self, 'F1_C0_ap_entry') else None,
            1:  lambda: self.F1_C1_ap_entry.get() if hasattr(self, 'F1_C1_ap_entry') else None,
            2:  lambda: self.F1_C2_ap_entry.get() if hasattr(self, 'F1_C2_ap_entry') else None,
            3:  lambda: self.F1_C3_ap_entry.get() if hasattr(self, 'F1_C3_ap_entry') else None,
            4:  lambda: self.F1_C4_ap_entry.get() if hasattr(self, 'F1_C4_ap_entry') else None,
            5:  lambda: self.F1_C5_ap_entry.get() if hasattr(self, 'F1_C5_ap_entry') else None,
            6:  lambda: self.F2_C0_ap_entry.get() if hasattr(self, 'F2_C0_ap_entry') else None,
            7:  lambda: self.F2_C1_ap_entry.get() if hasattr(self, 'F2_C1_ap_entry') else None,
            8:  lambda: self.F2_C2_ap_entry.get() if hasattr(self, 'F2_C2_ap_entry') else None,
            9:  lambda: self.F2_C3_ap_entry.get() if hasattr(self, 'F2_C3_ap_entry') else None,
            10:  lambda: self.F2_C4_ap_entry.get() if hasattr(self, 'F2_C4_ap_entry') else None,
            11:  lambda: self.F2_C5_ap_entry.get() if hasattr(self, 'F2_C5_ap_entry') else None
        }


        print("remove")
  

        selectedchannel = self.KP_rb_selection.get()
        selchankey = rb_dict[selectedchannel]
        print(f"{selectedchannel} --> KEY: {rb_dict[selectedchannel]}")  
        print(len(chan_arr_dict[selchankey]))  

        temparray = chan_arr_dict[selchankey]
        if len(temparray) == 0: return
        for row in chan_arr_dict[selchankey]:
            print(row)
            x1, y1, x2, y2 = self.canvas.coords(row[1])
            bboxid = self.canvas.create_rectangle(x1, y1, x2, y2, outline="white", fill="white",tags='bounding_box')

            mousex = self.canvas.canvasx(event.x)
            mousey = self.canvas.canvasy(event.y) 

            overlapping_items = self.canvas.find_overlapping(mousex, mousey, mousex, mousey)
            if overlapping_items:
                filteredarray = temparray
                for item in overlapping_items:
                    tag = self.canvas.gettags(item)
                    print(tag)
                    if 'bounding_box' in tag:
                        print("Clicked on oval:", row[1])
                        self.canvas.delete(row[1])
                        filteredarray = []
                        for trow in temparray:
                            if trow[1] != row[1]:
                                filteredarray.append(trow)
                temparray=filteredarray
            self.canvas.delete(bboxid)

 

        if selchankey == 0: self.kp_poly_F1C0_arr = temparray   
        if selchankey == 1: self.kp_poly_F1C1_arr = temparray     
        if selchankey == 2: self.kp_poly_F1C2_arr = temparray    
        if selchankey == 3: self.kp_poly_F1C3_arr = temparray    
        if selchankey == 4: self.kp_poly_F1C4_arr = temparray  
        if selchankey == 5: self.kp_poly_F1C5_arr = temparray  
        if selchankey == 6: self.kp_poly_F2C0_arr = temparray  
        if selchankey == 7: self.kp_poly_F2C1_arr = temparray  
        if selchankey == 8: self.kp_poly_F2C2_arr = temparray  
        if selchankey == 9: self.kp_poly_F2C3_arr = temparray  
        if selchankey == 10: self.kp_poly_F2C4_arr = temparray   
        if selchankey == 11: self.kp_poly_F2C5_arr = temparray  

        # Find all objects overlapping with the mouse click coordinates
     #   overlapping_items = self.canvas.find_overlapping(mousex, mousey, mousex, mousey)
      #  if overlapping_items:
       #     for item in overlapping_items:
        #        tag = self.canvas.gettags(item)
         #       if 'kp' in tag:
          #          print("Clicked on oval:", item)
    def removeall_chankp(self,chan):
        rb_dict = {
            "F1_C0": 0,
            "F1_C1": 1,
            "F1_C2": 2,
            "F1_C3": 3,
            "F1_C4": 4,
            "F1_C5": 5,
            "F2_C0": 6,
            "F2_C1": 7,
            "F2_C2": 8,
            "F2_C3": 9,
            "F2_C4": 10,
            "F2_C5": 11
        }

        chan_arr_dict = {
            0: self.kp_poly_F1C0_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            1: self.kp_poly_F1C1_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            2: self.kp_poly_F1C2_arr,  #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            3: self.kp_poly_F1C3_arr,   #[kpID,kpOvalID,nucInd,polyID,polyTextID]
            4: self.kp_poly_F1C4_arr, 
            5: self.kp_poly_F1C5_arr, 
            6: self.kp_poly_F2C0_arr, 
            7: self.kp_poly_F2C1_arr, 
            8: self.kp_poly_F2C2_arr, 
            9: self.kp_poly_F2C3_arr, 
            10: self.kp_poly_F2C4_arr, 
            11: self.kp_poly_F2C5_arr
        }

        selchankey = rb_dict[chan]

        for row in chan_arr_dict[selchankey]:
            self.canvas.delete(row[1])

        if selchankey == 0: self.kp_poly_F1C0_arr = []
        if selchankey == 1: self.kp_poly_F1C1_arr = []   
        if selchankey == 2: self.kp_poly_F1C2_arr = []    
        if selchankey == 3: self.kp_poly_F1C3_arr = [] 
        if selchankey == 4: self.kp_poly_F1C4_arr = []
        if selchankey == 5: self.kp_poly_F1C5_arr = [] 
        if selchankey == 6: self.kp_poly_F2C0_arr = []  
        if selchankey == 7: self.kp_poly_F2C1_arr = []
        if selchankey == 8: self.kp_poly_F2C2_arr = []   
        if selchankey == 9: self.kp_poly_F2C3_arr = []
        if selchankey == 10: self.kp_poly_F2C4_arr = []
        if selchankey == 11: self.kp_poly_F2C5_arr = []
    
    #-----MAKE ARROWS----------------------------------------- 
    def make_arrow(self,event):
        print("ARROW!")

        eventx = self.canvas.canvasx(event.x)
        eventy = self.canvas.canvasy(event.y)   
        if self.arrow_sp is None:
            # Store the starting point (first click)
            self.arrow_sp = (eventx, eventy)
            print(f"ARROW START {self.arrow_sp}")
        else:
            # Get the ending point (second click) and draw the arrow
            self.arrow_ep = (eventx, eventy)
            
            print(f"ARROW END {self.arrow_ep}")
            # Draw an arrow from the start point to the end point
            arrowid = self.canvas.create_line(self.arrow_sp[0], self.arrow_sp[1], self.arrow_ep[0], self.arrow_ep[1], arrow=tk.LAST, width=2,tags="arrow")
            

            #Get arrowIndex
            arrowInd = None         #the arrows global arrow index
            arrowarr = self.arrows_arr      #self.arrows_arr = [] #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
            arrowarr = sorted(arrowarr, key=lambda x: x[0]) #sort them in ascending order of arrowIndex
            if len(arrowarr) == 0: arrowInd = 1     #if the arrow array is empty assign it as 1
            else:
                missing_numbers = []
                expected_number = 1
                for row in arrowarr:
                    while row[0] > expected_number:
                        missing_numbers.append(expected_number)
                        expected_number += 1
                    expected_number = row[0] + 1
                if len(missing_numbers) > 0: arrowInd = missing_numbers[0]
                else: arrowInd = len(arrowarr) + 1
            
            
            arrowtextid = self.generate_arrow_text(arrowInd)

            aX1 = self.arrow_sp[0]
            aY1 = self.arrow_sp[1]
            aX2 = self.arrow_ep[0]
            aY2 = self.arrow_ep[1]

            print(self.canvas.coords(arrowtextid)[0])
            print(self.canvas.coords(arrowtextid)[1])
            atextX = self.canvas.coords(arrowtextid)[0]
            atextY = self.canvas.coords(arrowtextid)[1]



            newarrowinforow = [arrowInd,arrowid,aX1,aY1,aX2,aY2,arrowtextid,atextX,atextY]
            self.arrows_arr.append(newarrowinforow)

            print(self.arrows_arr)
            #self.arrows_arr = [] #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
            # Reset the start point for the next arrow
            self.arrow_sp = None       

    def generate_arrow_text(self,arrowindex):
        x1,y1 = self.arrow_sp
        x2,y2 = self.arrow_ep
        x = int((x1+x2)/2)
        y = int((y1+y2)/2)

        arrow_ID = arrowindex
        arrowtextid =self.canvas.create_text(x, y, text=str(arrow_ID), font=('Arial', 20), fill='magenta', tags='arrowtext')
        return arrowtextid
    
    def remove_arrow(self,event):
        print("remove arrow")
        mousex = self.canvas.canvasx(event.x)
        mousey = self.canvas.canvasy(event.y) 
        overlapping_items = self.canvas.find_overlapping(mousex, mousey, mousex, mousey)
        print(overlapping_items)
        if overlapping_items:
            for item in overlapping_items:
                tag = self.canvas.gettags(item)
                print(tag)
                if 'arrow' in tag:
                    print("Clicked on arrow:")
                    filterarr = []
                    for trow in self.arrows_arr: #self.arrows_arr = [] #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
                        if trow[1] != item:
                            filterarr.append(trow)
                        if trow[1] == item:
                            self.canvas.delete(trow[6])
                    self.arrows_arr = filterarr
                    print(self.arrows_arr)
                    self.canvas.delete(item)
                if 'arrowtext' in tag:
                    print("Clicked on arrowtext:")
                    filterarr = []
                    for trow in self.arrows_arr: #self.arrows_arr = [] #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
                        if trow[6] != item:
                            filterarr.append(trow)
                        if trow[6] == item:
                            self.canvas.delete(trow[1])
                    self.arrows_arr = filterarr
                    print(self.arrows_arr)
                    self.canvas.delete(item)
    #-----IMAGE_DISPLAY------------------------------------------ 
    def display_image(self, addimg):
        ''' Show image on the Canvas '''
        bbox1 = self.canvas.bbox(self.container)  # get image area
        # Remove 1 pixel shift at the sides of the bbox1
        bbox1 = (bbox1[0] + 1, bbox1[1] + 1, bbox1[2] - 1, bbox1[3] - 1)
        bbox2 = (self.canvas.canvasx(0),  # get visible area of the canvas
                self.canvas.canvasy(0),
                self.canvas.canvasx(self.canvas.winfo_width()),
                self.canvas.canvasy(self.canvas.winfo_height()))
        bbox = [min(bbox1[0], bbox2[0]), min(bbox1[1], bbox2[1]),  # get scroll region box
                max(bbox1[2], bbox2[2]), max(bbox1[3], bbox2[3])]
        if bbox[0] == bbox2[0] and bbox[2] == bbox2[2]:  # whole image in the visible area
            bbox[0] = bbox1[0]
            bbox[2] = bbox1[2]
        if bbox[1] == bbox2[1] and bbox[3] == bbox2[3]:  # whole image in the visible area
            bbox[1] = bbox1[1]
            bbox[3] = bbox1[3]
        self.canvas.configure(scrollregion=bbox)  # set scroll region

        print(f"bbox1 ({bbox1[0]},{bbox1[1]},{bbox1[2]},{bbox1[3]})")
        print(f"bbox2 ({bbox2[0]},{bbox2[1]},{bbox2[2]},{bbox2[3]})")
        print(f"bbox ({bbox[0]},{bbox[1]},{bbox[2]},{bbox[3]})")
        # Store bbox1 for later use
        self.bbox1 = bbox1
        self.bbox2 = bbox2

        x1 = max(bbox2[0] - bbox1[0], 0)  # get coordinates (x1,y1,x2,y2) of the image tile
        y1 = max(bbox2[1] - bbox1[1], 0)
        x2 = min(bbox2[2], bbox1[2]) - bbox1[0]
        y2 = min(bbox2[3], bbox1[3]) - bbox1[1]

        self.si_x1 = x1
        self.si_y1 = y1
        self.si_x2 = x2
        self.si_y2 = y2

        self.grey_image_offset_x = bbox1[0] - bbox2[0]
        self.grey_image_offset_y = bbox1[1] - bbox2[1]       

        print(f"x1,y1,x2,y2 {x1},{y1},{x2},{y2}")
        print(f"grey offset {self.grey_image_offset_x}, {self.grey_image_offset_y}")

        if int(x2 - x1) > 0 and int(y2 - y1) > 0:  # show image if it in the visible area
            x = min(int(x2 / self.imscale), self.width)   # sometimes it is larger on 1 pixel...
            y = min(int(y2 / self.imscale), self.height)  # ...and sometimes not
            #image = self.image.crop((int(x1 / self.imscale), int(y1 / self.imscale), x, y))

            image = addimg.crop((int(x1 / self.imscale), int(y1 / self.imscale), x, y))
           
            imagetk = ImageTk.PhotoImage(image.resize((int(x2 - x1), int(y2 - y1))))
            self.exportimg = imagetk
            
            self.imageid = self.canvas.create_image(max(bbox2[0], bbox1[0]), max(bbox2[1], bbox1[1]), anchor='nw', image=imagetk)
            
            #self.bgimage= self.canvas.lower(self.imageid)  # set image into background
            self.canvas.imagetk = imagetk  # keep an extra reference to prevent garbage-collection

            self.canvas.lower(self.imageid)


        print(self.canvas.coords(self.container))
        self.draw_scalebar()
    def toggle_image_channels(self, event=None):    #Updates off of BC sliders and toggles
            # Initialize added_image as None
        self.addedimage = self.f1f2_toggle_imgprocecessor()

        self.modimage = self.addedimage

        if self.addedimage is not None:
            self.addedimage = cv2.cvtColor(self.addedimage, cv2.COLOR_BGR2RGB)
            ##self.exportimg = self.addedimage.copy()
            addpil_image = Image.fromarray(self.addedimage)
            self.addimgmask = Image.new("L", addpil_image.size, int(255 * 1))
            self.addimgwopac = addpil_image.copy()
            self.addimgwopac.putalpha(self.addimgmask)      
            addimg_processed= np.array(self.addimgwopac)
            print(f"addimg shape {addimg_processed.shape}")
            addimg_processed = Image.fromarray(addimg_processed)

            self.display_image(addimg_processed)    
    def update_image_BC(self,image_in, brightness_slider, contrast_slider):
        # Get current brightness and contrast values
        brightness = brightness_slider.get()
        contrast = contrast_slider.get()
        
        # Normalize image to 16-bit range
        normalized_image = cv2.normalize(image_in, None, 0, 65535, cv2.NORM_MINMAX, dtype=cv2.CV_16U)
        # Convert image to 8-bit uint8 for display
        normalized_image_uint8 = (normalized_image / 256).astype(np.uint8)
        # Convert 16-bit grayscale to RGB
        if len(normalized_image_uint8.shape) == 2:
            normalized_image_uint8 = cv2.cvtColor(normalized_image_uint8, cv2.COLOR_GRAY2RGB)


        image = normalized_image_uint8.astype(np.float32)

        # Adjust brightness
        image += brightness

        # Adjust contrast
        if contrast != 0:
            factor = (259 * (contrast + 255)) / (255 * (259 - contrast))
            image = 128 + factor * (image - 128)

        # Clip values to [0, 255] and convert back to uint8
        image = np.clip(image, 0, 255).astype(np.uint8)

        # Convert to PIL Image
        outimage = Image.fromarray(image)

        return outimage
    def f1f2_toggle_imgprocecessor(self):
        #self.containeradded is the image size with padding in all 4 directions = f2offset
        #put f1 in the center (xy+abs(f2offset.xy))
        #put f2 in center + offset (so it is +/- sensitive)
        width_padcont= self.padimgwidth
        height_padcont =  self.padimgheight

        f2padx = int((width_padcont - self.width) /2)
        print(f"{width_padcont} - {self.width} /2 = {int((width_padcont - self.width) /2)}")

        f2pady = int((height_padcont - self.height) /2)


        f1padx = int((abs(f2padx)))
        f1pady = int((abs(f2pady)))

        #F2 OFFSET
        if self.f2_offset[0] < 0:
            f2padx_l = 0
            f2padx_r = 2*f2padx
        if self.f2_offset[0] > 0:
            f2padx_l = 2*f2padx
            f2padx_r = 0
        if self.f2_offset[0] == 0:
            f2padx_l = 0
            f2padx_r = 0

        if self.f2_offset[1] < 0:
            f2pady_u = 0
            f2pady_l = 2*f2pady
        if self.f2_offset[1] > 0:
            f2pady_u = 2*f2pady
            f2pady_l = 0
        if self.f2_offset[1] == 0:
            f2pady_u = 0
            f2pady_l = 0

        # 3 = BGR?
        padded_baseimg = np.zeros((width_padcont, height_padcont, 3), dtype=np.uint8)
        padded_baseimg = padded_baseimg.transpose(1,0,2)    #numpy goes by height/width rather than width, height


        print(f"padxy {f1padx},{f1pady} ---- {padded_baseimg.shape[0]},{padded_baseimg.shape[1]}")
    

        #Make a black image if there are no toggles on 

        padding_f1 = ((f1pady, f1pady), (f1padx, f1padx), (0, 0))
        padding_f2 = ((f2pady_u, f2pady_l), (f2padx_l, f2padx_r), (0, 0))


        if self.nochannelsselected(): return padded_baseimg.transpose(1,0,2)

        added_image = None
        processed_images = [padded_baseimg]
        #FILE1 CHANNELS-----------------------------------------------------
        if self.F1_C0_toggle.get() and self.F1_C0_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C0_MIP, self.f1_C0_brightness_slider, self.f1_C0_contrast_slider, padding_f1, 0, self))
        if self.F1_C1_toggle.get() and self.F1_C1_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C1_MIP, self.f1_C1_brightness_slider, self.f1_C1_contrast_slider, padding_f1, 1, self))
        if self.F1_C2_toggle.get() and self.F1_C2_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C2_MIP, self.f1_C2_brightness_slider, self.f1_C2_contrast_slider, padding_f1, 2, self))
        if self.F1_C3_toggle.get() and self.F1_C3_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C3_MIP, self.f1_C3_brightness_slider, self.f1_C3_contrast_slider, padding_f1, 3, self))
        if self.F1_C4_toggle.get() and self.F1_C4_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C4_MIP, self.f1_C4_brightness_slider, self.f1_C4_contrast_slider, padding_f1, 4, self))
        if self.F1_C5_toggle.get() and self.F1_C5_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C5_MIP, self.f1_C5_brightness_slider, self.f1_C5_contrast_slider, padding_f1, 5, self))
        if self.F2_C0_toggle.get() and self.F2_C0_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C0_MIP, self.f2_C0_brightness_slider, self.f2_C0_contrast_slider, padding_f2, 6, self))
        if self.F2_C1_toggle.get() and self.F2_C1_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C1_MIP, self.f2_C1_brightness_slider, self.f2_C1_contrast_slider, padding_f2, 7, self))
        if self.F2_C2_toggle.get() and self.F2_C2_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C2_MIP, self.f2_C2_brightness_slider, self.f2_C2_contrast_slider, padding_f2, 8, self))
        if self.F2_C3_toggle.get() and self.F2_C3_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C3_MIP, self.f2_C3_brightness_slider, self.f2_C3_contrast_slider, padding_f2, 9, self))
        if self.F2_C4_toggle.get() and self.F2_C4_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C4_MIP, self.f2_C4_brightness_slider, self.f2_C4_contrast_slider, padding_f2, 10, self))
        if self.F2_C5_toggle.get() and self.F2_C5_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C5_MIP, self.f2_C5_brightness_slider, self.f2_C5_contrast_slider, padding_f2, 11, self))
 
        #Transpose HEIGHTxWIDTH of images
      #  for i, img in enumerate(processed_images):
      #      if i == 0:  # Skip the first element
       #         continue
       #     else: 
       #         print(f"Transposing element {i} from shape {img.shape} to (2062, 2086, 3)")
        #        img = np.transpose(img, (1, 0, 2))  # Switch dimensions if needed
        #        processed_images[i] = img  # Update the list with the transposed image

        # Ensure all elements are numpy arrays and have the correct shape
        if all(isinstance(img, np.ndarray) for img in processed_images):
            added_image = np.sum(processed_images, axis=0)
            added_image = np.clip(added_image, 0, 255).astype(np.uint8)
        else:
            raise ValueError("One or more elements in added_images_list are not NumPy arrays")

        return added_image   
    
    def process_channel(self, mip, brightness_slider, contrast_slider, padding, color_index, f1):
        print(f"Processing channel {color_index}")
        if mip is not None:
            tempcopy = mip.copy()
            bc_adj_img = self.update_image_BC(tempcopy, brightness_slider, contrast_slider)
            ai_array = np.array(bc_adj_img)
            ai_array = self.color_nparr(ai_array, color_index)
            padded_image = np.pad(ai_array, padding, mode='constant')
            print(f"Finished processing channel {color_index}")
            return padded_image
        print(f"Skipped channel {color_index}")
        return None
    
    def process_channel_nopadding(self, mip, brightness_slider, contrast_slider, color_index):
        print(f"Processing channel {color_index}")
        if mip is not None:
            tempcopy = mip.copy()
            bc_adj_img = self.update_image_BC(tempcopy, brightness_slider, contrast_slider)
            ai_array = np.array(bc_adj_img)
            ai_array = self.color_nparr(ai_array, color_index)
            #padded_image = np.pad(ai_array, padding, mode='constant')
            print(f"Finished processing channel {color_index}")
            return ai_array
        print(f"Skipped channel {color_index}")
        return None
    
    def nochannelsselected(self):
        offcounter = 0
        if self.F1_C0_MIP is not None:
            offcounter -= 1
            if self.F1_C0_toggle.get() == False: offcounter += 1
        if self.F1_C1_MIP is not None:
            offcounter -= 1
            if self.F1_C1_toggle.get() == False: offcounter += 1
        if self.F1_C2_MIP is not None:
            offcounter -= 1
            if self.F1_C2_toggle.get() == False: offcounter += 1
        if self.F1_C3_MIP is not None:
            offcounter -= 1
            if self.F1_C3_toggle.get() == False: offcounter += 1
        if self.F1_C4_MIP is not None:
            offcounter -= 1
            if self.F1_C4_toggle.get() == False: offcounter += 1
        if self.F1_C5_MIP is not None:
            offcounter -= 1
            if self.F1_C5_toggle.get() == False: offcounter += 1
        if self.F2_C0_MIP is not None:
            offcounter -= 1
            if self.F2_C0_toggle.get() == False: offcounter += 1
        if self.F2_C1_MIP is not None:
            offcounter -= 1
            if self.F2_C1_toggle.get() == False: offcounter += 1
        if self.F2_C2_MIP is not None:
            offcounter -= 1
            if self.F2_C2_toggle.get() == False: offcounter += 1
        if self.F2_C3_MIP is not None:
            offcounter -= 1
            if self.F2_C3_toggle.get() == False: offcounter += 1
        if self.F2_C4_MIP is not None:
            offcounter -= 1
            if self.F2_C4_toggle.get() == False: offcounter += 1
        if self.F2_C5_MIP is not None:
            offcounter -= 1
            if self.F2_C5_toggle.get() == False: offcounter += 1
        if offcounter >= 0: return True
        else: return False

    def show_image(self, event=None):   #Updates of window resize and scrollbars
        added_image = None

            # Check if the images are not None and add them to the result
        if self.F1_C0_MIP is not None:
                if self.F1_C0_toggle.get():
                    self.F1_C0_MIP = np.array(self.F1_C0_MIP)
                    pil_image = Image.fromarray(self.F1_C0_MIP)
                    adjusted_image = self.update_image_BC(pil_image,self.f1_C0_brightness_slider,self.f1_C0_contrast_slider)
                    ai_array = np.array(adjusted_image)
                    added_image = ai_array.copy()

        if self.F1_C1_MIP is not None:
                if self.F1_C1_toggle.get():
                    if added_image is None:
                        self.F1_C1_MIP = np.array(self.F1_C1_MIP)
                        pil_image = Image.fromarray(self.F1_C1_MIP)
                        adjusted_image = self.update_image_BC(pil_image,self.f1_C1_brightness_slider,self.f1_C1_contrast_slider)
                        ai_array = np.array(adjusted_image)
                        added_image = ai_array.copy()
                    else:
                        # Ensure both images are numpy arrays
                        added_image = np.array(added_image)
                        self.F1_C1_MIP = np.array(self.F1_C1_MIP)
                        pil_image = Image.fromarray(self.F1_C1_MIP)
                        adjusted_image = self.update_image_BC(pil_image,self.f1_C1_brightness_slider,self.f1_C1_contrast_slider)
                        ai_array = np.array(adjusted_image)
                        added_image = cv2.add(added_image, ai_array)

        if self.F1_C2_MIP is not None:
                if self.F1_C2_toggle.get():
                    if added_image is None:
                        self.F1_C2_MIP = np.array(self.F1_C2_MIP)
                        pil_image = Image.fromarray(self.F1_C2_MIP)
                        adjusted_image = self.update_image_BC(pil_image,self.f1_C2_brightness_slider,self.f1_C2_contrast_slider)
                        ai_array = np.array(adjusted_image)
                        added_image = ai_array.copy()
                    else:
                        # Ensure both images are numpy arrays
                        added_image = np.array(added_image)
                        self.F1_C2_MIP = np.array(self.F1_C2_MIP)
                        pil_image = Image.fromarray(self.F1_C2_MIP)
                        adjusted_image = self.update_image_BC(pil_image,self.f1_C2_brightness_slider,self.f1_C2_contrast_slider)
                        ai_array = np.array(adjusted_image)
                        added_image = cv2.add(added_image, ai_array)

        if self.F1_C3_toggle.get() and self.F1_C3_MIP is not None:
                if self.F1_C3_toggle.get():
                    if added_image is None:
                        self.F1_C3_MIP = np.array(self.F1_C3_MIP)
                        pil_image = Image.fromarray(self.F1_C3_MIP)
                        adjusted_image = self.update_image_BC(pil_image,self.f1_C3_brightness_slider,self.f1_C3_contrast_slider)
                        ai_array = np.array(adjusted_image)
                        added_image = ai_array.copy()
                    else:
                        # Ensure both images are numpy arrays
                        added_image = np.array(added_image)
                        self.F1_C3_MIP = np.array(self.F1_C3_MIP)
                        pil_image = Image.fromarray(self.F1_C3_MIP)
                        adjusted_image = self.update_image_BC(pil_image,self.f1_C3_brightness_slider,self.f1_C3_contrast_slider)
                        ai_array = np.array(adjusted_image)
                        added_image = cv2.add(added_image, self.F1_C3_MIP)

            # Check if added_image is not None and perform further processing
        if added_image is not None:
                # Convert the image to RGB color space
                pil_image = Image.fromarray(added_image)
                image_in_arr = np.array(pil_image)
                image = Image.fromarray(image_in_arr)

        
        ''' Show image on the Canvas '''
        bbox1 = self.canvas.bbox(self.container)  # get image area
        # Remove 1 pixel shift at the sides of the bbox1
        bbox1 = (bbox1[0] + 1, bbox1[1] + 1, bbox1[2] - 1, bbox1[3] - 1)
        bbox2 = (self.canvas.canvasx(0),  # get visible area of the canvas
                self.canvas.canvasy(0),
                self.canvas.canvasx(self.canvas.winfo_width()),
                self.canvas.canvasy(self.canvas.winfo_height()))
        bbox = [min(bbox1[0], bbox2[0]), min(bbox1[1], bbox2[1]),  # get scroll region box
                max(bbox1[2], bbox2[2]), max(bbox1[3], bbox2[3])]
        if bbox[0] == bbox2[0] and bbox[2] == bbox2[2]:  # whole image in the visible area
            bbox[0] = bbox1[0]
            bbox[2] = bbox1[2]
        if bbox[1] == bbox2[1] and bbox[3] == bbox2[3]:  # whole image in the visible area
            bbox[1] = bbox1[1]
            bbox[3] = bbox1[3]
        self.canvas.configure(scrollregion=bbox)  # set scroll region

        # Store bbox1 for later use
        self.bbox1 = bbox1

        x1 = max(bbox2[0] - bbox1[0], 0)  # get coordinates (x1,y1,x2,y2) of the image tile
        y1 = max(bbox2[1] - bbox1[1], 0)
        x2 = min(bbox2[2], bbox1[2]) - bbox1[0]
        y2 = min(bbox2[3], bbox1[3]) - bbox1[1]
        if int(x2 - x1) > 0 and int(y2 - y1) > 0:  # show image if it in the visible area
            x = min(int(x2 / self.imscale), self.width)   # sometimes it is larger on 1 pixel...
            y = min(int(y2 / self.imscale), self.height)  # ...and sometimes not
            image = image.crop((int(x1 / self.imscale), int(y1 / self.imscale), x, y))
            #image = self.image.crop((int(x1 / self.imscale), int(y1 / self.imscale), x, y))
            imagetk = ImageTk.PhotoImage(image.resize((int(x2 - x1), int(y2 - y1))))
            imageid = self.canvas.create_image(max(bbox2[0], bbox1[0]), max(bbox2[1], bbox1[1]),
                                            anchor='nw', image=imagetk)
            self.canvas.lower(imageid)  # set image into background
            self.canvas.imagetk = imagetk  # keep an extra reference to prevent garbage-collection
        self.draw_scalebar() 
    def toggle_kp_visible(self):
        if self.F1_C0_toggle_KP.get():
            for row in self.kp_poly_F1C0_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F1C0_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F1_C1_toggle_KP.get():
            for row in self.kp_poly_F1C1_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F1C1_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F1_C2_toggle_KP.get():
            for row in self.kp_poly_F1C2_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F1C2_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F1_C3_toggle_KP.get():
            for row in self.kp_poly_F1C3_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F1C3_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F1_C4_toggle_KP.get():
            for row in self.kp_poly_F1C4_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F1C4_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F1_C5_toggle_KP.get():
            for row in self.kp_poly_F1C5_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F1C5_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        #FILE 2
        if self.F2_C0_toggle_KP.get():
            for row in self.kp_poly_F2C0_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F2C0_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F2_C1_toggle_KP.get(): 
            for row in self.kp_poly_F2C1_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F2C1_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F2_C2_toggle_KP.get():
            for row in self.kp_poly_F2C2_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F2C2_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F2_C3_toggle_KP.get():
            for row in self.kp_poly_F2C3_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F2C3_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F2_C4_toggle_KP.get():
            for row in self.kp_poly_F2C4_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F2C4_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
        if self.F2_C5_toggle_KP.get():
            for row in self.kp_poly_F2C5_arr:
                self.canvas.itemconfigure(row[1], state="normal")
        else: 
            for row in self.kp_poly_F2C5_arr:
                self.canvas.itemconfigure(row[1], state="hidden")
    #-----Scalebar------------------------------------------------
    def toggle_scalebar(self):
        if self.scalebar_visible:
            # Hide the scale bar
            self.canvas.delete("scalebar")
        else:
            # Show the scale bar
            self.draw_scalebar()  # Adjust the length of the scale bar as needed
        self.scalebar_visible = not self.scalebar_visible
    def draw_scalebar(self):
        #1px =  0.108333333333333 microns
        # Determine the position of the scale bar relative to the image
        # For example, you may want to place it at the bottom-right corner
        self.canvas.delete('scalebar')

        if self.scalebar_toggle_var.get():
            px_per_micron = 0.108333333333333
            scalebar_microns = 10
            scale_bar_width = scalebar_microns / px_per_micron * self.imscale
            scale_bar_height = 5  # Height of the scale bar rectangle


            # Calculate the coordinates for the scale bar
            window_width = self.master.winfo_width() 
            window_height = self.master.winfo_height() 
            scale_bar_x0 = self.bbox2[0] + 20    # Adjust as needed
            scale_bar_y0 = self.bbox2[3] - 20  # Adjust as needed
            scale_bar_x1 = scale_bar_x0 + scale_bar_width   # Adjust as needed
            scale_bar_y1 = scale_bar_y0 - scale_bar_height  # Adjust as needed

            # Draw the scale bar
            self.scalebarwidget = self.canvas.create_rectangle(scale_bar_x0, scale_bar_y0, scale_bar_x1, scale_bar_y1, fill="white", tags="scalebar")

            # Optionally, add text to indicate the length of the scale bar
            scale_text_x = scale_bar_x0 + 50
            scale_text_y = scale_bar_y0 - 15  # Adjust as needed
            self.canvas.create_text(scale_text_x, scale_text_y, text=f"{scalebar_microns} um ({round(scale_bar_width,1)} px)", fill="white", tags="scalebar")
    #-----Initialization------------------------------------------
    def get_MIP_file(self,target_dir,channels,channum):
        print(f"get mip channel length = {len(channels)} --- {channels} ---- {channum}")
        if channum >= len(channels): return None

        filepath = os.path.join(target_dir,f"C{channum}_{channels[channum]}")

        print(filepath)
        target_dir = filepath

        files = os.listdir(target_dir)
        if len(files) == 0: return None
        else:
            image_name= os.listdir(target_dir)[0]
            image_path = os.path.join(target_dir,image_name)
            print(image_path)
            return cv2.imread(image_path, cv2.IMREAD_UNCHANGED)
    def bc_frame_maker(self):
        self.bc_mframe = tk.Frame(self.control_window)
        self.bc_mframe.grid(row=5,column=0)

        brightness_min = -200
        brightness_max = 200
        contrast_min = -200
        contrast_max = 200
        debugtog = False

        self.bc_f1maker(brightness_max,brightness_min,contrast_max,contrast_min,debugtog)

        self.bcspacerframe = tk.Frame(self.bc_mframe)
        self.bcspacerframe.grid(row=5,column=7,padx=20)

        self.bc_f2maker(brightness_max,brightness_min,contrast_max,contrast_min,debugtog)


    def bc_f1maker(self,brightness_max,brightness_min,contrast_max,contrast_min,debugtog):

        self.bc1frame = tk.Frame(self.bc_mframe)
        self.bc1frame.grid(row=5,column=5)

        #CHANNEL TOGGLES
        self.F1_C0_toggle_checkbox = None if self.F1_C0_MIP is None else tk.Checkbutton(self.bc1frame, text=f"{self.f1nametag}_C0_{self.f1channels[0]}", variable=self.F1_C0_toggle, command=self.toggle_image_channels).grid(row=50, column=50, sticky='w')
        self.F1_C1_toggle_checkbox = None if self.F1_C1_MIP is None else tk.Checkbutton(self.bc1frame, text=f"{self.f1nametag}_C1_{self.f1channels[1]}", variable=self.F1_C1_toggle, command=self.toggle_image_channels).grid(row=100, column=50, sticky='w')
        self.F1_C2_toggle_checkbox = None if self.F1_C2_MIP is None else tk.Checkbutton(self.bc1frame, text=f"{self.f1nametag}_C2_{self.f1channels[2]}", variable=self.F1_C2_toggle, command=self.toggle_image_channels).grid(row=150, column=50, sticky='w')
        self.F1_C3_toggle_checkbox = None if self.F1_C3_MIP is None else tk.Checkbutton(self.bc1frame, text=f"{self.f1nametag}_C3_{self.f1channels[3]}", variable=self.F1_C3_toggle, command=self.toggle_image_channels).grid(row=200, column=50, sticky='w')
        self.F1_C4_toggle_checkbox = None if self.F1_C4_MIP is None else tk.Checkbutton(self.bc1frame, text=f"{self.f1nametag}_C4_{self.f1channels[4]}", variable=self.F1_C4_toggle, command=self.toggle_image_channels).grid(row=250, column=50, sticky='w')
        self.F1_C5_toggle_checkbox = None if self.F1_C5_MIP is None else tk.Checkbutton(self.bc1frame, text=f"{self.f1nametag}_C5_{self.f1channels[5]}", variable=self.F1_C5_toggle, command=self.toggle_image_channels).grid(row=300, column=50, sticky='w')
    
        #COLOR PICK SQUARES
        self.F1_C0_cpick_square = None if self.F1_C0_MIP is None else tk.Label(self.bc1frame, width=2, height=1, bg=self.chancolorarr[0], relief="solid")
        if self.F1_C0_MIP is not None: self.F1_C0_cpick_square.grid(row=50, column=60, sticky='w')
        self.F1_C1_cpick_square = None if self.F1_C1_MIP is None else tk.Label(self.bc1frame, width=2, height=1, bg=self.chancolorarr[1], relief="solid")
        if self.F1_C1_MIP is not None: self.F1_C1_cpick_square.grid(row=100, column=60, sticky='w')
        self.F1_C2_cpick_square = None if self.F1_C2_MIP is None else tk.Label(self.bc1frame, width=2, height=1, bg=self.chancolorarr[2], relief="solid")
        if self.F1_C2_MIP is not None: self.F1_C2_cpick_square.grid(row=150, column=60, sticky='w')
        self.F1_C3_cpick_square = None if self.F1_C3_MIP is None else tk.Label(self.bc1frame, width=2, height=1, bg=self.chancolorarr[3], relief="solid")
        if self.F1_C3_MIP is not None: self.F1_C3_cpick_square.grid(row=200, column=60, sticky='w')
        self.F1_C4_cpick_square = None if self.F1_C4_MIP is None else tk.Label(self.bc1frame, width=2, height=1, bg=self.chancolorarr[4], relief="solid")
        if self.F1_C4_MIP is not None: self.F1_C4_cpick_square.grid(row=250, column=60, sticky='w')
        self.F1_C5_cpick_square = None if self.F1_C5_MIP is None else tk.Label(self.bc1frame, width=2, height=1, bg=self.chancolorarr[5], relief="solid")
        if self.F1_C5_MIP is not None: self.F1_C5_cpick_square.grid(row=300, column=60, sticky='w')

        #COLOR PICK BUTTON
        self.F1_C0_cpick_b = None if self.F1_C0_MIP is None else tk.Button(self.bc1frame, text=f"{self.f1nametag}_C0_{self.f1channels[0]}_COLOR", command=lambda: self.pick_color(1, 0,self.F1_C0_cpick_square))
        if self.F1_C0_MIP is not None: self.F1_C0_cpick_b.grid(row=50, column=55, sticky='w', pady=10)
        self.F1_C1_cpick_b = None if self.F1_C1_MIP is None else tk.Button(self.bc1frame, text=f"{self.f1nametag}_C1_{self.f1channels[1]}_COLOR", command=lambda: self.pick_color(1, 1,self.F1_C1_cpick_square))
        if self.F1_C1_MIP is not None: self.F1_C1_cpick_b.grid(row=100, column=55, sticky='w', pady=10)
        self.F1_C2_cpick_b = None if self.F1_C2_MIP is None else tk.Button(self.bc1frame, text=f"{self.f1nametag}_C2_{self.f1channels[2]}_COLOR", command=lambda: self.pick_color(1, 2,self.F1_C2_cpick_square))
        if self.F1_C2_MIP is not None: self.F1_C2_cpick_b.grid(row=150, column=55, sticky='w', pady=10)
        self.F1_C3_cpick_b = None if self.F1_C3_MIP is None else tk.Button(self.bc1frame, text=f"{self.f1nametag}_C3_{self.f1channels[3]}_COLOR", command=lambda: self.pick_color(1, 3,self.F1_C3_cpick_square))
        if self.F1_C3_MIP is not None: self.F1_C3_cpick_b.grid(row=200, column=55, sticky='w', pady=10)
        self.F1_C4_cpick_b = None if self.F1_C4_MIP is None else tk.Button(self.bc1frame, text=f"{self.f1nametag}_C4_{self.f1channels[4]}_COLOR", command=lambda: self.pick_color(1, 4,self.F1_C4_cpick_square))
        if self.F1_C4_MIP is not None: self.F1_C4_cpick_b.grid(row=250, column=55, sticky='w', pady=10)
        self.F1_C5_cpick_b = None if self.F1_C5_MIP is None else tk.Button(self.bc1frame, text=f"{self.f1nametag}_C5_{self.f1channels[5]}_COLOR", command=lambda: self.pick_color(1, 5,self.F1_C5_cpick_square))
        if self.F1_C5_MIP is not None: self.F1_C5_cpick_b.grid(row=300, column=55, sticky='w', pady=10)

        #F1C0--------------------------------------------------------
        if self.F1_C0_MIP is not None or debugtog==True:
            # Create a Label for Brightness
            self.f1_C0_brightness_label = tk.Label(self.bc1frame, text=f"Brightness")
            self.f1_C0_brightness_label.grid(row=60, column=50,sticky='s')
            # Create a Scale widget for Brightness
            self.f1_C0_brightness_slider = tk.Scale(self.bc1frame, from_=brightness_min, to=brightness_max, length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C0_brightness_slider.set(0)  # Set default value
            self.f1_C0_brightness_slider.grid(row=60, column=55)
            # Create a Label for Contrast
            self.f1_C0_contrast_label = tk.Label(self.bc1frame, text=f"Contrast")
            self.f1_C0_contrast_label.grid(row=70, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f1_C0_contrast_slider = tk.Scale(self.bc1frame, from_=contrast_min, to=contrast_max, length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C0_contrast_slider.set(0)  # Set default value
            self.f1_C0_contrast_slider.grid(row=70, column=55)
        #F1C1--------------------------------------------------------
        if self.F1_C1_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f1_C1_brightness_label = tk.Label(self.bc1frame, text=f"Brightness")
            self.f1_C1_brightness_label.grid(row=110, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f1_C1_brightness_slider = tk.Scale(self.bc1frame, from_=brightness_min, to=brightness_max, length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C1_brightness_slider.set(0)  # Set default value
            self.f1_C1_brightness_slider.grid(row=110, column=55)
            # Create a Label for Contrast
            self.f1_C1_contrast_label = tk.Label(self.bc1frame, text="Contrast")
            self.f1_C1_contrast_label.grid(row=120, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f1_C1_contrast_slider = tk.Scale(self.bc1frame, from_=contrast_min, to=contrast_max,length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C1_contrast_slider.set(0)  # Set default value
            self.f1_C1_contrast_slider.grid(row=120, column=55)
        #F1C2--------------------------------------------------------
        if self.F1_C2_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f1_C2_brightness_label = tk.Label(self.bc1frame, text="Brightness")
            self.f1_C2_brightness_label.grid(row=160, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f1_C2_brightness_slider = tk.Scale(self.bc1frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C2_brightness_slider.set(0)  # Set default value
            self.f1_C2_brightness_slider.grid(row=160, column=55)
            # Create a Label for Contrast
            self.f1_C2_contrast_label = tk.Label(self.bc1frame, text="Contrast")
            self.f1_C2_contrast_label.grid(row=170, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f1_C2_contrast_slider = tk.Scale(self.bc1frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C2_contrast_slider.set(0)  # Set default value
            self.f1_C2_contrast_slider.grid(row=170, column=55)
        #F1C3--------------------------------------------------------
        if self.F1_C3_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f1_C3_brightness_label = tk.Label(self.bc1frame, text="Brightness")
            self.f1_C3_brightness_label.grid(row=210, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f1_C3_brightness_slider = tk.Scale(self.bc1frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C3_brightness_slider.set(0)  # Set default value
            self.f1_C3_brightness_slider.grid(row=210, column=55)
            # Create a Label for Contrast
            self.f1_C3_contrast_label = tk.Label(self.bc1frame, text="Contrast")
            self.f1_C3_contrast_label.grid(row=220, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f1_C3_contrast_slider= tk.Scale(self.bc1frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C3_contrast_slider.set(0)  # Set default value
            self.f1_C3_contrast_slider.grid(row=220, column=55)
        #F1C4--------------------------------------------------------
        if self.F1_C4_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f1_C4_brightness_label = tk.Label(self.bc1frame, text="Brightness")
            self.f1_C4_brightness_label.grid(row=260, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f1_C4_brightness_slider = tk.Scale(self.bc1frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C4_brightness_slider.set(0)  # Set default value
            self.f1_C4_brightness_slider.grid(row=260, column=55)
            # Create a Label for Contrast
            self.f1_C4_contrast_label = tk.Label(self.bc1frame, text="Contrast")
            self.f1_C4_contrast_label.grid(row=270, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f1_C4_contrast_slider= tk.Scale(self.bc1frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C4_contrast_slider.set(0)  # Set default value
            self.f1_C4_contrast_slider.grid(row=270, column=55)
        #F1C5--------------------------------------------------------
        if self.F1_C5_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f1_C5_brightness_label = tk.Label(self.bc1frame, text="Brightness")
            self.f1_C5_brightness_label.grid(row=310, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f1_C5_brightness_slider = tk.Scale(self.bc1frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C5_brightness_slider.set(0)  # Set default value
            self.f1_C5_brightness_slider.grid(row=310, column=55)
            # Create a Label for Contrast
            self.f1_C5_contrast_label = tk.Label(self.bc1frame, text="Contrast")
            self.f1_C5_contrast_label.grid(row=320, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f1_C5_contrast_slider= tk.Scale(self.bc1frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f1_C5_contrast_slider.set(0)  # Set default value
            self.f1_C5_contrast_slider.grid(row=320, column=55)

    def bc_f2maker(self,brightness_max,brightness_min,contrast_max,contrast_min,debugtog):
        self.bc2frame = tk.Frame(self.bc_mframe)
        self.bc2frame.grid(row=5,column=10)

        #CHANNEL TOGGLES
        self.F2_C0_toggle_checkbox = None if self.F2_C0_MIP is None else tk.Checkbutton(self.bc2frame, text=f"{self.f2nametag}_C0_{self.f2channels[0]}", variable=self.F2_C0_toggle, command=self.toggle_image_channels).grid(row=50, column=50, sticky='w')
        self.F2_C1_toggle_checkbox = None if self.F2_C1_MIP is None else tk.Checkbutton(self.bc2frame, text=f"{self.f2nametag}_C1_{self.f2channels[1]}", variable=self.F2_C1_toggle, command=self.toggle_image_channels).grid(row=100, column=50, sticky='w')
        self.F2_C2_toggle_checkbox = None if self.F2_C2_MIP is None else tk.Checkbutton(self.bc2frame, text=f"{self.f2nametag}_C2_{self.f2channels[2]}", variable=self.F2_C2_toggle, command=self.toggle_image_channels).grid(row=150, column=50, sticky='w')
        self.F2_C3_toggle_checkbox = None if self.F2_C3_MIP is None else tk.Checkbutton(self.bc2frame, text=f"{self.f2nametag}_C3_{self.f2channels[3]}", variable=self.F2_C3_toggle, command=self.toggle_image_channels).grid(row=200, column=50, sticky='w')
        self.F2_C4_toggle_checkbox = None if self.F2_C4_MIP is None else tk.Checkbutton(self.bc2frame, text=f"{self.f2nametag}_C4_{self.f2channels[4]}", variable=self.F2_C4_toggle, command=self.toggle_image_channels).grid(row=250, column=50, sticky='w')
        self.F2_C5_toggle_checkbox = None if self.F2_C5_MIP is None else tk.Checkbutton(self.bc2frame, text=f"{self.f2nametag}_C5_{self.f2channels[5]}", variable=self.F2_C5_toggle, command=self.toggle_image_channels).grid(row=300, column=50, sticky='w')
    
        #COLOR PICK SQUARES
        self.F2_C0_cpick_square = None if self.F2_C0_MIP is None else tk.Label(self.bc2frame, width=2, height=1, bg=self.chancolorarr[6], relief="solid")
        if self.F2_C0_MIP is not None: self.F2_C0_cpick_square.grid(row=50, column=60, sticky='w')
        self.F2_C1_cpick_square = None if self.F2_C1_MIP is None else tk.Label(self.bc2frame, width=2, height=1, bg=self.chancolorarr[7], relief="solid")
        if self.F2_C1_MIP is not None: self.F2_C1_cpick_square.grid(row=100, column=60, sticky='w')
        self.F2_C2_cpick_square = None if self.F2_C2_MIP is None else tk.Label(self.bc2frame, width=2, height=1, bg=self.chancolorarr[8], relief="solid")
        if self.F2_C2_MIP is not None: self.F2_C2_cpick_square.grid(row=150, column=60, sticky='w')
        self.F2_C3_cpick_square = None if self.F2_C3_MIP is None else tk.Label(self.bc2frame, width=2, height=1, bg=self.chancolorarr[9], relief="solid")
        if self.F2_C3_MIP is not None: self.F2_C3_cpick_square.grid(row=200, column=60, sticky='w')
        self.F2_C4_cpick_square = None if self.F2_C4_MIP is None else tk.Label(self.bc2frame, width=2, height=1, bg=self.chancolorarr[10], relief="solid")
        if self.F2_C4_MIP is not None: self.F2_C4_cpick_square.grid(row=250, column=60, sticky='w')
        self.F2_C5_cpick_square = None if self.F2_C5_MIP is None else tk.Label(self.bc2frame, width=2, height=1, bg=self.chancolorarr[11], relief="solid")
        if self.F2_C5_MIP is not None: self.F2_C5_cpick_square.grid(row=300, column=60, sticky='w')

        #COLOR PICK BUTTON
        self.F2_C0_cpick_b = None if self.F2_C0_MIP is None else tk.Button(self.bc2frame, text=f"{self.f2nametag}_C0_{self.f2channels[0]}_COLOR", command=lambda: self.pick_color(2, 6,self.F2_C0_cpick_square)).grid(row=50, column=55, sticky='w',pady=10)
        self.F2_C1_cpick_b = None if self.F2_C1_MIP is None else tk.Button(self.bc2frame, text=f"{self.f2nametag}_C1_{self.f2channels[1]}_COLOR", command=lambda: self.pick_color(2, 7,self.F2_C1_cpick_square)).grid(row=100, column=55, sticky='w',pady=10)
        self.F2_C2_cpick_b = None if self.F2_C2_MIP is None else tk.Button(self.bc2frame, text=f"{self.f2nametag}_C2_{self.f2channels[2]}_COLOR", command=lambda: self.pick_color(2, 8,self.F2_C2_cpick_square)).grid(row=150, column=55, sticky='w',pady=10)
        self.F2_C3_cpick_b = None if self.F2_C3_MIP is None else tk.Button(self.bc2frame, text=f"{self.f2nametag}_C3_{self.f2channels[3]}_COLOR", command=lambda: self.pick_color(2, 9,self.F2_C3_cpick_square)).grid(row=200, column=55, sticky='w',pady=10)   
        self.F2_C4_cpick_b = None if self.F2_C4_MIP is None else tk.Button(self.bc2frame, text=f"{self.f2nametag}_C4_{self.f2channels[4]}_COLOR", command=lambda: self.pick_color(2, 10,self.F2_C4_cpick_square)).grid(row=250, column=55, sticky='w',pady=10)      
        self.F2_C5_cpick_b = None if self.F2_C5_MIP is None else tk.Button(self.bc2frame, text=f"{self.f2nametag}_C5_{self.f2channels[5]}_COLOR", command=lambda: self.pick_color(2, 11,self.F2_C5_cpick_square)).grid(row=300, column=55, sticky='w',pady=10)

        #F2C0--------------------------------------------------------
        if self.F2_C0_MIP is not None or debugtog==True:
            # Create a Label for Brightness
            self.f2_C0_brightness_label = tk.Label(self.bc2frame, text=f"Brightness")
            self.f2_C0_brightness_label.grid(row=60, column=50,sticky='s')
            # Create a Scale widget for Brightness
            self.f2_C0_brightness_slider = tk.Scale(self.bc2frame, from_=brightness_min, to=brightness_max, length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C0_brightness_slider.set(0)  # Set default value
            self.f2_C0_brightness_slider.grid(row=60, column=55)
            # Create a Label for Contrast
            self.f2_C0_contrast_label = tk.Label(self.bc2frame, text=f"Contrast")
            self.f2_C0_contrast_label.grid(row=70, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f2_C0_contrast_slider = tk.Scale(self.bc2frame, from_=contrast_min, to=contrast_max, length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C0_contrast_slider.set(0)  # Set default value
            self.f2_C0_contrast_slider.grid(row=70, column=55)
        #F2C1--------------------------------------------------------
        if self.F2_C1_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f2_C1_brightness_label = tk.Label(self.bc2frame, text=f"Brightness")
            self.f2_C1_brightness_label.grid(row=110, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f2_C1_brightness_slider = tk.Scale(self.bc2frame, from_=brightness_min, to=brightness_max, length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C1_brightness_slider.set(0)  # Set default value
            self.f2_C1_brightness_slider.grid(row=110, column=55)
            # Create a Label for Contrast
            self.f2_C1_contrast_label = tk.Label(self.bc2frame, text="Contrast")
            self.f2_C1_contrast_label.grid(row=120, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f2_C1_contrast_slider = tk.Scale(self.bc2frame, from_=contrast_min, to=contrast_max,length=150, orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C1_contrast_slider.set(0)  # Set default value
            self.f2_C1_contrast_slider.grid(row=120, column=55)
        #F2C2--------------------------------------------------------
        if self.F2_C2_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f2_C2_brightness_label = tk.Label(self.bc2frame, text="Brightness")
            self.f2_C2_brightness_label.grid(row=160, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f2_C2_brightness_slider = tk.Scale(self.bc2frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C2_brightness_slider.set(0)  # Set default value
            self.f2_C2_brightness_slider.grid(row=160, column=55)
            # Create a Label for Contrast
            self.f2_C2_contrast_label = tk.Label(self.bc2frame, text="Contrast")
            self.f2_C2_contrast_label.grid(row=170, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f2_C2_contrast_slider = tk.Scale(self.bc2frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C2_contrast_slider.set(0)  # Set default value
            self.f2_C2_contrast_slider.grid(row=170, column=55)
        #F2C3--------------------------------------------------------
        if self.F2_C3_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f2_C3_brightness_label = tk.Label(self.bc2frame, text="Brightness")
            self.f2_C3_brightness_label.grid(row=210, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f2_C3_brightness_slider = tk.Scale(self.bc2frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C3_brightness_slider.set(0)  # Set default value
            self.f2_C3_brightness_slider.grid(row=210, column=55)
            # Create a Label for Contrast
            self.f2_C3_contrast_label = tk.Label(self.bc2frame, text="Contrast")
            self.f2_C3_contrast_label.grid(row=220, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f2_C3_contrast_slider= tk.Scale(self.bc2frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C3_contrast_slider.set(0)  # Set default value
            self.f2_C3_contrast_slider.grid(row=220, column=55)
        #F2C4--------------------------------------------------------
        if self.F2_C4_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f2_C4_brightness_label = tk.Label(self.bc2frame, text="Brightness")
            self.f2_C4_brightness_label.grid(row=260, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f2_C4_brightness_slider = tk.Scale(self.bc2frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C4_brightness_slider.set(0)  # Set default value
            self.f2_C4_brightness_slider.grid(row=260, column=55)
            # Create a Label for Contrast
            self.f2_C4_contrast_label = tk.Label(self.bc2frame, text="Contrast")
            self.f2_C4_contrast_label.grid(row=270, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f2_C4_contrast_slider= tk.Scale(self.bc2frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C4_contrast_slider.set(0)  # Set default value
            self.f2_C4_contrast_slider.grid(row=270, column=55)
        #F2C5--------------------------------------------------------
        if self.F2_C5_MIP is not None or debugtog == True:
            # Create a Label for Brightness
            self.f2_C5_brightness_label = tk.Label(self.bc2frame, text="Brightness")
            self.f2_C5_brightness_label.grid(row=310, column=50, sticky='s')
            # Create a Scale widget for Brightness
            self.f2_C5_brightness_slider = tk.Scale(self.bc2frame, from_=brightness_min, to=brightness_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C5_brightness_slider.set(0)  # Set default value
            self.f2_C5_brightness_slider.grid(row=310, column=55)
            # Create a Label for Contrast
            self.f2_C5_contrast_label = tk.Label(self.bc2frame, text="Contrast")
            self.f2_C5_contrast_label.grid(row=320, column=50,sticky='s')
            # Create a Scale widget for Contrast
            self.f2_C5_contrast_slider= tk.Scale(self.bc2frame, from_=contrast_min, to=contrast_max,length=150,  orient="horizontal", command=lambda value: self.toggle_image_channels())
            self.f2_C5_contrast_slider.set(0)  # Set default value
            self.f2_C5_contrast_slider.grid(row=320, column=55)

    def chan_kp_autopicker_framemaker(self):
        kpchan_frame = tk.Frame(self.control_window)
        kpchan_frame.grid(row=100,column=0,columnspan=3)

        self.kpchantitle_label = tk.Label(kpchan_frame, text="------------------------------------------------------------Channel ROI Autopicking------------------------------------------------------------")
        self.kpchantitle_label.grid(row=0, column=0, columnspan=8, padx=5, pady=5)

        rb_options = []   #radio button to define channel to pick kp manually

        # Create a variable to hold the selected option
        var = tk.StringVar()

        # Create and place the Radiobuttons
        self.enableautopicking = False
        self.enableanalysistogs = False

        debugtog = False

        #HEADERS:
        self.kprbheader = tk.Label(kpchan_frame, text="Channel ROI").grid(row=1,column=0)
        self.kptoggleheader = tk.Label(kpchan_frame,text="Channel ROI Visible").grid(row=1,column=1)
        self.kproisizeheader = tk.Label(kpchan_frame,text="ROI Radius").grid(row=1,column=2,columnspan=2)
        self.kpheadspacer1 = tk.Label(kpchan_frame,text="").grid(row=1,column=3)
        self.kpcolocheader = tk.Label(kpchan_frame,text="Coloc Z Thresh").grid(row=1,column=4,columnspan=2)
        self.kpheadspacer2 = tk.Label(kpchan_frame,text="").grid(row=1,column=5)
        self.chanoutheader = tk.Label(kpchan_frame,text="Channel Analysis")
        self.chanoutheader .grid(row=1,column=7)
        self.kpautopickheader = tk.Label(kpchan_frame,text="Autopick Channel ROIs")
        self.kpautopickheader.grid(row=1,column=8)
        self.kpremoveroiheader = tk.Label(kpchan_frame,text="Remove ALL Channel ROIs:").grid(row=1,column=10)

        if self.enableautopicking == False: self.kpautopickheader.grid_remove()
        if self.enableanalysistogs == False: self.chanoutheader.grid_remove() 
        #F1C0--------------------------------------------------------
        if self.F1_C0_MIP is not None or debugtog == True:
            rb_options.append("F1_C0")
            #F1_C0 AP Label
            #self.F1_C0_autopicker_label = tk.Label(kpchan_frame, text="F1_C0").grid(row=2, column=1, padx=5, pady=5)
            self.F1_C0_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[0]}_ROI", variable=self.F1_C0_toggle_KP, command=self.toggle_kp_visible)
            self.F1_C0_KP_toggle_cb.grid(row=2, column=1, sticky='w')
            #F1_C0 AP Entry
            self.F1_C0_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C0_ap_entry.insert(0,"0.5")
            self.F1_C0_ap_entry.grid(row=2,column=2,sticky='e')
            self.F1_C0_ap_entry_label = tk.Label(kpchan_frame, text="um") 
            self.F1_C0_ap_entry_label.grid(row=2,column=3, sticky="w") 
            #F1_COB COLOC ENTRY
            self.F1_C0_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C0_coloc_entry.insert(0, "1.0")
            self.F1_C0_coloc_entry.grid(row=2,column=4,sticky='e')
            self.F1_C0_coloc_entry_label = tk.Label(kpchan_frame, text="um") 
            self.F1_C0_coloc_entry_label.grid(row=2,column=5, sticky='w') 
            #F1_COB Analysis Toggle
            self.F1_C0_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[0]}_ROI", variable=self.F1_C0_out_toggle, command=None)
            self.F1_C0_out_toggle_cb.grid(row=2, column=7, sticky='w')
            if self.enableanalysistogs == False: self.F1_C0_out_toggle_cb.grid_remove() 
            #F1_C0 AP Button
            self.F1_C0_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C0 ROIs", command=lambda: self.kp_autopick(0,self.F1_C0_ap_entry.get()))
            self.F1_C0_ap_button.grid(row=2,column=8)
            if self.enableautopicking == False: self.F1_C0_ap_button.grid_remove()
            #F1_C0 AP Remove All Button
            self.F1_C0_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F1_C0 ROIs",command=lambda: self.removeall_chankp("F1_C0"))
            self.F1_C0_ap_remove_button.grid(row=2,column=10)
            
        #F1C1--------------------------------------------------------
        if self.F1_C1_MIP is not None or debugtog == True:
            rb_options.append("F1_C1")
            #F1_C1 AP Label
            #self.F1_C1_autopicker_label = tk.Label(kpchan_frame, text="C1")
            #self.F1_C1_autopicker_label.grid(row=3, column=1, padx=5, pady=5)
            self.F1_C1_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[1]}_ROI", variable=self.F1_C1_toggle_KP, command=self.toggle_kp_visible).grid(row=3, column=1, sticky='w')
            #F1_C1 AP Entry
            self.F1_C1_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C1_ap_entry.insert(0,"0.5")
            self.F1_C1_ap_entry.grid(row=3,column=2, sticky='e')
            self.F1_C1_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=3,column=3, sticky="w") 
            #F1_C1 COLOC ENTRY
            self.F1_C1_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C1_coloc_entry.insert(0, "1.0")
            self.F1_C1_coloc_entry.grid(row=3,column=4,sticky='e')
            self.F1_C1_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=3,column=5, sticky="w") 
            #F1_C1 Analysis Toggle
            self.F1_C1_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[1]}_ROI", variable=self.F1_C1_out_toggle, command=None)
            self.F1_C1_out_toggle_cb.grid(row=3, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F1_C1_out_toggle_cb.grid_remove() 
            #F1_C1 AP Button
            self.F1_C1_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C1 ROIs", command=lambda: self.kp_autopick(1,self.F1_C1_ap_entry.get()))
            self.F1_C1_ap_button.grid(row=3,column=8)
            if self.enableautopicking == False: self.F1_C1_ap_button.grid_remove()
            #F1_C1 AP Remove All Button
            self.F1_C1_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F1_C1 ROIs", command=lambda: self.removeall_chankp("F1_C1")).grid(row=3,column=10)
        #F1C2--------------------------------------------------------
        if self.F1_C2_MIP is not None or debugtog == True:
            rb_options.append("F1_C2")
            #F1_C2 AP Label
            #self.F1_C2_autopicker_label = tk.Label(kpchan_frame, text="F1_C2").grid(row=4, column=1, padx=5, pady=5)
            self.F1_C2_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[2]}_ROI", variable=self.F1_C2_toggle_KP, command=self.toggle_kp_visible).grid(row=4, column=1, sticky='w')
            #F1_C2 AP Entry
            self.F1_C2_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C2_ap_entry.insert(0,"0.5")
            self.F1_C2_ap_entry.grid(row=4,column=2,sticky='e')
            self.F1_C2_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=4,column=3, sticky="w")   
            #F1_C2 COLOC ENTRY
            self.F1_C2_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C2_coloc_entry.insert(0, "1.0")
            self.F1_C2_coloc_entry.grid(row=4,column=4,sticky='e')
            self.F1_C2_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=4,column=5, sticky="w")  
            #F1_C2 Analysis Toggle
            self.F1_C2_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[2]}_ROI", variable=self.F1_C2_out_toggle, command=None)
            self.F1_C2_out_toggle_cb.grid(row=4, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F1_C2_out_toggle_cb.grid_remove() 
            #F1_C2 AP Button
            self.F1_C2_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C2 ROIs", command=lambda: self.kp_autopick(1,self.F1_C2_ap_entry.get()))
            self.F1_C2_ap_button.grid(row=4,column=8)
            if self.enableautopicking == False: self.F1_C2_ap_button.grid_remove()
            #F1_C2 AP Remove All Button
            self.F1_C2_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F1_C2 ROIs", command=lambda: self.removeall_chankp("F1_C2")).grid(row=4,column=10)
        #F1C3--------------------------------------------------------
        if self.F1_C3_MIP is not None or debugtog == True:
            rb_options.append("F1_C3")
            #F1_C3 AP Label
            #self.F1_C3_autopicker_label = tk.Label(kpchan_frame, text="F1_C3").grid(row=5, column=1, padx=5, pady=5)
            self.F1_C3_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[3]}_ROI", variable=self.F1_C3_toggle_KP, command=self.toggle_kp_visible).grid(row=5, column=1, sticky='w')
            #F1_C3 AP Entry
            self.F1_C3_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C3_ap_entry.insert(0, "0.5")
            self.F1_C3_ap_entry.grid(row=5,column=2,sticky='e')
            self.F1_C3_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=5,column=3, sticky="w") 
            #F1_C2R COLOC ENTRY
            self.F1_C3_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C3_coloc_entry.insert(0, "1.0")
            self.F1_C3_coloc_entry.grid(row=5,column=4,sticky='e')
            self.F1_C3_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=5,column=5, sticky="w")  
            #F1_C3 Analysis Toggle
            self.F1_C3_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[3]}_ROI", variable=self.F1_C3_out_toggle, command=None)
            self.F1_C3_out_toggle_cb.grid(row=5, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F1_C3_out_toggle_cb.grid_remove()
            #F1_C3 AP Button
            self.F1_C3_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C3 ROIs", command=lambda: self.kp_autopick(1,self.F1_C3_ap_entry.get()))
            self.F1_C3_ap_button.grid(row=5,column=8)
            if self.enableautopicking == False: self.F1_C3_ap_button.grid_remove()
            #F1_C3 AP Remove All Button
            self.F1_C3_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F1_C3 ROIs", command=lambda: self.removeall_chankp("F1_C3")).grid(row=5,column=10)
        #F1C4--------------------------------------------------------        
        if self.F1_C4_MIP is not None or debugtog == True:
            rb_options.append("F1_C4")
            #F1_C4 AP Label
            #self.F1_C4_autopicker_label = tk.Label(kpchan_frame, text="F1_C4").grid(row=5, column=1, padx=5, pady=5)
            self.F1_C4_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[4]}_ROI", variable=self.F1_C4_toggle_KP, command=self.toggle_kp_visible).grid(row=10, column=1, sticky='w')
            #F1_C4 AP Entry
            self.F1_C4_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C4_ap_entry.insert(0, "0.5")
            self.F1_C4_ap_entry.grid(row=10,column=2,sticky='e')
            self.F1_C4_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=10,column=3, sticky="w") 
            #F1_C2R COLOC ENTRY
            self.F1_C4_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C4_coloc_entry.insert(0, "1.0")
            self.F1_C4_coloc_entry.grid(row=10,column=4,sticky='e')
            self.F1_C4_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=10,column=5, sticky="w")  
            #F1_C4 Analysis Toggle
            self.F1_C4_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[4]}_ROI", variable=self.F1_C4_out_toggle, command=None)
            self.F1_C4_out_toggle_cb.grid(row=10, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F1_C4_out_toggle_cb.grid_remove()
            #F1_C4 AP Button
            self.F1_C4_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C4 ROIs", command=lambda: self.kp_autopick(1,self.F1_C4_ap_entry.get()))
            self.F1_C4_ap_button.grid(row=10,column=8)
            if self.enableautopicking == False: self.F1_C4_ap_button.grid_remove()
            #F1_C4 AP Remove All Button
            self.F1_C4_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F1_C4 ROIs", command=lambda: self.removeall_chankp("F1_C4")).grid(row=10,column=10)
        #F1C5--------------------------------------------------------        
        if self.F1_C5_MIP is not None or debugtog == True:
            rb_options.append("F1_C5")
            #F1_C5 AP Label
            #self.F1_C5_autopicker_label = tk.Label(kpchan_frame, text="F1_C5").grid(row =15, column=1, padx=5, pady=5)
            self.F1_C5_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[5]}_ROI", variable=self.F1_C5_toggle_KP, command=self.toggle_kp_visible).grid(row =15, column=1, sticky='w')
            #F1_C5 AP Entry
            self.F1_C5_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C5_ap_entry.insert(0, "0.5")
            self.F1_C5_ap_entry.grid(row =15,column=2,sticky='e')
            self.F1_C5_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row =15,column=3, sticky="w") 
            #F1_C2R COLOC ENTRY
            self.F1_C5_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F1_C5_coloc_entry.insert(0, "1.0")
            self.F1_C5_coloc_entry.grid(row =15,column=4,sticky='e')
            self.F1_C5_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row =15,column=5, sticky="w")  
            #F1_C5 Analysis Toggle
            self.F1_C5_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[5]}_ROI", variable=self.F1_C5_out_toggle, command=None)
            self.F1_C5_out_toggle_cb.grid(row =15, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F1_C5_out_toggle_cb.grid_remove()
            #F1_C5 AP Button
            self.F1_C5_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C5 ROIs", command=lambda: self.kp_autopick(1,self.F1_C5_ap_entry.get()))
            self.F1_C5_ap_button.grid(row =15,column=8)
            if self.enableautopicking == False: self.F1_C5_ap_button.grid_remove()
            #F1_C5 AP Remove All Button
            self.F1_C5_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F1_C5 ROIs", command=lambda: self.removeall_chankp("F1_C5")).grid(row =15,column=10)
        
        
        #------------------------------------------------------------
        #F2C0--------------------------------------------------------
        if self.F2_C0_MIP is not None or debugtog == True:
            rb_options.append("F2_C0")
            #F2_C0 AP Label
            #self.F2_C0_autopicker_label = tk.Label(kpchan_frame, text="F2_C0").grid(row=2, column=1, padx=5, pady=5)
            self.F2_C0_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[0]}_ROI", variable=self.F2_C0_toggle_KP, command=self.toggle_kp_visible)
            self.F2_C0_KP_toggle_cb.grid(row=20, column=1, sticky='w')
            #F2_C0 AP Entry
            self.F2_C0_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C0_ap_entry.insert(0,"0.5")
            self.F2_C0_ap_entry.grid(row=20,column=2,sticky='e')
            self.F2_C0_ap_entry_label = tk.Label(kpchan_frame, text="um") 
            self.F2_C0_ap_entry_label.grid(row=20,column=3, sticky="w") 
            #F2_COB COLOC ENTRY
            self.F2_C0_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C0_coloc_entry.insert(0, "1.0")
            self.F2_C0_coloc_entry.grid(row=20,column=4,sticky='e')
            self.F2_C0_coloc_entry_label = tk.Label(kpchan_frame, text="um") 
            self.F2_C0_coloc_entry_label.grid(row=20,column=5, sticky='w') 
            #F2_C0_Analysis Toggle
            self.F2_C0_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[0]}_ROI", variable=self.F2_C0_out_toggle, command=None)
            self.F2_C0_out_toggle_cb.grid(row=20, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F2_C0_out_toggle_cb.grid_remove()
            #F2_C0 AP Button
            self.F2_C0_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C3 ROIs", command=lambda: self.kp_autopick(1,self.F2_C0_ap_entry.get()))
            self.F2_C0_ap_button.grid(row=20,column=8)
            if self.enableautopicking == False: self.F2_C0_ap_button.grid_remove()
            #F2_C0 AP Button
            self.F2_C0_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C3 ROIs", command=lambda: self.kp_autopick(1,self.F2_C0_ap_entry.get()))
            self.F2_C0_ap_button.grid(row=20,column=8)
            if self.enableautopicking == False: self.F2_C0_ap_button.grid_remove()
            #F1_C0 AP Button
            self.F2_C0_ap_button = tk.Button(kpchan_frame, text="Autopick F1_C3 ROIs", command=lambda: self.kp_autopick(1,self.F2_C0_ap_entry.get()))
            self.F2_C0_ap_button.grid(row=20,column=8)
            if self.enableautopicking == False: self.F2_C0_ap_button.grid_remove()
            #F2_C0 AP Remove All Button
            self.F2_C0_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F2_C0 ROIs",command=lambda: self.removeall_chankp("F2_C0"))
            self.F2_C0_ap_remove_button.grid(row=20,column=10)
        #F2C1--------------------------------------------------------
        if self.F2_C1_MIP is not None or debugtog==True:
            rb_options.append("F2_C1")
            #F2_C1 AP Label
            #self.F2_C1_autopicker_label = tk.Label(kpchan_frame, text="C1").grid(row=3, column=1, padx=5, pady=5)
            self.F2_C1_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[1]}_ROI", variable=self.F2_C1_toggle_KP, command=self.toggle_kp_visible).grid(row=25, column=1, sticky='w')
            #F2_C1 AP Entry
            self.F2_C1_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C1_ap_entry.insert(0, "0.5")
            self.F2_C1_ap_entry.grid(row=25,column=2,sticky='e')
            self.F2_C1_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=25,column=3, sticky="w") 
            #F2_C1 COLOC ENTRY
            self.F2_C1_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C1_coloc_entry.insert(0, "1.0")
            self.F2_C1_coloc_entry.grid(row=25,column=4,sticky='e')
            self.F2_C1_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=25,column=5, sticky="w")   
            #F2_C1 Analysis Toggle
            self.F2_C1_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[1]}_ROI", variable=self.F2_C1_out_toggle, command=None)
            self.F2_C1_out_toggle_cb.grid(row=25, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F2_C1_out_toggle_cb.grid_remove()
            #F2_C1 AP Button
            self.F2_C1_ap_button = tk.Button(kpchan_frame, text="Autopick F2_C1 ROIs", command=lambda: self.kp_autopick(1,self.F2_C1_ap_entry.get()))
            self.F2_C1_ap_button.grid(row=25,column=8)
            if self.enableautopicking == False: self.F2_C1_ap_button.grid_remove()
            #F2_C1 AP Remove All Button
            self.F2_C1_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F2_C1 ROIs", command=lambda: self.removeall_chankp("F2_C1")).grid(row=25,column=10)
        #F2C2--------------------------------------------------------
        if self.F2_C2_MIP is not None or debugtog==True:
            rb_options.append("F2_C2")
            #F2_C2 AP Label
            #self.F2_C2_autopicker_label = tk.Label(kpchan_frame, text="F2_C2").grid(row=4, column=1, padx=5, pady=5)
            self.F2_C2_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[2]}_ROI", variable=self.F2_C2_toggle_KP, command=self.toggle_kp_visible).grid(row=30, column=1, sticky='w')
            #F2_C2 AP Entry
            self.F2_C2_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C2_ap_entry.insert(0, "0.5")
            self.F2_C2_ap_entry.grid(row=30,column=2,sticky='e')
            self.F2_C2_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=30,column=3, sticky="w") 
            #F2_C2 COLOC ENTRY
            self.F2_C2_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C2_coloc_entry.insert(0, "1.0")
            self.F2_C2_coloc_entry.grid(row=30,column=4,sticky='e')
            self.F2_C2_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=30,column=5, sticky="w")  
            #F2_C2 Analysis Toggle
            self.F2_C2_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[2]}_ROI", variable=self.F2_C2_out_toggle, command=None)
            self.F2_C2_out_toggle_cb.grid(row=30, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F2_C2_out_toggle_cb.grid_remove()
            #F2_C2 AP Button
            self.F2_C2_ap_button = tk.Button(kpchan_frame, text="Autopick F2_C2 ROIs", command=lambda: self.kp_autopick(1,self.F2_C2_ap_entry.get()))
            self.F2_C2_ap_button.grid(row=30,column=8)
            if self.enableautopicking == False: self.F2_C2_ap_button.grid_remove()
            #F2_C2 AP Remove All Button
            self.F2_C2_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F2_C2 ROIs", command=lambda: self.removeall_chankp("F2_C2")).grid(row=30,column=10)
        #F2C3--------------------------------------------------------
        if self.F2_C3_MIP is not None or debugtog==True:
            rb_options.append("F2_C3")
            #C3 AP Label
            #self.F2_C3_autopicker_label = tk.Label(kpchan_frame, text="F2_C3").grid(row=35, column=1, padx=5, pady=5)
            self.F2_C3_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[3]}_ROI", variable=self.F2_C3_toggle_KP, command=self.toggle_kp_visible).grid(row=35, column=1, sticky='w')
            #C3 AP Entry
            self.F2_C3_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C3_ap_entry.insert(0, "0.5")
            self.F2_C3_ap_entry.grid(row=35,column=2,sticky='e')
            self.F2_C3_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=35,column=3, sticky="w") 
            #F2_C3 COLOC ENTRY
            self.F2_C3_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C3_coloc_entry.insert(0, "1.0")
            self.F2_C3_coloc_entry.grid(row=35,column=4,sticky='e')
            self.F2_C3_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=35,column=5, sticky="w")   
            #F2_C3 Analysis Toggle
            self.F2_C3_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[3]}_ROI", variable=self.F2_C3_out_toggle, command=None)
            self.F2_C3_out_toggle_cb.grid(row=35, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F2_C3_out_toggle_cb.grid_remove()
            #F2_C3 AP Button
            self.F2_C3_ap_button = tk.Button(kpchan_frame, text="Autopick F2_C3 ROIs", command=lambda: self.kp_autopick(1,self.F2_C3_ap_entry.get()))
            self.F2_C3_ap_button.grid(row=35,column=8)
            if self.enableautopicking == False: self.F2_C3_ap_button.grid_remove()
            #F2_C3 AP Remove All Button
            self.F2_C3_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F2_C3 ROIs", command=lambda: self.removeall_chankp("F2_C3")).grid(row=35,column=10)
        #F2C4--------------------------------------------------------
        if self.F2_C4_MIP is not None or debugtog==True:
            rb_options.append("F2_C4")
            #C4 AP Label
            #self.F2_C4_autopicker_label = tk.Label(kpchan_frame, text="F2_C4").grid(row=40, column=1, padx=5, pady=5)
            self.F2_C4_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[4]}_ROI", variable=self.F2_C4_toggle_KP, command=self.toggle_kp_visible).grid(row=40, column=1, sticky='w')
            #C4 AP Entry
            self.F2_C4_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C4_ap_entry.insert(0, "0.5")
            self.F2_C4_ap_entry.grid(row=40,column=2,sticky='e')
            self.F2_C4_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=40,column=3, sticky="w") 
            #F2_C4 COLOC ENTRY
            self.F2_C4_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C4_coloc_entry.insert(0, "1.0")
            self.F2_C4_coloc_entry.grid(row=40,column=4,sticky='e')
            self.F2_C4_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=40,column=5, sticky="w")   
            #F2_C4 Analysis Toggle
            self.F2_C4_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[4]}_ROI", variable=self.F2_C4_out_toggle, command=None)
            self.F2_C4_out_toggle_cb.grid(row=40, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F2_C4_out_toggle_cb.grid_remove()
            #F2_C4 AP Button
            self.F2_C4_ap_button = tk.Button(kpchan_frame, text="Autopick F2_C4 ROIs", command=lambda: self.kp_autopick(1,self.F2_C4_ap_entry.get()))
            self.F2_C4_ap_button.grid(row=40,column=8)
            if self.enableautopicking == False: self.F2_C4_ap_button.grid_remove()
            #F2_C4 AP Remove All Button
            self.F2_C4_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F2_C4 ROIs", command=lambda: self.removeall_chankp("F2_C4")).grid(row=40,column=10)
        #F2C5--------------------------------------------------------
        if self.F2_C5_MIP is not None or debugtog==True:
            rb_options.append("F2_C5")
            #C5 AP Label
            #self.F2_C5_autopicker_label = tk.Label(kpchan_frame, text="F2_C5").grid(row=45, column=1, padx=5, pady=5)
            self.F2_C5_KP_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[5]}_ROI", variable=self.F2_C5_toggle_KP, command=self.toggle_kp_visible).grid(row=45, column=1, sticky='w')
            #C5 AP Entry
            self.F2_C5_ap_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C5_ap_entry.insert(0, "0.5")
            self.F2_C5_ap_entry.grid(row=45,column=2,sticky='e')
            self.F2_C5_ap_entry_label = tk.Label(kpchan_frame, text="um").grid(row=45,column=3, sticky="w") 
            #F2_C5 COLOC ENTRY
            self.F2_C5_coloc_entry = tk.Entry(kpchan_frame, width=10)
            self.F2_C5_coloc_entry.insert(0, "1.0")
            self.F2_C5_coloc_entry.grid(row=45,column=4,sticky='e')
            self.F2_C5_coloc_entry_label = tk.Label(kpchan_frame, text="um").grid(row=45,column=5, sticky="w")   
            #F2_C5 Analysis Toggle
            self.F2_C5_out_toggle_cb = tk.Checkbutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[5]}_ROI", variable=self.F2_C5_out_toggle, command=None)
            self.F2_C5_out_toggle_cb.grid(row=45, column=7, sticky='w') 
            if self.enableanalysistogs == False: self.F2_C5_out_toggle_cb.grid_remove()
            #F2_C5 AP Button
            self.F2_C5_ap_button = tk.Button(kpchan_frame, text="Autopick F2_C5 ROIs", command=lambda: self.kp_autopick(1,self.F2_C5_ap_entry.get()))
            self.F2_C5_ap_button.grid(row=45,column=8)
            if self.enableautopicking == False: self.F2_C5_ap_button.grid_remove()
            #F2_C5 AP Remove All Button
            self.F2_C5_ap_remove_button = tk.Button(kpchan_frame, text="Remove ALL F2_C5 ROIs", command=lambda: self.removeall_chankp("F2_C5")).grid(row=45,column=10)



        #Manual KP Picking RadioButton
        for index, option in enumerate(rb_options):
            # Create radio buttons with associated options
            if option == "F1_C0":
                F1C0rb = tk.Radiobutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[0]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F1C0rb.grid(row=2, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F1_C1":
                F1C1rb = tk.Radiobutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[1]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F1C1rb.grid(row=3, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F1_C2":
                F1C2rb = tk.Radiobutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[2]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F1C2rb.grid(row=4, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F1_C3":
                F1C3rb = tk.Radiobutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[3]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F1C3rb.grid(row=5, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F1_C4":
                F1C4rb = tk.Radiobutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[4]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F1C4rb.grid(row=10, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F1_C5":
                F1C5rb = tk.Radiobutton(kpchan_frame, text=f"{self.f1nametag}_{self.f1channels[5]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F1C5rb.grid(row=15, column=0, sticky=tk.W)  # Adjust row number as needed

            elif option == "F2_C0":
                F2C0Brb = tk.Radiobutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[0]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F2C0Brb.grid(row=20, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F2_C1":
                F2C1Grb = tk.Radiobutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[1]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F2C1Grb.grid(row=25, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F2_C2":
                F2C2Rrb = tk.Radiobutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[2]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F2C2Rrb.grid(row=30, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F2_C3":
                F2C3rb = tk.Radiobutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[3]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F2C3rb.grid(row=35, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F2_C4":
                F2C4rb = tk.Radiobutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[4]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F2C4rb.grid(row=40, column=0, sticky=tk.W)  # Adjust row number as needed
            elif option == "F2_C5":
                F2C5rb = tk.Radiobutton(kpchan_frame, text=f"{self.f2nametag}_{self.f2channels[5]}_ROI", variable=self.KP_rb_selection, value=option, command=None)
                F2C5rb.grid(row=45, column=0, sticky=tk.W)  # Adjust row number as needed
 
    #----COLOR_PICK----------------------------------------------
    def pick_color(self,fnum,cnum,square):
        print(f"{fnum} -- {cnum} -- {square}")

        color = colorchooser.askcolor()[1]  # Ask the user to choose a color and get the hex code

        if color == None: return
        if color:
            #color_label.config(foreground=color)  # Set the foreground color of the Label
            print("color chosen")

        print(f"HEX {color}")
        square.config(bg=color)
        self.chancolorarr[cnum] = color

        print(self.chancolorarr)

        # Remove the hash symbol if present
        hex_color = color.lstrip('#')
        # Convert the hex string to RGB tuple
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

        print(f"RGB: {rgb}")

        self.toggle_image_channels()
    def color_nparr(self, in_image, chan):
        hex_color = self.chancolorarr[chan]
        hex_code = hex_color.lstrip('#')
        rgb_color = tuple(int(hex_code[i:i+2], 16) for i in (0, 2, 4))
        grayscale_image = np.dot(in_image[...,:3], [0.2989, 0.5870, 0.1140]).astype(np.uint8)
        colored_image = np.zeros((grayscale_image.shape[0], grayscale_image.shape[1], 3), dtype=np.uint8)
        for i in range(3): colored_image[:, :, i] = grayscale_image * (rgb_color[2-i] / 255.0)
        return colored_image

    #-----SCROLLING CANVAS WITH SCROLLBARS------------------------    
    def scroll_y(self, *args, **kwargs):
        ''' Scroll canvas vertically and redraw the image '''
        self.canvas.yview(*args, **kwargs)  # scroll vertically

        self.toggle_image_channels()  # redraw the image
        #self.display_image(self.modimage)
    def scroll_x(self, *args, **kwargs):
        ''' Scroll canvas horizontally and redraw the image '''
        self.canvas.xview(*args, **kwargs)  # scroll horizontally

        self.toggle_image_channels() # redraw the image
        #self.display_image(self.modimage)
    #-------#DRAG AND ZOOM------------------------------------------------------
    def move_from(self, event):
        ''' Remember previous coordinates for scrolling with the mouse '''
        self.canvas.scan_mark(event.x, event.y)
    def move_to(self, event):
        ''' Drag (move) canvas to the new position '''
        self.canvas.scan_dragto(event.x, event.y, gain=1)
        #
        self.toggle_image_channels()  # redraw the image
        #self.display_image(self.modimage)
    def zoom(self, event):
        print("scrolling")
        ''' Zoom with mouse wheel '''
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        bbox = self.canvas.bbox(self.container)  # get image area
        if bbox[0] < x < bbox[2] and bbox[1] < y < bbox[3]: pass  # Ok! Inside the image
        else: return  # zoom only inside image area
        scale = 1.0
        # Respond to Linux (event.num) or Windows (event.delta) wheel event
        if event.num == 5 or event.delta > 0:  # scroll down
            print("down)")
            i = min(self.width, self.height)
            if int(i * self.imscale) < 30: return  # image is less than 30 pixels
            self.imscale /= self.delta
            scale        /= self.delta
        if event.num == 4 or event.delta < 0:  # scroll up
            print("up")
            i = min(self.canvas.winfo_width(), self.canvas.winfo_height())
            if i < self.imscale: return  # 1 pixel is bigger than the visible area
            self.imscale *= self.delta
            scale        *= self.delta

        print(self.imscale)

        self.canvas.scale('all', x, y, scale, scale)  # rescale all canvas objects
        #self.canvas.scale('all', x, y, scale, scale)  # rescale all canvas objects
        self.toggle_image_channels()
       # self.display_image(self.modimage)

        
    #-------------------------------------------------------------
    def motion(self, event):
        # Calculate the mouse position in the original image
        mouse_x = self.canvas.canvasx(event.x)
        mouse_y = self.canvas.canvasy(event.y)

        mouse_x = self.canvas.canvasx(event.x)/self.imscale
        mouse_y = self.canvas.canvasy(event.y)/self.imscale

       # x_lab = "{:.2f}".format(mouse_x)
       # y_lab = "{:.2f}".format(mouse_y)
        x_lab = float(mouse_x)
        y_lab = float(mouse_y)
        #self.mousepos_label.config(text=f"Mouse Position: ({mouse_x}, {mouse_y})")
       # self.mousepos_label.config(text=f"Mouse Position: ({x_lab:6.2f}, {y_lab:6.2f})")
        self.mousepos_label.config(text=f"Mouse Position: ({float(x_lab):07.2f}, {float(y_lab):07.2f})")

       # self.mousepos_label.config(text=f"Mouse Position: ({x_lab:03}, {y_lab:03})")


    #-------------------------------------------------------
    #END KP PICKING
    def get_ROI_entry_float(self,entry):
        miproi_dict = {
            'F1C0': self.F1_C0_ap_entry if hasattr(self, 'F1_C0_ap_entry') else None,
            'F1C1': self.F1_C1_ap_entry if hasattr(self, 'F1_C1_ap_entry') else None, 
            'F1C2': self.F1_C2_ap_entry if hasattr(self, 'F1_C2_ap_entry') else None,
            'F1C3': self.F1_C3_ap_entry if hasattr(self, 'F1_C3_ap_entry') else None,
            'F1C4': self.F1_C4_ap_entry if hasattr(self, 'F1_C4_ap_entry') else None,
            'F1C5': self.F1_C5_ap_entry if hasattr(self, 'F1_C5_ap_entry') else None,

            'F2C0': self.F2_C0_ap_entry if hasattr(self, 'F2_C0_ap_entry') else None,
            'F2C1': self.F2_C1_ap_entry if hasattr(self, 'F2_C1_ap_entry') else None,
            'F2C2': self.F2_C2_ap_entry if hasattr(self, 'F2_C2_ap_entry') else None,
            'F2C3': self.F2_C3_ap_entry if hasattr(self, 'F2_C3_ap_entry') else None,
            'F2C4': self.F2_C4_ap_entry if hasattr(self, 'F2_C4_ap_entry') else None,
            'F2C5': self.F2_C5_ap_entry if hasattr(self, 'F2_C5_ap_entry') else None
        }

        if miproi_dict[entry]:
            try:
                return float(miproi_dict[entry].get())
            except ValueError:
                return -1
        else:
            return -1
    def get_COLOC_entry_float(self,entry):
        mipcoloc_dict = {
            'F1C0': self.F1_C0_coloc_entry if hasattr(self, 'F1_C0_ap_entry') else None,
            'F1C1': self.F1_C1_coloc_entry if hasattr(self, 'F1_C1_ap_entry') else None,
            'F1C2': self.F1_C2_coloc_entry if hasattr(self, 'F1_C2_ap_entry') else None,
            'F1C3': self.F1_C3_coloc_entry if hasattr(self, 'F1_C3_ap_entry') else None,
            'F1C4': self.F1_C4_coloc_entry if hasattr(self, 'F1_C4_ap_entry') else None,
            'F1C5': self.F1_C5_coloc_entry if hasattr(self, 'F1_C5_ap_entry') else None,

            'F2C0': self.F2_C0_coloc_entry if hasattr(self, 'F2_C0_ap_entry') else None,
            'F2C1': self.F2_C1_coloc_entry if hasattr(self, 'F2_C1_ap_entry') else None,
            'F2C2': self.F2_C2_coloc_entry if hasattr(self, 'F2_C2_ap_entry') else None,
            'F2C3': self.F2_C3_coloc_entry if hasattr(self, 'F2_C3_ap_entry') else None,
            'F2C4': self.F2_C4_coloc_entry if hasattr(self, 'F2_C4_ap_entry') else None,
            'F2C5': self.F2_C5_coloc_entry if hasattr(self, 'F2_C5_ap_entry') else None,            
        }

        if mipcoloc_dict[entry]:
            try:
                return float(mipcoloc_dict[entry].get())
            except ValueError:
                return -1
        else:
            return -1  
    def finalize_kppicking(self):
    
    #    print(f"F1C0 KP -> {self.kp_poly_F1C0_arr}")
    #    print(f"F1C1 KP -> {self.kp_poly_F1C1_arr}")  
    #    print(f"F1C2 KP -> {self.kp_poly_F1C2_arr}")  
     #   print(f"F1C3 KP -> {self.kp_poly_F1C3_arr}") 
     #   print(f"F1C4 KP -> {self.kp_poly_F1C4_arr}")  
     #   print(f"F1C5 KP -> {self.kp_poly_F1C5_arr}") 

    #    print(f"F2C0 KP -> {self.kp_poly_F2C0_arr}")  
     #   print(f"F2C1 KP -> {self.kp_poly_F2C1_arr}")  
    #    print(f"F2C2 KP -> {self.kp_poly_F2C2_arr}")  
     #   print(f"F2C3 KP -> {self.kp_poly_F2C3_arr}")  
    #    print(f"F2C4 KP -> {self.kp_poly_F2C4_arr}")  
    #    print(f"F2C5 KP -> {self.kp_poly_F2C5_arr}") 

    #   print("finished kp picking")

     #   self.update_nuc_count
     #   dyn_data.nucleuscount = self.nuccount

        
        #Set Coordinates back to base
      #  print(f"canvas coords {self.canvas.coords(self.container)}")
     #   print(f"canvas padded coords {self.canvas.coords(self.containerpadded)}")
     #   print(self.bbox)

     #   print(f" padx,pady,imscale {self.padx},{self.pady},{self.imscale}")

        self.F1_exp_arr = []   #array to iterate through for recalculating scaling and offset
        self.F2_exp_arr = []   #array to iterate through for recalculating scaling and offset

        #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
        self.F1_exp_arr.append(self.kp_poly_F1C0_arr) 
        self.F1_exp_arr.append(self.kp_poly_F1C1_arr) 
        self.F1_exp_arr.append(self.kp_poly_F1C2_arr) 
        self.F1_exp_arr.append(self.kp_poly_F1C3_arr) 
        self.F1_exp_arr.append(self.kp_poly_F1C4_arr) 
        self.F1_exp_arr.append(self.kp_poly_F1C5_arr) 

        self.F2_exp_arr.append(self.kp_poly_F2C0_arr) 
        self.F2_exp_arr.append(self.kp_poly_F2C1_arr) 
        self.F2_exp_arr.append(self.kp_poly_F2C2_arr) 
        self.F2_exp_arr.append(self.kp_poly_F2C3_arr) 
        self.F2_exp_arr.append(self.kp_poly_F2C4_arr) 
        self.F2_exp_arr.append(self.kp_poly_F2C5_arr) 

      #  print(f"F1exparr {self.F1_exp_arr}")
      #  print(f"F2exparr {self.F2_exp_arr}")

        
        


        #Move and rescale the canvas and kps
        dx = -self.canvas.coords(self.container)[0]
        dy = -self.canvas.coords(self.container)[1]
        self.canvas.move('all', dx,dy)
        self.canvas.scale('all', 0,0, 1/self.imscale, 1/self.imscale)
  
        


      #  print(f"canvas coords post {self.canvas.coords(self.container)}")
       # print(f"canvas padded coords post {self.canvas.coords(self.containerpadded)}")

        #SHIFT KPS in F1 and F2 into final export array self.f_exp_arr

        self.f_exp_arr = []
        self.f1_draw_arr = [] #for drawing on output images
        self.f2_draw_arr = [] #for drawing on output images
       # print(f"padx,pady {self.padx},{self.pady}")
        #File 1 recalculate kp_x and kp_y
        for chanarr in self.F1_exp_arr:
            temparray = []
            for row in chanarr:
               kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = row #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
        #       print(f"{kp_x},{kp_y}")
        #       print(f"padx,pady {self.padx},{self.pady}")

               kpcoords = self.canvas.coords(kpOvalID)
               t_kpx = (kpcoords[0]+kpcoords[2])/2
               t_kpy = (kpcoords[1]+kpcoords[3])/2
               f_kp_x = t_kpx - abs(self.padx)
               f_kp_y = t_kpy - abs(self.pady)

               if f_kp_x <= 0 or f_kp_y <= 0: pass
               else:
                temprow = kpID,kpOvalID,f_kp_x,f_kp_y,nucInd,polyID,polyTextID
                temparray.append(temprow)

            self.f_exp_arr.append(temparray)
            self.f1_draw_arr.append(temparray)
        #File 2 recalculate kp_x and kp_y
        for chanarr in self.F2_exp_arr:
            temparray = []
            for row in chanarr:
                kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = row #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
        #        print(f"{kpID} - {kp_x},{kp_y}")
        #        print(f"padx,pady {self.padx},{self.pady}")

                kpcoords = self.canvas.coords(kpOvalID)
                t_kpx = (kpcoords[0]+kpcoords[2])/2
                t_kpy = (kpcoords[1]+kpcoords[3])/2
                f_kp_x = t_kpx - abs(self.padx)
                f_kp_y = t_kpy - abs(self.pady)

                if self.padx < 0:
                    f_kp_x = f_kp_x + abs(self.padx)
                if self.padx > 0:
                    f_kp_x = f_kp_x - abs(self.padx) 
                if self.padx == 0:
                    f_kp_x = f_kp_x

                if self.pady < 0:
                    f_kp_y = f_kp_y + abs(self.pady)
                if self.pady > 0:
                    f_kp_y = f_kp_y - abs(self.pady) 
                if self.pady == 0:
                    f_kp_y = f_kp_y

                if f_kp_x <= 0 or f_kp_y <= 0: pass
                else:
                    temprow = kpID,kpOvalID,f_kp_x,f_kp_y,nucInd,polyID,polyTextID
                    temparray.append(temprow)

            self.f_exp_arr.append(temparray)
            self.f2_draw_arr.append(temparray)
        
        #Arrows recalculate coords
        self.arrow_exp_arr = []
        for row in self.arrows_arr:
            temparray = []
            arrowind, arrowid,x1, y1, x2, y2,atextid,atextx,atexty = row #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
            coords = self.canvas.coords(arrowid)
            f_x1 = coords[0] - abs(self.padx)
            f_y1 = coords[1] - abs(self.pady)
            f_x2 = coords[2] - abs(self.padx)
            f_y2 = coords[3] - abs(self.pady)
            temprow = arrowind, arrowid,f_x1, f_y1, f_x2, f_y2,atextid,atextx,atexty
            temparray.append(temprow)

            self.arrow_exp_arr.append(temparray)
     #   print(f"f_exparr {self.f_exp_arr}")
     #   print(f"F1exparr {self.F1_exp_arr}")
     #   print(f"F2exparr {self.F2_exp_arr}")

        #1- EXPORT BUNDLE - KPS PER CHANNEL AND NUCLEI
        dyn_data.kpchan_kpnuc_xbundle = self.f_exp_arr
        #------------------------#
        #2- EXPORT BUNDLE - ROI RADIUS OF EACH CHANNEL
        # NEED TO DO

        self.F1C0_ROI_expvalue = self.get_ROI_entry_float('F1C0') if self.F1_C0_MIP is not None else -1
        self.F1C1_ROI_expvalue = self.get_ROI_entry_float('F1C1') if self.F1_C1_MIP is not None else -1
        self.F1C2_ROI_expvalue = self.get_ROI_entry_float('F1C2') if self.F1_C2_MIP is not None else -1
        self.F1C3_ROI_expvalue = self.get_ROI_entry_float('F1C3') if self.F1_C3_MIP is not None else -1
        self.F1C4_ROI_expvalue = self.get_ROI_entry_float('F1C4') if self.F1_C4_MIP is not None else -1
        self.F1C5_ROI_expvalue = self.get_ROI_entry_float('F1C5') if self.F1_C5_MIP is not None else -1

        self.F2C0_ROI_expvalue = self.get_ROI_entry_float('F2C0') if self.F2_C0_MIP is not None else -1
        self.F2C1_ROI_expvalue = self.get_ROI_entry_float('F2C1') if self.F2_C1_MIP is not None else -1
        self.F2C2_ROI_expvalue = self.get_ROI_entry_float('F2C2') if self.F2_C2_MIP is not None else -1
        self.F2C3_ROI_expvalue = self.get_ROI_entry_float('F2C3') if self.F2_C3_MIP is not None else -1
        self.F2C4_ROI_expvalue = self.get_ROI_entry_float('F2C3') if self.F2_C4_MIP is not None else -1
        self.F2C5_ROI_expvalue = self.get_ROI_entry_float('F2C3') if self.F2_C5_MIP is not None else -1

        self.ROI_xbundle = []
        self.ROI_xbundle.append(self.F1C0_ROI_expvalue)
        self.ROI_xbundle.append(self.F1C1_ROI_expvalue)
        self.ROI_xbundle.append(self.F1C2_ROI_expvalue)
        self.ROI_xbundle.append(self.F1C3_ROI_expvalue)
        self.ROI_xbundle.append(self.F1C4_ROI_expvalue)
        self.ROI_xbundle.append(self.F1C5_ROI_expvalue)

        self.ROI_xbundle.append(self.F2C0_ROI_expvalue)
        self.ROI_xbundle.append(self.F2C1_ROI_expvalue)
        self.ROI_xbundle.append(self.F2C2_ROI_expvalue)
        self.ROI_xbundle.append(self.F2C3_ROI_expvalue)
        self.ROI_xbundle.append(self.F2C4_ROI_expvalue)
        self.ROI_xbundle.append(self.F2C5_ROI_expvalue)

        dyn_data.kpchan_ROIradius_xbundle = self.ROI_xbundle

        #------------------------#
        #3- EXPORT BUNDLE - COLOCALIZATION TOLERANCE CUTOFF
        # NEED TO DO

        self.F1C0_COLOC_expvalue = self.get_COLOC_entry_float('F1C0') if self.F1_C0_MIP is not None else -1
        self.F1C1_COLOC_expvalue = self.get_COLOC_entry_float('F1C1') if self.F1_C1_MIP is not None else -1
        self.F1C2_COLOC_expvalue = self.get_COLOC_entry_float('F1C2') if self.F1_C2_MIP is not None else -1
        self.F1C3_COLOC_expvalue = self.get_COLOC_entry_float('F1C3') if self.F1_C3_MIP is not None else -1
        self.F1C4_COLOC_expvalue = self.get_COLOC_entry_float('F1C4') if self.F1_C4_MIP is not None else -1
        self.F1C5_COLOC_expvalue = self.get_COLOC_entry_float('F1C5') if self.F1_C5_MIP is not None else -1

        self.F2C0_COLOC_expvalue = self.get_COLOC_entry_float('F2C0') if self.F2_C0_MIP is not None else -1
        self.F2C1_COLOC_expvalue = self.get_COLOC_entry_float('F2C1') if self.F2_C1_MIP is not None else -1
        self.F2C2_COLOC_expvalue = self.get_COLOC_entry_float('F2C2') if self.F2_C2_MIP is not None else -1
        self.F2C3_COLOC_expvalue = self.get_COLOC_entry_float('F2C3') if self.F2_C3_MIP is not None else -1
        self.F2C4_COLOC_expvalue = self.get_COLOC_entry_float('F2C4') if self.F2_C4_MIP is not None else -1
        self.F2C5_COLOC_expvalue = self.get_COLOC_entry_float('F2C5') if self.F2_C5_MIP is not None else -1

        self.COLOC_xbundle = []
        self.COLOC_xbundle.append(self.F1C0_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F1C1_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F1C2_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F1C3_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F1C4_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F1C5_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F2C0_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F2C1_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F2C2_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F2C3_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F2C4_COLOC_expvalue)
        self.COLOC_xbundle.append(self.F2C5_COLOC_expvalue)


        dyn_data.kpchan_coloc_xbundle= self.COLOC_xbundle
        #------------------------#
        #4- EXPORT BUNDLE - ANALYSIS TOGGLES OUT
        self.F1C0_outtoggle_expvalue = self.F1_C0_out_toggle.get() if self.F1_C0_MIP is not None else -1
        self.F1C1_outtoggle_expvalue = self.F1_C1_out_toggle.get() if self.F1_C1_MIP is not None else -1
        self.F1C2_outtoggle_expvalue = self.F1_C2_out_toggle.get() if self.F1_C2_MIP is not None else -1
        self.F1C3_outtoggle_expvalue = self.F1_C3_out_toggle.get() if self.F1_C3_MIP is not None else -1
        self.F1C4_outtoggle_expvalue = self.F1_C4_out_toggle.get() if self.F1_C4_MIP is not None else -1
        self.F1C5_outtoggle_expvalue = self.F1_C5_out_toggle.get() if self.F1_C5_MIP is not None else -1  
        self.F2C0_outtoggle_expvalue = self.F2_C0_out_toggle.get() if self.F2_C0_MIP is not None else -1
        self.F2C1_outtoggle_expvalue = self.F2_C1_out_toggle.get() if self.F2_C1_MIP is not None else -1
        self.F2C2_outtoggle_expvalue = self.F2_C2_out_toggle.get() if self.F2_C2_MIP is not None else -1
        self.F2C3_outtoggle_expvalue = self.F2_C3_out_toggle.get() if self.F2_C3_MIP is not None else -1
        self.F2C4_outtoggle_expvalue = self.F2_C4_out_toggle.get() if self.F2_C4_MIP is not None else -1
        self.F2C5_outtoggle_expvalue = self.F2_C5_out_toggle.get() if self.F2_C5_MIP is not None else -1

        self.analysistoggle_xbundle = []
        self.analysistoggle_xbundle.append(self.F1C0_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F1C1_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F1C2_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F1C3_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F1C4_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F1C5_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F2C0_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F2C1_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F2C2_outtoggle_expvalue)
        self.analysistoggle_xbundle.append(self.F2C3_outtoggle_expvalue)        
        self.analysistoggle_xbundle.append(self.F2C4_outtoggle_expvalue)  
        self.analysistoggle_xbundle.append(self.F2C5_outtoggle_expvalue)  

        dyn_data.kpchan_analysistoggle_xbundle= self.analysistoggle_xbundle
        #------------------------#
        #5- EXPORT BUNDLE - ARROWS
        dyn_data.arrows_xbundle = self.arrows_arr #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]

        #------------------------#

        self.final_fullcapture()

      #  print(dyn_data.kpchan_kpnuc_xbundle)  
      #  print(dyn_data.kpchan_ROIradius_xbundle)       
       # print(dyn_data.kpchan_coloc_xbundle)   
      #  print(dyn_data.kpchan_analysistoggle_xbundle)          
        #----
        print("FINISHED KP PICKING")
        mainapp.toCalculations()
        #self.master.destroy()
    #------------------------------------------------------------
    #Capture full screen 
    def final_fullcapture(self):

        #GENERATE FULL SIZED IMG COMPOSITE
        f1f2compimg = self.generate_final_f1f2compimg()
        f1compimg, f1mipbundle = self.generate_final_f1_imgs()
        f2compimg, f2mipbundle = self.generate_final_f2_imgs()

        self.draw_scalebar_imgs(f1f2compimg)
        self.draw_scalebar_imgs(f1compimg)
        self.draw_scalebar_imgs(f2compimg)
        for mip in f1mipbundle: self.draw_scalebar_imgs(mip)
        for mip in f2mipbundle: self.draw_scalebar_imgs(mip)
        #-------------------------------------------------------------------#
        #-------------------------------------------------------------------#
        #-------------------------------------------------------------------#
        #F1 - F2 COMPOSITE --------#
        #-NO WIDGET-#
        export_dir = "Outputs/IMAGES"
        output_name = f"f1f2_composite_NOwidget.tiff"
        output_path = os.path.join(export_dir, output_name)
        f1f2compimg.save(output_path, 'TIFF')
        print(f"Image saved at: {output_path}")
        #-WIDGET-#
        #DRAW F1F2 MIP COMPOSITE
        f1f2WIDGETimg = f1f2compimg.copy()
        self.drawf1f2composite(f1f2WIDGETimg)
        #SAVE the WIDGET VERSION ------
        # Define the export directory and output path
        export_dir = "Outputs/IMAGES"
        output_name = f"f1f2_composite_widget.tiff"
        output_path = os.path.join(export_dir, output_name)
        f1f2WIDGETimg.save(output_path, 'TIFF')
        print(f"Image saved at: {output_path}")
        #-------------------------------------------------------------------#
        #-------------------------------------------------------------------#
        #F1 COMPOSITE
        #-NO WIDGET-#
        export_dir = "Outputs/IMAGES"
        output_name = f"f1composite.tiff"
        output_path = os.path.join(export_dir, output_name)
        f1compimg.save(output_path, 'TIFF')
        print(f"Image saved at: {output_path}")
        #-WIDGET-#
        f1compWIDGETimg = f1compimg.copy()
        self.drawf1composite(f1compWIDGETimg)
        #SAVE the WIDGET VERSION ------
        # Define the export directory and output path
        export_dir = "Outputs/IMAGES"
        output_name = f"f1composite_WIDGET.tiff"
        output_path = os.path.join(export_dir, output_name)
        f1compWIDGETimg.save(output_path, 'TIFF')
        print(f"Image saved at: {output_path}")
        #-------------------------------------------------------------------#
        #F1 SINGLE MIPS
        #-NO WIDGET and WIDGET-#
        self.f1mipsaver(f1mipbundle)
        #-------------------------------------------------------------------#
        #-------------------------------------------------------------------#
        #F2 COMPOSITE
        #-NO WIDGET-#
        export_dir = "Outputs/IMAGES"
        output_name = f"f2composite.tiff"
        output_path = os.path.join(export_dir, output_name)
        f2compimg.save(output_path, 'TIFF')
        print(f"Image saved at: {output_path}")
        #-WIDGET-#
        f2compWIDGETimg = f2compimg.copy()
        self.drawf2composite(f2compWIDGETimg)
        #SAVE the WIDGET VERSION ------
        # Define the export directory and output path
        export_dir = "Outputs/IMAGES"
        output_name = f"f2composite_WIDGET.tiff"
        output_path = os.path.join(export_dir, output_name)
        f2compWIDGETimg.save(output_path, 'TIFF')
        print(f"Image saved at: {output_path}")
        #-------------------------------------------------------------------#
        #F2 SINGLE MIPS
        #-NO WIDGET and WIDGET-#
        self.f2mipsaver(f2mipbundle)
        #-------------------------------------------------------------------#
        self.master.update() 
        return
    
    def f1mipsaver(self,mipbundle):
        chancounter = 0
        for mip in mipbundle:
            #NO WIDGET
            exportmip = mip
            channame = dyn_data.f1channels[chancounter]
            export_dir = "Outputs/IMAGES"
            output_name = f"f1c{chancounter}_{channame}.tiff"
            output_path = os.path.join(export_dir, output_name)
            exportmip.save(output_path, 'TIFF')
            print(f"Image saved at: {output_path}")
            #WIDGET
            exportmip_widget = exportmip.copy()
            self.drawf1mipwidgets(exportmip_widget,chancounter)
            channame = dyn_data.f1channels[chancounter]
            export_dir = "Outputs/IMAGES"
            output_name = f"f1c{chancounter}_{channame}_WIDGET.tiff"
            output_path = os.path.join(export_dir, output_name)
            exportmip_widget.save(output_path, 'TIFF')
            print(f"Image saved at: {output_path}")
            ####
            chancounter +=1

    def f2mipsaver(self,mipbundle):
        chancounter = 0
        for mip in mipbundle:
            #NO WIDGET
            exportmip = mip
            channame = dyn_data.f2channels[chancounter]
            export_dir = "Outputs/IMAGES"
            output_name = f"f2c{chancounter}_{channame}.tiff"
            output_path = os.path.join(export_dir, output_name)
            exportmip.save(output_path, 'TIFF')
            print(f"Image saved at: {output_path}")
            #WIDGET
            exportmip_widget = exportmip.copy()
            self.drawf2mipwidgets(exportmip_widget,chancounter)
            channame = dyn_data.f2channels[chancounter]
            export_dir = "Outputs/IMAGES"
            output_name = f"f2c{chancounter}_{channame}_WIDGET.tiff"
            output_path = os.path.join(export_dir, output_name)
            exportmip_widget.save(output_path, 'TIFF')
            print(f"Image saved at: {output_path}")
            ####
            chancounter +=1

    def drawf2mipwidgets(self,f2mip,chancounter):
        draw = ImageDraw.Draw(f2mip)
        self.labelF2ROIs_singlechan(draw,chancounter)
        object_ids = self.canvas.find_all()   
        shift_x = -abs(self.padx)
        shift_y = -abs(self.pady)
        if self.padx < 0:
            shift_x = (2*abs(self.padx))
        if self.padx > 0:
            shift_x = -(2*abs(self.padx))
        if self.padx == 0:
            shift_x = 0
        if self.pady < 0:
            shift_y = (2*abs(self.pady))
        if self.pady > 0:
            shift_y = -(2*abs(self.pady) )
        if self.pady == 0:
            shift_y = 0
        for obj_id in object_ids:
            obj_type = self.canvas.type(obj_id)
            if "polygon" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                fill_color = None # Default to 'white'
                outline_color = "white"  # Default to 'black'
                # Shift the coordinates
                shifted_coords = [(x + shift_x, y + shift_y) for x, y in zip(coords[::2], coords[1::2])]
                # Flatten the list of tuples back to a single list of coordinates
                shifted_coords_flat = [coord for xy in shifted_coords for coord in xy]
                draw.polygon(shifted_coords_flat, fill=fill_color, outline=outline_color)
            if "polygontext" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                newx = coords[0] + shift_x
                newy = coords[1] + shift_y
                newcoords=[newx,newy]
                id_text = self.canvas.itemcget(obj_id, 'text')
                fill_color = None # Default to 'white'
                ptextfont = ImageFont.load_default()#ImageFont.truetype("arial.ttf", 12)
                draw.text(newcoords, id_text, fill='white', font=ptextfont)

    def drawf1mipwidgets(self,f1mip,chancounter):
        draw = ImageDraw.Draw(f1mip)
        self.labelF1ROIs_singlechan(draw,chancounter)
        object_ids = self.canvas.find_all()   
        shift_x = -abs(self.padx)
        shift_y = - abs(self.pady)
        for obj_id in object_ids:
            obj_type = self.canvas.type(obj_id)
            if "polygon" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                fill_color = None # Default to 'white'
                outline_color = "white"  # Default to 'black'
                # Shift the coordinates
                shifted_coords = [(x + shift_x, y + shift_y) for x, y in zip(coords[::2], coords[1::2])]
                # Flatten the list of tuples back to a single list of coordinates
                shifted_coords_flat = [coord for xy in shifted_coords for coord in xy]
                draw.polygon(shifted_coords_flat, fill=fill_color, outline=outline_color)
            if "polygontext" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                newx = coords[0] + shift_x
                newy = coords[1] + shift_y
                newcoords=[newx,newy]
                id_text = self.canvas.itemcget(obj_id, 'text')
                fill_color = None # Default to 'white'
                ptextfont = ImageFont.load_default()#ImageFont.truetype("arial.ttf", 12)
                draw.text(newcoords, id_text, fill='white', font=ptextfont)

    def drawf2composite(self,f2compimg):
        draw = ImageDraw.Draw(f2compimg)
        self.labelF2ROIs(draw)
        
        object_ids = self.canvas.find_all()   

        shift_x = -abs(self.padx)
        shift_y = -abs(self.pady)

        print(f"padxy {self.padx} - {self.pady}")
        if self.padx < 0:
            shift_x = (2*abs(self.padx))
        if self.padx > 0:
            shift_x = -(2*abs(self.padx))
        if self.padx == 0:
            shift_x = 0

        if self.pady < 0:
            shift_y = (2*abs(self.pady))
        if self.pady > 0:
            shift_y = -(2*abs(self.pady) )
        if self.pady == 0:
            shift_y = 0

        print(f"shiftxy - {shift_x} - {shift_y}")

        for obj_id in object_ids:
            obj_type = self.canvas.type(obj_id)
            if "polygon" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                fill_color = None # Default to 'white'
                outline_color = "white"  # Default to 'black'
                # Shift the coordinates
                shifted_coords = [(x + shift_x, y + shift_y) for x, y in zip(coords[::2], coords[1::2])]
                # Flatten the list of tuples back to a single list of coordinates
                shifted_coords_flat = [coord for xy in shifted_coords for coord in xy]
                draw.polygon(shifted_coords_flat, fill=fill_color, outline=outline_color)
            if "polygontext" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                newx = coords[0] + shift_x
                newy = coords[1] + shift_y
                newcoords=[newx,newy]
                id_text = self.canvas.itemcget(obj_id, 'text')
                fill_color = None # Default to 'white'
                ptextfont = ImageFont.load_default()#ImageFont.truetype("arial.ttf", 12)
                draw.text(newcoords, id_text, fill='white', font=ptextfont)

    def drawf1composite(self,f1compimg):
        draw = ImageDraw.Draw(f1compimg)
        self.labelF1ROIs(draw)
        
        object_ids = self.canvas.find_all()   

        shift_x = -abs(self.padx)
        shift_y = - abs(self.pady)
        for obj_id in object_ids:
            obj_type = self.canvas.type(obj_id)
            if "polygon" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                fill_color = None # Default to 'white'
                outline_color = "white"  # Default to 'black'
                # Shift the coordinates
                shifted_coords = [(x + shift_x, y + shift_y) for x, y in zip(coords[::2], coords[1::2])]
                # Flatten the list of tuples back to a single list of coordinates
                shifted_coords_flat = [coord for xy in shifted_coords for coord in xy]
                draw.polygon(shifted_coords_flat, fill=fill_color, outline=outline_color)
            if "polygontext" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                newx = coords[0] + shift_x
                newy = coords[1] + shift_y
                newcoords=[newx,newy]
                id_text = self.canvas.itemcget(obj_id, 'text')
                fill_color = None # Default to 'white'
                ptextfont = ImageFont.load_default()#ImageFont.truetype("arial.ttf", 12)
                draw.text(newcoords, id_text, fill='white', font=ptextfont)
                    
    def drawf1f2composite(self,compimg):
        object_ids = self.canvas.find_all()   
        draw = ImageDraw.Draw(compimg)
        
        # Get all canvas objects
        object_ids = self.canvas.find_all()
        for obj_id in object_ids:
            obj_type = self.canvas.type(obj_id)

            if "polygon" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                fill_color = None # Default to 'white'
                outline_color = "white"  # Default to 'black'
                draw.polygon(coords, fill=fill_color, outline=outline_color)
            if "polygontext" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                id_text = self.canvas.itemcget(obj_id, 'text')
                fill_color = None # Default to 'white'
                ptextfont = ImageFont.load_default()#ImageFont.truetype("arial.ttf", 12)
                draw.text(coords, id_text, fill='white', font=ptextfont)
            if "arrow" in self.canvas.gettags(obj_id):
                print("arrow")
                coords = self.canvas.coords(obj_id)
                draw.line([(coords[0], coords[1]), (coords[2], coords[3])], fill="white", width=3)
                # Calculate the direction of the arrow shaft
                dx = coords[2] - coords[0]
                dy = coords[3] - coords[1]
                
                # Calculate the angle of the shaft
                angle = math.atan2(dy, dx)
                
                arrowhead_angle = 30
                arrowhead_length = 10
                # Calculate the points for the arrowhead
                angle1 = angle + math.radians(arrowhead_angle)
                angle2 = angle - math.radians(arrowhead_angle)
                
                x1 = coords[2] - arrowhead_length * math.cos(angle1)
                y1 = coords[3] - arrowhead_length * math.sin(angle1)
                
                x2 = coords[2] - arrowhead_length * math.cos(angle2)
                y2 = coords[3] - arrowhead_length * math.sin(angle2)
                
                # Draw the arrowhead (a triangle)
                draw.polygon([(coords[2],coords[3]), (x1, y1), (x2, y2)], fill="white")
            if "arrowtext" in self.canvas.gettags(obj_id):
                coords = self.canvas.coords(obj_id)
                id_text = self.canvas.itemcget(obj_id, 'text')
                fill_color = None # Default to 'white'
                atextfont = ImageFont.load_default()#ImageFont.truetype("arial.ttf", 12)
                draw.text(coords, id_text, fill='magenta', font=atextfont)

        #Now label all the ROIs
        self.labelROIsImage(draw)

    def labelF1ROIs_singlechan(self,draw,chancounter):
        roi_color_dict = {
            
            0:  "orange",       #F1 - DAPI - COB
            1:  "red",          #F1- FITC - C1G
            2:  "green",        #F1 - RED - C2R
            3:  "yellow",       #F1 - FARRED - C3M
            4:  "magenta",        #F2 - DAPI - C0B
            5:  "indigo",       #F2 - FITC - C1G
            6:  "white",       #F2 - RED - C2R
            7:  "cyan",           #F2 - FARRED = C3M
            8:  "purple",
            9:  "pink",
            10: "gray",
            11: "tan",
        }

        ppm = 0.108333333333333 

        #DRAW FILE 1 SINGLE CHAN ROIS
        f1drawarr = []

        if self.f1_draw_arr[chancounter] == []: return
        channame = dyn_data.f1channels[chancounter]
        chanradius = self.ROI_xbundle[chancounter]
        px_radius = chanradius/ppm
        roicolor = roi_color_dict[chancounter]      

        for roirow in self.f1_draw_arr[chancounter]: #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
            kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = roirow
            newroidraw = [kpID,channame,roicolor,kp_x,kp_y,px_radius]
            f1drawarr.append(newroidraw)

        for roi in f1drawarr: #[kpID,channame,roicolor,kp_x,kp_y,px_radius]
            #DRAW CIRCLE
            topleft_x = int(roi[3]- roi[5])
            topleft_y = int(roi[4]- roi[5])
            botright_x = int(roi[3]+ roi[5])
            botright_y = int(roi[4]+ roi[5])
            coords = [(topleft_x,topleft_y),(botright_x,botright_y)]
            draw.ellipse(coords, fill=None, outline=roi[2])

            #DRAW ID
            text= f"1-{roi[1]}-{roi[0]}"
            coords = (int(roi[3]-25),int(roi[4]-25))
            draw.text(coords, text, fill=roi[2], font=ImageFont.load_default())
    def labelF1ROIs(self,draw):
        roi_color_dict = {
            
            0:  "orange",       #F1 - DAPI - COB
            1:  "red",          #F1- FITC - C1G
            2:  "green",        #F1 - RED - C2R
            3:  "yellow",       #F1 - FARRED - C3M
            4:  "magenta",        #F2 - DAPI - C0B
            5:  "indigo",       #F2 - FITC - C1G
            6:  "white",       #F2 - RED - C2R
            7:  "cyan",           #F2 - FARRED = C3M
            8:  "purple",
            9:  "pink",
            10: "gray",
            11: "tan",
        }
        print("f1channels")
        print(dyn_data.f1channels)
        numf1channels = len(dyn_data.f1channels)

        ppm = 0.108333333333333 

        #DRAW FILE 1
        chancounter = 0
        f1drawarr = []
        print(dyn_data.f1channels)
        for channelset in self.f1_draw_arr:     
            print(chancounter)
            print(channelset)
            if channelset == []: 
                chancounter += 1
                continue
            channame = dyn_data.f1channels[chancounter]
            chanradius = self.ROI_xbundle[chancounter]
            px_radius = chanradius/ppm
            roicolor = roi_color_dict[chancounter]
            for roirow in channelset: #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
                kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = roirow
                newroidraw = [kpID,channame,roicolor,kp_x,kp_y,px_radius]
                f1drawarr.append(newroidraw)
            chancounter += 1

        for roi in f1drawarr: #[kpID,channame,roicolor,kp_x,kp_y,px_radius]
            #DRAW CIRCLE
            topleft_x = int(roi[3]- roi[5])
            topleft_y = int(roi[4]- roi[5])
            botright_x = int(roi[3]+ roi[5])
            botright_y = int(roi[4]+ roi[5])
            coords = [(topleft_x,topleft_y),(botright_x,botright_y)]
            draw.ellipse(coords, fill=None, outline=roi[2])

            #DRAW ID
            text= f"1-{roi[1]}-{roi[0]}"
            coords = (int(roi[3]-25),int(roi[4]-25))
            draw.text(coords, text, fill=roi[2], font=ImageFont.load_default())

    def labelF2ROIs_singlechan(self,draw,chancounter):
        roi_color_dict = {
            
            0:  "orange",       #F1 - DAPI - COB
            1:  "red",          #F1- FITC - C1G
            2:  "green",        #F1 - RED - C2R
            3:  "yellow",       #F1 - FARRED - C3M
            4:  "magenta",        #F2 - DAPI - C0B
            5:  "indigo",       #F2 - FITC - C1G
            6:  "white",       #F2 - RED - C2R
            7:  "cyan",           #F2 - FARRED = C3M
            8:  "purple",
            9:  "pink",
            10: "gray",
            11: "tan",
        }

        ppm = 0.108333333333333 

        #DRAW FILE 1 SINGLE CHAN ROIS
        f2drawarr = []
        counterf2 = chancounter+len(dyn_data.f1channels)

        if self.f2_draw_arr[chancounter] == []: return
        channame = dyn_data.f2channels[chancounter]
        chanradius = self.ROI_xbundle[6+chancounter]
        px_radius = chanradius/ppm
        roicolor = roi_color_dict[6+chancounter]      

        print(f"{chanradius} - {px_radius} -  {ppm}")
        for roirow in self.f2_draw_arr[chancounter]: #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
            kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = roirow
            newroidraw = [kpID,channame,roicolor,kp_x,kp_y,px_radius]
            f2drawarr.append(newroidraw)

        for roi in f2drawarr: #[kpID,channame,roicolor,kp_x,kp_y,px_radius]
            #DRAW CIRCLE
            print(f"row {roi}")
            topleft_x = int(roi[3]- roi[5])
            topleft_y = int(roi[4]- roi[5])
            botright_x = int(roi[3]+ roi[5])
            botright_y = int(roi[4]+ roi[5])
            print(f"ROI - {topleft_x} - {topleft_y} - {botright_x} - {botright_y}")
            coords = [(topleft_x,topleft_y),(botright_x,botright_y)]
            draw.ellipse(coords, fill=None, outline=roi[2])

            #DRAW ID
            text= f"2-{roi[1]}-{roi[0]}"
            coords = (int(roi[3]-25),int(roi[4]-25))
            draw.text(coords, text, fill=roi[2], font=ImageFont.load_default())
    def labelF2ROIs(self,draw):
        roi_color_dict = {
            
            0:  "orange",       #F1 - DAPI - COB
            1:  "red",          #F1- FITC - C1G
            2:  "green",        #F1 - RED - C2R
            3:  "yellow",       #F1 - FARRED - C3M
            4:  "magenta",        #F2 - DAPI - C0B
            5:  "indigo",       #F2 - FITC - C1G
            6:  "white",       #F2 - RED - C2R
            7:  "cyan",           #F2 - FARRED = C3M
            8:  "purple",
            9:  "pink",
            10: "gray",
            11: "tan",
        }
        print("f2channels")
        print(dyn_data.f2channels)
        numf2channels = len(dyn_data.f2channels)

        ppm = 0.108333333333333 

        #DRAW FILE 2
        chancounter = 0
        counterf2 = chancounter+len(dyn_data.f1channels)
        f2drawarr = []
        print(dyn_data.f2channels)
        for channelset in self.f2_draw_arr:     
            print(chancounter)
            print(channelset)
            if channelset == []: 
                chancounter += 1
                counterf2 += 1
                continue
            channame = dyn_data.f2channels[chancounter]
            chanradius = self.ROI_xbundle[6+chancounter]
            px_radius = chanradius/ppm
            roicolor = roi_color_dict[6+chancounter]
            for roirow in channelset: #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
                kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = roirow
                newroidraw = [kpID,channame,roicolor,kp_x,kp_y,px_radius]
                f2drawarr.append(newroidraw)
            chancounter += 1
            counterf2 += 1

        for roi in f2drawarr: #[kpID,channame,roicolor,kp_x,kp_y,px_radius]
            #DRAW CIRCLE
            topleft_x = int(roi[3]- roi[5])
            topleft_y = int(roi[4]- roi[5])
            botright_x = int(roi[3]+ roi[5])
            botright_y = int(roi[4]+ roi[5])
            coords = [(topleft_x,topleft_y),(botright_x,botright_y)]
            draw.ellipse(coords, fill=None, outline=roi[2])

            #DRAW ID
            text= f"2-{roi[1]}-{roi[0]}"
            coords = (int(roi[3]-25),int(roi[4]-25))
            draw.text(coords, text, fill=roi[2], font=ImageFont.load_default())


    def labelROIsImage(self,draw):
        roi_color_dict = {
            
            0:  "orange",       #F1 - DAPI - COB
            1:  "red",          #F1- FITC - C1G
            2:  "green",        #F1 - RED - C2R
            3:  "yellow",       #F1 - FARRED - C3M
            4:  "magenta",        #F2 - DAPI - C0B
            5:  "indigo",       #F2 - FITC - C1G
            6:  "white",       #F2 - RED - C2R
            7:  "cyan",           #F2 - FARRED = C3M
            8:  "purple",
            9:  "pink",
            10: "gray",
            11: "tan",
        }
        print("f1channels")
        print(dyn_data.f1channels)
        numf1channels = len(dyn_data.f1channels)
        print("f2channels")
        print(dyn_data.f2channels)
        numf2channels = len(dyn_data.f2channels)

        print(self.f_exp_arr)


        #kp --> each row in f_exp_arr #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
        #roi radius =    #-1 if empty --> self.ROI_xbundle

        ppm = 0.108333333333333 

        #DRAW FILE 1
        chancounter = 0
        f1drawarr = []
        print(dyn_data.f1channels)
        for channelset in self.F1_exp_arr:
            print(chancounter)
            print(channelset)
            if channelset == []: 
                chancounter += 1
                continue
            channame = dyn_data.f1channels[chancounter]
            chanradius = self.ROI_xbundle[chancounter]
            print(f"chanradius {chanradius} - c {chancounter}")
            px_radius = chanradius/ppm
            roicolor = roi_color_dict[chancounter]
            for roirow in channelset: #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
                kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = roirow
                newroidraw = [kpID,channame,roicolor,kp_x,kp_y,px_radius]
                f1drawarr.append(newroidraw)
            chancounter += 1
        print(f1drawarr)
        for roi in f1drawarr: #[kpID,channame,roicolor,kp_x,kp_y,px_radius]
            #DRAW CIRCLE
            topleft_x = int(roi[3]- roi[5])
            topleft_y = int(roi[4]- roi[5])
            botright_x = int(roi[3]+ roi[5])
            botright_y = int(roi[4]+ roi[5])
            coords = [(topleft_x,topleft_y),(botright_x,botright_y)]
            print(f"coords dX {botright_x-topleft_x} dy {botright_y-topleft_y}")
            draw.ellipse(coords, fill=None, outline=roi[2])

            #DRAW ID
            text= f"1-{roi[1]}-{roi[0]}"
            coords = (int(roi[3]-25),int(roi[4]-25))
            draw.text(coords, text, fill=roi[2], font=ImageFont.load_default())


        #DRAW FILE 2
        chancounter = 0
        counterf2 = chancounter+len(dyn_data.f1channels)
        f2drawarr = []
        print(dyn_data.f2channels)
        for channelset in self.F2_exp_arr:
            print(chancounter)
            print(channelset)
            if channelset == []: 
                chancounter += 1
                counterf2 += 1
                continue
            channame = dyn_data.f2channels[chancounter]
            chanradius = self.ROI_xbundle[6+chancounter]
            px_radius = chanradius/ppm
            roicolor = roi_color_dict[6+chancounter]
            for roirow in channelset: #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
                kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = roirow
                newroidraw = [kpID,channame,roicolor,kp_x,kp_y,px_radius]
                f2drawarr.append(newroidraw)
            chancounter += 1
            counterf2 += 1

        for roi in f2drawarr: #[kpID,channame,roicolor,kp_x,kp_y,px_radius]
            #DRAW CIRCLE
            topleft_x = int(roi[3]- roi[5])
            topleft_y = int(roi[4]- roi[5])
            botright_x = int(roi[3]+ roi[5])
            botright_y = int(roi[4]+ roi[5])
            coords = [(topleft_x,topleft_y),(botright_x,botright_y)]
            draw.ellipse(coords, fill=None, outline=roi[2])

            #DRAW ID
            text= f"2-{roi[1]}-{roi[0]}"
            coords = (int(roi[3]-25),int(roi[4]-25))
            draw.text(coords, text, fill=roi[2], font=ImageFont.load_default())

    def generate_final_f1f2compimg(self):
        width_padcont= self.padimgwidth
        height_padcont =  self.padimgheight

        f2padx = int((width_padcont - self.width) /2)
        print(f"{width_padcont} - {self.width} /2 = {int((width_padcont - self.width) /2)}")

        f2pady = int((height_padcont - self.height) /2)


        f1padx = int((abs(f2padx)))
        f1pady = int((abs(f2pady)))

        #F2 OFFSET
        if self.f2_offset[0] < 0:
            f2padx_l = 0
            f2padx_r = 2*f2padx
        if self.f2_offset[0] > 0:
            f2padx_l = 2*f2padx
            f2padx_r = 0
        if self.f2_offset[0] == 0:
            f2padx_l = 0
            f2padx_r = 0

        if self.f2_offset[1] < 0:
            f2pady_u = 0
            f2pady_l = 2*f2pady
        if self.f2_offset[1] > 0:
            f2pady_u = 2*f2pady
            f2pady_l = 0
        if self.f2_offset[1] == 0:
            f2pady_u = 0
            f2pady_l = 0

        # 3 = BGR?
        padded_baseimg = np.zeros((width_padcont, height_padcont, 3), dtype=np.uint8)
        padded_baseimg = padded_baseimg.transpose(1,0,2)    #numpy goes by height/width rather than width, height


        print(f"padxy {f1padx},{f1pady} ---- {padded_baseimg.shape[0]},{padded_baseimg.shape[1]}")
    

        #Make a black image if there are no toggles on 

        padding_f1 = ((f1pady, f1pady), (f1padx, f1padx), (0, 0))
        padding_f2 = ((f2pady_u, f2pady_l), (f2padx_l, f2padx_r), (0, 0))


        if self.nochannelsselected(): return padded_baseimg.transpose(1,0,2)

        added_image = None
        processed_images = [padded_baseimg]
        #FILE1 CHANNELS-----------------------------------------------------
        if self.F1_C0_toggle.get() and self.F1_C0_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C0_MIP, self.f1_C0_brightness_slider, self.f1_C0_contrast_slider, padding_f1, 0, self))
        if self.F1_C1_toggle.get() and self.F1_C1_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C1_MIP, self.f1_C1_brightness_slider, self.f1_C1_contrast_slider, padding_f1, 1, self))
        if self.F1_C2_toggle.get() and self.F1_C2_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C2_MIP, self.f1_C2_brightness_slider, self.f1_C2_contrast_slider, padding_f1, 2, self))
        if self.F1_C3_toggle.get() and self.F1_C3_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C3_MIP, self.f1_C3_brightness_slider, self.f1_C3_contrast_slider, padding_f1, 3, self))
        if self.F1_C4_toggle.get() and self.F1_C4_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C4_MIP, self.f1_C4_brightness_slider, self.f1_C4_contrast_slider, padding_f1, 4, self))
        if self.F1_C5_toggle.get() and self.F1_C5_MIP is not None:
            processed_images.append(self.process_channel(self.F1_C5_MIP, self.f1_C5_brightness_slider, self.f1_C5_contrast_slider, padding_f1, 5, self))
        if self.F2_C0_toggle.get() and self.F2_C0_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C0_MIP, self.f2_C0_brightness_slider, self.f2_C0_contrast_slider, padding_f2, 6, self))
        if self.F2_C1_toggle.get() and self.F2_C1_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C1_MIP, self.f2_C1_brightness_slider, self.f2_C1_contrast_slider, padding_f2, 7, self))
        if self.F2_C2_toggle.get() and self.F2_C2_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C2_MIP, self.f2_C2_brightness_slider, self.f2_C2_contrast_slider, padding_f2, 8, self))
        if self.F2_C3_toggle.get() and self.F2_C3_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C3_MIP, self.f2_C3_brightness_slider, self.f2_C3_contrast_slider, padding_f2, 9, self))
        if self.F2_C4_toggle.get() and self.F2_C4_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C4_MIP, self.f2_C4_brightness_slider, self.f2_C4_contrast_slider, padding_f2, 10, self))
        if self.F2_C5_toggle.get() and self.F2_C5_MIP is not None:
            processed_images.append(self.process_channel(self.F2_C5_MIP, self.f2_C5_brightness_slider, self.f2_C5_contrast_slider, padding_f2, 11, self))
 
        if all(isinstance(img, np.ndarray) for img in processed_images):
            added_image = np.sum(processed_images, axis=0)
            added_image = np.clip(added_image, 0, 255).astype(np.uint8)
        else:
            raise ValueError("One or more elements in added_images_list are not NumPy arrays")

        self.modimage = self.addedimage

        if self.addedimage is not None:
           # self.addedimage = cv2.cvtColor(self.addedimage, cv2.COLOR_BGR2RGB)
            ##self.exportimg = self.addedimage.copy()
            addpil_image = Image.fromarray(self.addedimage)
            self.addimgmask = Image.new("L", addpil_image.size, int(255 * 1))
            self.addimgwopac = addpil_image.copy()
            self.addimgwopac.putalpha(self.addimgmask)      
            addimg_processed= np.array(self.addimgwopac)
            print(f"addimg shape {addimg_processed.shape}")
            addimg_processed = Image.fromarray(addimg_processed)
        
        return addimg_processed

    def generate_final_f1_imgs(self):

        imgheight = self.height
        imgwidth = self.width
        padded_baseimg = np.zeros((imgwidth,imgheight, 3), dtype=np.uint8)
        padded_baseimg = padded_baseimg.transpose(1,0,2)    #numpy goes by height/width rather than width, height
        print(f"baseimg shape = {padded_baseimg.shape[0]},{padded_baseimg.shape[1]}")
    
        #Make a black image if there are no toggles on 
        padding = 0

        if self.nochannelsselected(): return padded_baseimg.transpose(1,0,2)

        self.f1added_image = None
        processed_images = [padded_baseimg]
        f1mips = [] 
        #FILE1 CHANNELS-----------------------------------------------------
        if self.F1_C0_toggle.get() and self.F1_C0_MIP is not None:
            f1c0_img = self.process_channel_nopadding(self.F1_C0_MIP, self.f1_C0_brightness_slider, self.f1_C0_contrast_slider, 0)
            processed_images.append(f1c0_img)
            f1mips.append(f1c0_img)
        if self.F1_C1_toggle.get() and self.F1_C1_MIP is not None:
            f1c1_img = self.process_channel_nopadding(self.F1_C1_MIP, self.f1_C1_brightness_slider, self.f1_C1_contrast_slider, 1)
            processed_images.append(f1c1_img)
            f1mips.append(f1c1_img)
        if self.F1_C2_toggle.get() and self.F1_C2_MIP is not None:
            f1c2_img = self.process_channel_nopadding(self.F1_C2_MIP, self.f1_C2_brightness_slider, self.f1_C2_contrast_slider, 2)
            processed_images.append(f1c2_img)
            f1mips.append(f1c2_img)
        if self.F1_C3_toggle.get() and self.F1_C3_MIP is not None:
            f1c3_img = self.process_channel_nopadding(self.F1_C3_MIP, self.f1_C3_brightness_slider, self.f1_C3_contrast_slider, 3)
            processed_images.append(f1c3_img)
            f1mips.append(f1c3_img)
        if self.F1_C4_toggle.get() and self.F1_C4_MIP is not None:
            f1c4_img = self.process_channel_nopadding(self.F1_C4_MIP, self.f1_C4_brightness_slider, self.f1_C4_contrast_slider, 4)
            processed_images.append(f1c4_img)
            f1mips.append(f1c4_img)
        if self.F1_C5_toggle.get() and self.F1_C5_MIP is not None:
            f1c5_img = self.process_channel_nopadding(self.F1_C5_MIP, self.f1_C5_brightness_slider, self.f1_C5_contrast_slider, 5)
            processed_images.append(f1c5_img)
            f1mips.append(f1c5_img)

        print("PROCESSED SHAPE")
        print(processed_images[0].shape)

        #PROCESS SINGLE-MIPS
        mipexport = []
        for mip in f1mips:
            mip2 = np.clip(mip, 0, 255).astype(np.uint8)
            mip2 = cv2.cvtColor(mip2, cv2.COLOR_BGR2RGB)
            mip2addpil_image = Image.fromarray(mip2)
            mip2addimgmask = Image.new("L", mip2addpil_image.size, int(255 * 1))
            mip2addimgwopac = mip2addpil_image.copy()
            mip2addimgwopac.putalpha(mip2addimgmask)      
            mip2addimg_processed= np.array(mip2addimgwopac)
            print(f"addimg shape F1 {mip2addimg_processed.shape}")
            mip2addimg_processed = Image.fromarray(mip2addimg_processed)
            mipexport.append(mip2addimg_processed)
        #PROCESS COMPOSITE
        if all(isinstance(img, np.ndarray) for img in processed_images):
            self.f1added_image = np.sum(processed_images, axis=0)
            self.f1added_image = np.clip(self.f1added_image, 0, 255).astype(np.uint8)
        else:
            raise ValueError("One or more elements in added_images_list are not NumPy arrays")

        if self.f1added_image is not None:
            self.f1added_image = cv2.cvtColor(self.f1added_image, cv2.COLOR_BGR2RGB)
            ##self.exportimg = self.addedimage.copy()
            addpil_image = Image.fromarray(self.f1added_image)
            addimgmask = Image.new("L", addpil_image.size, int(255 * 1))
            addimgwopac = addpil_image.copy()
            addimgwopac.putalpha(addimgmask)      
            addimg_processed= np.array(addimgwopac)
            print(f"addimg shape F1 {addimg_processed.shape}")
            addimg_processed = Image.fromarray(addimg_processed)
        
        return addimg_processed, mipexport
    def generate_final_f2_imgs(self):

        imgheight = self.height
        imgwidth = self.width
        padded_baseimg = np.zeros((imgwidth,imgheight, 3), dtype=np.uint8)
        padded_baseimg = padded_baseimg.transpose(1,0,2)    #numpy goes by height/width rather than width, height
        print(f"baseimg shape = {padded_baseimg.shape[0]},{padded_baseimg.shape[1]}")
    
        #Make a black image if there are no toggles on 
        padding = 0

        if self.nochannelsselected(): return padded_baseimg.transpose(1,0,2)

        self.f2added_image = None
        processed_images = [padded_baseimg]
        f2mips = [] 
        #FILE2 CHANNELS-----------------------------------------------------
        if self.F2_C0_toggle.get() and self.F2_C0_MIP is not None:
            f2c0_img = self.process_channel_nopadding(self.F2_C0_MIP, self.f2_C0_brightness_slider, self.f2_C0_contrast_slider, 0)
            processed_images.append(f2c0_img)
            f2mips.append(f2c0_img)
        if self.F2_C1_toggle.get() and self.F2_C1_MIP is not None:
            f2c1_img = self.process_channel_nopadding(self.F2_C1_MIP, self.f2_C1_brightness_slider, self.f2_C1_contrast_slider, 1)
            processed_images.append(f2c1_img)
            f2mips.append(f2c1_img)
        if self.F2_C2_toggle.get() and self.F2_C2_MIP is not None:
            f2c2_img = self.process_channel_nopadding(self.F2_C2_MIP, self.f2_C2_brightness_slider, self.f2_C2_contrast_slider, 2)
            processed_images.append(f2c2_img)
            f2mips.append(f2c2_img)
        if self.F2_C3_toggle.get() and self.F2_C3_MIP is not None:
            f2c3_img = self.process_channel_nopadding(self.F2_C3_MIP, self.f2_C3_brightness_slider, self.f2_C3_contrast_slider, 3)
            processed_images.append(f2c3_img)
            f2mips.append(f2c3_img)
        if self.F2_C4_toggle.get() and self.F2_C4_MIP is not None:
            f2c4_img = self.process_channel_nopadding(self.F2_C4_MIP, self.f2_C4_brightness_slider, self.f2_C4_contrast_slider, 4)
            processed_images.append(f2c4_img)
            f2mips.append(f2c4_img)
        if self.F2_C5_toggle.get() and self.F2_C5_MIP is not None:
            f2c5_img = self.process_channel_nopadding(self.F2_C5_MIP, self.f2_C5_brightness_slider, self.f2_C5_contrast_slider, 5)
            processed_images.append(f2c5_img)
            f2mips.append(f2c5_img)

        print("PROCESSED SHAPE")
        print(processed_images[0].shape)

        #PROCESS SINGLE-MIPS
        mipexport = []
        for mip in f2mips:
            mip2 = np.clip(mip, 0, 255).astype(np.uint8)
            mip2 = cv2.cvtColor(mip2, cv2.COLOR_BGR2RGB)
            mip2addpil_image = Image.fromarray(mip2)
            mip2addimgmask = Image.new("L", mip2addpil_image.size, int(255 * 1))
            mip2addimgwopac = mip2addpil_image.copy()
            mip2addimgwopac.putalpha(mip2addimgmask)      
            mip2addimg_processed= np.array(mip2addimgwopac)
            print(f"addimg shape f2 {mip2addimg_processed.shape}")
            mip2addimg_processed = Image.fromarray(mip2addimg_processed)
            mipexport.append(mip2addimg_processed)
        #PROCESS COMPOSITE
        if all(isinstance(img, np.ndarray) for img in processed_images):
            self.f2added_image = np.sum(processed_images, axis=0)
            self.f2added_image = np.clip(self.f2added_image, 0, 255).astype(np.uint8)
        else:
            raise ValueError("One or more elements in added_images_list are not NumPy arrays")

        if self.f2added_image is not None:
            self.f2added_image = cv2.cvtColor(self.f2added_image, cv2.COLOR_BGR2RGB)
            ##self.exportimg = self.addedimage.copy()
            addpil_image = Image.fromarray(self.f2added_image)
            addimgmask = Image.new("L", addpil_image.size, int(255 * 1))
            addimgwopac = addpil_image.copy()
            addimgwopac.putalpha(addimgmask)      
            addimg_processed= np.array(addimgwopac)
            print(f"addimg shape f2 {addimg_processed.shape}")
            addimg_processed = Image.fromarray(addimg_processed)
        
        return addimg_processed, mipexport

    def draw_scalebar_imgs(self,img):
        #.shape is height,width
        img_w,img_h = img.size
        #1px =  0.108333333333333 microns
        draw = ImageDraw.Draw(img)
        # Determine the position of the scale bar relative to the image
        px_per_micron = 0.108333333333333
        scalebar_microns = 10
        scale_bar_width = scalebar_microns / px_per_micron 
        scale_bar_height = 5  # Height of the scale bar rectangle

        # Calculate the coordinates for the scale bar
        scale_bar_x0 = 0 + 20    # Adjust as needed
        scale_bar_y0 = img_h - 20  # Adjust as needed
        scale_bar_x1 = scale_bar_x0 + scale_bar_width   # Adjust as needed
        scale_bar_y1 = scale_bar_y0 + scale_bar_height  # Adjust as needed

        # Draw the scale bar
        coords = [(scale_bar_x0,scale_bar_y0),(scale_bar_x1,scale_bar_y1)]

        draw.rectangle(coords, fill='white', outline=None, width=0)

        #TEXT 
        scale_text_x = scale_bar_x0 + 10
        scale_text_y = scale_bar_y0 - 15  # Adjust as needed
        coords = [scale_text_x,scale_text_y]
        text=f"{scalebar_microns} um ({round(scale_bar_width,1)} px)"
        ptextfont = ImageFont.load_default()#ImageFont.truetype("arial.ttf", 12)
        draw.text(coords, text, fill='white', font=ptextfont)
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class Calculations():
    def __init__(self,master):

        self.master = master
        self.master.title("ZFISHER --- Calculations")
        self.master.geometry("500x500")
        self.master.config(bg="lightblue")

        
        self.starttime = str(datetime.now())
        self.endtime = "???"
        # Configure the grid to center the buttons
        master.grid_rowconfigure(0, weight=1)
        master.grid_rowconfigure(1, weight=1)
        master.grid_rowconfigure(2, weight=1)
        master.grid_rowconfigure(3, weight=1)
        master.grid_columnconfigure(0, weight=1)
        master.grid_columnconfigure(1, weight=1)
        master.grid_columnconfigure(2, weight=1)
        master.grid_columnconfigure(3, weight=1)

        self.titlelabel = tk.Label(master, text="Performing Calculations... ", background="black", foreground="white", font=("Helvetica", 20))
        self.titlelabel.grid(row=0, column=1, sticky="nsew", padx=0, pady=20, columnspan=2)

        self.starttime_label = tk.Label(master, text=f"Start Time: {self.starttime}")
        self.starttime_label.grid(row=1, column=1, sticky="nsew", padx=0, pady=2, columnspan=2)
        self.endtime_label = tk.Label(master, text=f"End Time: {self.endtime}")
        self.endtime_label.grid(row=2, column=1, sticky="nsew", padx=0, pady=2, columnspan=2)        
        #PROGRESS BAR
        self.progress1 = tk.IntVar()
        self.progress_label1 = tk.Label(master, text="PROGRESS: 0%")
        self.progress_label1.grid(row=3, column=1, sticky="nsew")
        self.progressbar1 = ttk.Progressbar(master, orient="horizontal", length=200, mode="determinate", variable=self.progress1)
        self.progressbar1.grid(row=3, column=2, sticky="nsew")



        # Create a queue for progress updates
        self.progress_queue = queue.Queue()

        #INIT - DIRECTORIES
        self.inputs_directory_folder = "Inputs"
        self.inputs_directory_F1 = "FILE_1"
        self.inputs_directory_F2 = "FILE_2"
        self.processing_directory_folder = "Processing"
        self.p_MIP_directory = "MIP"
        self.p_MIP_raw_directory = "RAW_MIP"
        self.p_MIP_masked_directory = "Masked"
        self.p_zslices_directory = "zslices"
        self.outputs_directory_folder = "Outputs"
        self.p_C0_directory = "C0"
        self.p_C1_directory = "C1"
        self.p_C2_directory = "C2"
        self.p_C3_directory = "C3"    
        self.p_C4_directory = "C4"
        self.p_C5_directory = "C5"

        #INIT - GATHER ALL PARAMTERS FOR CALCULATIONS--------------------

        self.kpchan_kpnuc_xbundle = dyn_data.kpchan_kpnuc_xbundle  #for bundle of each channel kp, an array = [kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
        self.kpchan_ROIradius_xbundle = dyn_data.kpchan_ROIradius_xbundle   # need to get in picking
        self.kpchan_coloc_xbundle = dyn_data.kpchan_coloc_xbundle   # need to get in picking
        self.f2_offset = dyn_data.f_offset
        self.f_offset = dyn_data.f_offset #[x,y] between 2 files   --> NOT NEEDED Already corrected in Picking Output
        self.kpanalysistoggles = dyn_data.kpchan_analysistoggle_xbundle #Should we do coloc calculations on this channel
        
        self.F1C0_zslices = []
        self.F1C1_zslices = []
        self.F1C2_zslices = []
        self.F1C3_zslices = []
        self.F1C4_zslices = []
        self.F1C5_zslices = []

        self.F2C0_zslices = []
        self.F2C1_zslices = []
        self.F2C2_zslices = []
        self.F2C3_zslices = []
        self.F2C4_zslices = []
        self.F2C5_zslices = []

        #also need Z-stack spacing
        #also need nm per pixel

        print(self.kpchan_kpnuc_xbundle)


        self.F1_nd2_path = os.path.join(self.inputs_directory_folder,self.inputs_directory_F1,os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F1))[0])
        self.F2_nd2_path = os.path.join(self.inputs_directory_folder,self.inputs_directory_F2,os.listdir(os.path.join(self.inputs_directory_folder,self.inputs_directory_F2))[0])



        self.f1channels = [None,None,None,None,None,None,]
        self.f2channels = [None,None,None,None,None,None,]

        with nd2reader.ND2Reader(self.F1_nd2_path) as nd2:
            self.F1_metadata = nd2.metadata
            print(self.F1_metadata)

            self.f1channels = nd2.metadata['channels']
        

        self.ppm = 1/self.F1_metadata['pixel_microns']
        self.mpp = self.F1_metadata['pixel_microns']
        print(f"Microns per pixel: {self.F1_metadata['pixel_microns']}")
        print(f"Pixels per micron: {self.ppm}")
        print(f"Microns per pixel: {1/self.ppm}")
        F1_z_coordinates = self.F1_metadata['z_coordinates']
        self.F1_z_spacing = abs(F1_z_coordinates[0])-abs(F1_z_coordinates[1])
        print(f"F1 Z spacing uM: {self.F1_z_spacing}")

        with nd2reader.ND2Reader(self.F2_nd2_path) as nd2:
            self.F2_metadata = nd2.metadata
            print(self.F2_metadata)
            F2_z_coordinates = self.F2_metadata['z_coordinates']
            self.F2_z_spacing = abs(F2_z_coordinates[0])-abs(F2_z_coordinates[1])

            self.f2channels = nd2.metadata['channels']

        print(f"F2 Z spacing uM: {self.F2_z_spacing}")

        #-----------------------------
        self.colocarr =  []
        self.noncolocarr = []

        self.threadmain = threading.Thread(target=self.processing_main)
        self.threadmain.start()

 
    def processing_main(self):

        self.thread1 = threading.Thread(target=self.z_slices_loader(self.kpanalysistoggles))
        self.thread1.start()
        self.thread1.join()

        #GET SETS OF ALL PAIR COMBINATIONS FOR ANALYSIS
        self.analysispairs = self.generate_pairs(self.kpanalysistoggles)
        
        self.roi_paircount = len(self.analysispairs)


        #FIND MAX INTENSITY OF EACH PAIR
        
        self.maxZ_finder(self.kpchan_kpnuc_xbundle)

        self.titlelabel.configure(text="Parsing colocalization of ROI max slices")
        self.master.update_idletasks()

        self.thread2 = threading.Thread(target=self.coloc_analysis_parser(self.analysispairs))
        self.thread2.start()
        self.thread2.join()
        
        self.titlelabel.configure(text="Calculations Complete.")
        self.master.update_idletasks()
        self.finishcalculations()
        #self.ocombbut = tk.Button(self.master, text="Export Datasheet/ Finish", command=self.finishcalculations, background="black", foreground='red', font=("Helvetica", 20))
        #self.ocombbut.grid(row=2, column=1, sticky="nsew", padx=30, pady=50)
        


    def finishcalculations(self):
        self.titlelabel.configure(text="Analysis Complete, \n please find output file in \n 'Outputs' directory of specified path",font=("Helvetica", 12))
        #self.ocombbut.grid_remove()
        self.progress_label1.configure(text="100% COMPLETE")
        self.endtime = str(datetime.now())
        self.endtime_label.configure(text=f"End Time: {self.endtime}")
        self.master.config(bg="green")
        self.master.update_idletasks()
        mainapp.toOutputExport()
    #----------------------------------------------
    #LOAD via CV2 AND SORT ZSLICES INTO ARRAY 
    def sort_key(self,filename):
        # Extract the number between "z" and "_.tif" in the filename
        start_index = filename.find('z') + 1
        end_index = filename.find('_.tif')
        number_str = filename[start_index:end_index]
        
        # Convert the extracted number string to an integer
        return int(number_str)  
    
    def z_slices_loader(self,chantogs):
        zslice_dict = {
            0: self.F1C0_zslices,
            1: self.F1C1_zslices,
            2: self.F1C2_zslices,
            3: self.F1C3_zslices,
            4: self.F1C4_zslices,
            5: self.F1C5_zslices,
            6: self.F2C0_zslices,
            7: self.F2C1_zslices,
            8: self.F2C2_zslices,
            9: self.F2C3_zslices,
            10: self.F2C4_zslices,
            11: self.F2C5_zslices
        }
        zslice_dir_dict = {
            0: 'Processing/FILE_1/zslices/c0B',
            1: 'Processing/FILE_1/zslices/c1G',
            2: 'Processing/FILE_1/zslices/c2R',
            3: 'Processing/FILE_1/zslices/c3M',
            4: 'Processing/FILE_2/zslices/c0B',
            5: 'Processing/FILE_2/zslices/c1G',
            6: 'Processing/FILE_2/zslices/c2R',
            7: 'Processing/FILE_2/zslices/c3M',
            8: 'Processing/FILE_2/zslices/c3M',
            9: 'Processing/FILE_2/zslices/c3M',
            10: 'Processing/FILE_2/zslices/c3M',
            11: 'Processing/FILE_2/zslices/c3M',
        }

        print(f"CHAN TOGS {chantogs}")

        self.titlelabel.configure(text="Loading Channel Z Slices for Analysis")
        self.master.config(bg="yellow")
        self.master.update_idletasks()

        #Update the zslice directory list if the channel is toggled

        activechannels = 0
        processedchannels = 0
        for index, i in enumerate(chantogs):
            if i == True:  
                print(f"index {index}")     
                if index < 6:
                    zslice_dir_dict[index] = f'Processing/FILE_1/zslices/C{index}_{self.f1channels[index]}'
                if index >= 6:
                    zslice_dir_dict[index] = f'Processing/FILE_2/zslices/C{index-6}_{self.f2channels[index-6]}'
                activechannels += 1
        for index, i in enumerate(chantogs):
            if i == True:
                chan_all_slices = os.listdir(zslice_dir_dict[index])

                sorted_slices = sorted(chan_all_slices, key=self.sort_key)
                print(sorted_slices)
                for slice in sorted_slices:
                    path = os.path.join(zslice_dir_dict[index],slice)
                    unchanged_slice = cv2.imread(path, cv2.IMREAD_UNCHANGED)
                    #grey_slice = cv2.cvtColor(unchanged_slice,cv2.COLOR_BGR2GRAY)
                    grey_slice = unchanged_slice
                    zslice_dict[index].append(grey_slice)
                # Update progress
                processedchannels += 1
                progress_percent = (processedchannels / activechannels) * 100
                self.progress1.set(progress_percent)
                self.progress_label1.config(text=f"SLICE LOADING PROGRESS: {int(progress_percent)}%")
                self.master.update_idletasks()    
            print(f"Channel {index} --> #slices = {len(zslice_dict[index])}")
    

    #----------------------------------------------
    #GET SETS OF ALL PAIR COMBINATIONS FOR ANALYSIS
    def generate_pairs(self, chantogs):
        i_dict = {
            0 : 'F1C0B',
            1 : 'F1C1G',
            2 : 'F1C2R',
            3 : 'F1C3M',
            4 : 'F2C0B',
            5 : 'F2C1G',
            6 : 'F2C2R',
            7 : 'F2C3M'
        }

        print(chantogs)

        self.titlelabel.configure(text="Generating all possible ROI pairs.")
        self.master.update_idletasks()

        valid_indices = [i for i, value in enumerate(chantogs) if value not in (False, -1)]

        # Generate combinations of pairs from valid indices
        pairs = list(combinations(valid_indices, 2))

        print(pairs)
        return pairs    
    #----------------------------------------------
    #FIND Z MAX SLICE OF ALL KPS
    def maxZ_finder(self, kpchannucbundle):
        zslice_dict = {
            0: self.F1C0_zslices,
            1: self.F1C1_zslices,
            2: self.F1C2_zslices,
            3: self.F1C3_zslices,
            4: self.F1C4_zslices,
            5: self.F1C5_zslices,
            6: self.F2C0_zslices,
            7: self.F2C1_zslices,
            8: self.F2C2_zslices,
            9: self.F2C3_zslices,
            10: self.F2C4_zslices,
            11: self.F2C5_zslices
        }  

        channeldict = {
            0: 'F1C0',
            1: 'F1C1',
            2: 'F1C2',
            3: 'F1C3',
            4: 'F1C4',
            5: 'F1C5',
            6: 'F2C0',
            7: 'F2C1',
            8: 'F2C2',
            9: 'F2C3',
            10: 'F2C4',
            11: 'F2C5'
        }


        self.titlelabel.configure(text="Finding max intensity of each ROI.")
        self.master.config(bg="red")
        self.master.update_idletasks()
        # Calculate total number of keypoints for the progress bar
        total_keypoints = sum(len(chan_i) for chan_i in kpchannucbundle)
        processed_keypoints = 0

        self.kpchan_kpnucxyz_xbundle = [] #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID,z_slice,max_intensity_pixel]
        for index, chan_i in enumerate(kpchannucbundle):
            chanroi_radius = self.kpchan_ROIradius_xbundle[index]
            chanbundle = []
            for row in chan_i:
                kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = row#[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
                pixelradius = math.ceil(chanroi_radius * self.ppm)
                radius = pixelradius
                center = int(kp_x), int(kp_y)

                chan_slices = zslice_dict[index]

                max_intensity =-1
                max_slice_index = ''
                for sliceindex, slice in enumerate(chan_slices, start=1):
                    gray_image_slice = chan_slices[sliceindex-1].copy()
                    y, x = np.ogrid[:gray_image_slice.shape[0], :gray_image_slice.shape[1]]
                    mask_circle = (x - center[0]) ** 2 + (y - center[1]) ** 2 > radius ** 2
                    gray_image_slice[mask_circle] = 0  # Set the area outside the circle to 0
                    roi = gray_image_slice
                    
                    max_value = np.max(roi)

                    print(f" SLICE {sliceindex} - {max_value}")

                    if max_value > max_intensity:
                        max_slice_index = sliceindex
                        max_intensity = max_value
                print(f"max intensity slice {max_slice_index}")
                print(f"max intensity slice counter {max_intensity}")     
    
                temprow = kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID, max_slice_index, max_intensity,channeldict[index]
                chanbundle.append(temprow)
                # Update progress
                processed_keypoints += 1
                progress_percent = (processed_keypoints / total_keypoints) * 100
                self.progress1.set(progress_percent)
                self.progress_label1.config(text=f"MAX INTENSITY ROI PROGRESS: {int(progress_percent)}%")
                self.master.update_idletasks()

            self.kpchan_kpnucxyz_xbundle.append(chanbundle) 
        dyn_data.kpchan_kpnucxyz_xbundle = self.kpchan_kpnucxyz_xbundle
    #----------------------------------------------   
    def coloc_analysis_parser(self,pairspacket):

        completed_pairs = 0
        total_pairs = len(self.analysispairs)

        zslice_dict = {
            0: self.F1C0_zslices,
            1: self.F1C1_zslices,
            2: self.F1C2_zslices,
            3: self.F1C3_zslices,
            4: self.F1C4_zslices,
            5: self.F1C5_zslices,
            6: self.F2C0_zslices,
            7: self.F2C1_zslices,
            8: self.F2C2_zslices,
            9: self.F2C3_zslices,
            10: self.F2C4_zslices,
            11: self.F2C5_zslices
        }  
        
        channeldict = {
            0: 'F1C0',
            1: 'F1C1',
            2: 'F1C2',
            3: 'F1C3',
            4: 'F1C4',
            5: 'F1C5',
            6: 'F2C0',
            7: 'F2C1',
            8: 'F2C2',
            9: 'F2C3',
            10: 'F2C4',
            11: 'F2C5'
        }

        f1spac = self.F1_z_spacing
        f2spac = self.F2_z_spacing

        self.kpchan_kpnuc_xbundle = dyn_data.kpchan_kpnucxyz_xbundle  #for bundle of each channel kp, an array = [kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID,kp_z,max_intens]
        self.kpchan_ROIradius_xbundle = dyn_data.kpchan_ROIradius_xbundle   # need to get in picking
        self.kpchan_coloc_xbundle = dyn_data.kpchan_coloc_xbundle   # need to get in picking       

        #Get each set of pairs
        for pair in pairspacket:
            chan1 = pair[0]
            chan2 = pair[1]
        
            chan1kpset = self.kpchan_kpnuc_xbundle[chan1]
            chan2kpset = self.kpchan_kpnuc_xbundle[chan2]

            chan1roi = self.kpchan_ROIradius_xbundle[chan1]
            chan2roi = self.kpchan_ROIradius_xbundle[chan2]

            f1coloc = self.kpchan_coloc_xbundle[chan1]
            f2coloc = self.kpchan_coloc_xbundle[chan2]

            f2offset = self.f2_offset
            #iterate through each keypoint of chan1

            #Get slice count 
            
            f1slicepardir = 'Processing/FILE_1/zslices'
            f2slicepardir = 'Processing/FILE_2/zslices'
            f1slicecountf = os.listdir(f1slicepardir)[0]
            f2slicecountf = os.listdir(f2slicepardir)[0]

            f1slicecount = len(os.listdir(os.path.join(f1slicepardir,f1slicecountf)))
            f2slicecount = len(os.listdir(os.path.join(f2slicepardir,f2slicecountf)))

            for row in chan1kpset:     # [kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID,kp_z,max_intens]
                a_kpID, a_kpoval, a_kp_x,a_kp_y, a_nucID, a_pid, a_ptid, a_kp_z, a_max_ints,channel = row
                x1 = a_kp_x 
                y1 = a_kp_y
                for row in chan2kpset:
                    b_kpID, b_kpoval, b_kp_x,b_kp_y, b_nucID, b_pid, b_ptid, b_kp_z, b_max_ints,channel = row
                    x2 = b_kp_x + f2offset[0]
                    y2 = b_kp_y + f2offset[1]
                    xdelta = (x2 - x1) * self.mpp
                    ydelta = (y2- y1)  * self.mpp
                    
                    print("SPACCHOICE")
                    print(f"f1 count {f1slicecount}")
                    print(f"f2 count {f2slicecount}")
                    if f2slicecount > f1slicecount: 
                        correctionfactor = f1slicecount/f2slicecount
                        spacchoice = f2spac
                        print(f"SPACCOICE {spacchoice} - f2spac {f2spac}")
                    if f1slicecount > f2slicecount: 
                        correctionfactor = f2slicecount/f1slicecount
                        spacchoice = f1spac
                        print(f"SPACCOICE {spacchoice} - f1spac {f1spac}")
                    if f2slicecount == f1slicecount: 
                        correctionfactor = 1
                        spacchoice = f1spac
                        print(f"EQUAL SPACCOICE {spacchoice} - f1spac {f1spac}")

                    
                    zdelta = (b_kp_z - a_kp_z)* spacchoice * correctionfactor

                    #zdelta = z2-z1 * (small/distance)

                    distance = math.sqrt((xdelta)**2 + (ydelta)**2 + (zdelta)**2) ################# um

                    if f1coloc < f2coloc: 
                        coloc_choice = f1coloc
                        print(f"coloc choice f1 {coloc_choice}")
                    if f2coloc < f1coloc: 
                        coloc_choice = f2coloc
                        print(f"coloc choice f2 {coloc_choice}")
                    if f2coloc == f1coloc: 
                        coloc_choice = f1coloc
                        print(f"coloc choice f1 {coloc_choice}")
                
                    if distance > coloc_choice: 
                        print("NONCOLOCALIZATION FOUND")
                        ntemprow = [a_nucID,distance,channeldict[chan1],a_kpID,a_kp_x,a_kp_y,a_kp_z,a_max_ints,channeldict[chan2],b_kpID,b_kp_x,b_kp_y,b_kp_z,b_max_ints]
                        self.noncolocarr.append(ntemprow)
                    if distance <= coloc_choice:
                        print("COLOCALIZATION FOUND")
                        temprow = [a_nucID,distance,channeldict[chan1],a_kpID,a_kp_x,a_kp_y,a_kp_z,a_max_ints,channeldict[chan2],b_kpID,b_kp_x,b_kp_y,b_kp_z,b_max_ints]
                        self.colocarr.append(temprow)  #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        
        completed_pairs += 1
        progress = int((completed_pairs / total_pairs) * 100)
        #self.progress_queue.put(progress)

        dyn_data.kpchan_colocalized_xbundle = self.colocarr  #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        dyn_data.kpchan_noncolocalized_xbundle = self.noncolocarr #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        print(f"dyn_data noncoloc {dyn_data.kpchan_noncolocalized_xbundle}")
        print(f"dyn_data coloc {dyn_data.kpchan_colocalized_xbundle}")

        print(f"FINISHED with calcs")

       # mainapp.toOutputExport()
    def zslicemaxfinder(self, chan, chan_slices, kp_x, kp_y, roiradius):
        pixelradius = math.ceil(roiradius * self.ppm)
        radius = pixelradius
        center = int(kp_x), int(kp_y)

        roi_arr = []

        max_intensity =-1
        max_slice_index = ''

        for sliceindex, slice in enumerate(chan_slices, start=1):
            gray_image_slice = chan_slices[sliceindex-1].copy()
            y, x = np.ogrid[:gray_image_slice.shape[0], :gray_image_slice.shape[1]]
            mask_circle = (x - center[0]) ** 2 + (y - center[1]) ** 2 > radius ** 2
            gray_image_slice[mask_circle] = 0  # Set the area outside the circle to 0
            roi = gray_image_slice
            
            max_value = np.max(roi)

            print(f" SLICE {sliceindex} - {max_value}")

            if max_value > max_intensity:
                max_slice_index = sliceindex
                max_intensity = max_value




 

        print(f"max intensity slice {max_slice_index}")
        print(f"max intensity slice counter {max_intensity}")     
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class WriteExcel():
    def __init__(self,mainframe):

        self.filename = 'AIPoutput_.xlsx'

        self.data1 = {'Column1': [1, 2, 3], 'Column2': ['A', 'B', 'C']}
        self.data2 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data3 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data4 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data5 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data6 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}

        self.metadata_generator() #TAB 1 - METADATA
        self.ROIpercell_data_generator() #TAB 2 - ROI PER CELL
        self.allrois_generator() #TAB3 - ALL ROIS
        self.noncolocalized_generator() #TAB4 - NONCOLOCALIZED ROIS
        self.colocalized_generator() #TAB5 - COLOZALIZED ROIS
        self.arrows_generator() #TAB6 - COLOZALIZED ROIS

        #TAB 1 - METADATA
        #max_length_data1 = len(self.data1)
        #self.data1_padded = {key: value + [None] * (max_length_data1 - len(value)) for key, value in enumerate(self.data1)}

        #TAB 2 - ROI PER CELL
       # max_length_data2 = len(self.data2)
       # self.data2_padded = {key: value + [None] * (max_length_data2 - len(value)) for key, value in enumerate(self.data2)}

        #TAB3 - ALL ROIS
       # max_length_data3 = len(self.data3)
       # self.data3_padded = {key: value + [None] * (max_length_data3 - len(value)) for key, value in enumerate(self.data3)}

        #TAB4 - NONCOLOCALIZED ROIS
       # max_length_data4 = len(self.data4)
       # self.data4_padded = {key: value + [None] * (max_length_data4 - len(value)) for key, value in enumerate(self.data4)}

        #TAB5 - COLOCALIZED ROIs
       # max_length_data5 = len(self.data5)
       # self.data5_padded = {key: value + [None] * (max_length_data5 - len(value)) for key, value in enumerate(self.data5)}

        self.writesheetwb(self.data1,self.data2,self.data3,self.data4,self.data5,self.data6)

    def metadata_generator(self):   #TAB 1
        #TAB 1
        currentdate = str(datetime.now())
        versionnumber = dyn_data.aip_version_number
        file1_filename = os.listdir('Inputs/FILE_1')[0]
        file2_filename = os.listdir('Inputs/FILE_2')[0]
        nuclei_count = dyn_data.nucleuscount
        foffset = dyn_data.f_offset
        f1nametag = dyn_data.f1nametag
        f2nametag = dyn_data.f2nametag

        f1slicepardir = 'Processing/FILE_1/zslices'
        f2slicepardir = 'Processing/FILE_2/zslices'
        f1slicecountf = os.listdir(f1slicepardir)[0]
        f2slicecountf = os.listdir(f2slicepardir)[0]

        f1slicecount = len(os.listdir(os.path.join(f1slicepardir,f1slicecountf)))
        f2slicecount = len(os.listdir(os.path.join(f2slicepardir,f2slicecountf)))

        atoggles = dyn_data.kpchan_analysistoggle_xbundle
        ROIradii = dyn_data.kpchan_ROIradius_xbundle   
        coloctols = dyn_data.kpchan_coloc_xbundle

        self.data1 = [
            ('DATE', str(currentdate)),
            ('AIP_VERSION#', str(versionnumber)),
            ('File1', str(file1_filename)),
            ('File2', str(file2_filename)),
            ('F1_Nametag', str(f1nametag)),
            ('F2_Nametag', str(f2nametag)),
            ('#Nuclei', str(nuclei_count)),
            ('File2_Offset(pixels)', str(foffset)),
            ('File1_#Slices', str(f1slicecount)),
            ('File2_#Slices', str(f2slicecount)),
            ('F1C0_Toggle', str(atoggles[0])),
            ('F1C1_Toggle', str(atoggles[1])),
            ('F1C2_Toggle', str(atoggles[2])),
            ('F1C3_Toggle', str(atoggles[3])),
            ('F1C4_Toggle', str(atoggles[4])),
            ('F1C5_Toggle', str(atoggles[5])),
            ('F2C0_Toggle', str(atoggles[6])),
            ('F2C1_Toggle', str(atoggles[7])),
            ('F2C2_Toggle', str(atoggles[8])),
            ('F2C3_Toggle', str(atoggles[9])),
            ('F2C4_Toggle', str(atoggles[10])),
            ('F2C5_Toggle', str(atoggles[11])),
            ('F1C0_ROI_Radius(um)', str(ROIradii[0])),
            ('F1C1_ROI_Radius(um)', str(ROIradii[1])),
            ('F1C2_ROI_Radius(um)', str(ROIradii[2])),
            ('F1C3_ROI_Radius(um)', str(ROIradii[3])),
            ('F1C4_ROI_Radius(um)', str(ROIradii[4])),
            ('F1C5_ROI_Radius(um)', str(ROIradii[5])),
            ('F2C0_ROI_Radius(um)', str(ROIradii[6])),
            ('F2C1_ROI_Radius(um)', str(ROIradii[7])),
            ('F2C2_ROI_Radius(um)', str(ROIradii[8])),
            ('F2C3_ROI_Radius(um)', str(ROIradii[9])),
            ('F2C4_ROI_Radius(um)', str(ROIradii[10])),
            ('F2C5_ROI_Radius(um)', str(ROIradii[11])),
            ('F1C0_Colocalization_Tolerance(um)', str(coloctols[0])),
            ('F1C1_Colocalization_Tolerance(um)', str(coloctols[1])),
            ('F1C2_Colocalization_Tolerance(um)', str(coloctols[2])),
            ('F1C3_Colocalization_Tolerance(um)', str(coloctols[3])),
            ('F1C4_Colocalization_Tolerance(um)', str(coloctols[4])),
            ('F1C5_Colocalization_Tolerance(um)', str(coloctols[5])),
            ('F2C0_Colocalization_Tolerance(um)', str(coloctols[6])),
            ('F2C1_Colocalization_Tolerance(um)', str(coloctols[7])),
            ('F2C2_Colocalization_Tolerance(um)', str(coloctols[8])),
            ('F2C3_Colocalization_Tolerance(um)', str(coloctols[9])),
            ('F2C4_Colocalization_Tolerance(um)', str(coloctols[10])),
            ('F2C5_Colocalization_Tolerance(um)', str(coloctols[11]))
        ]
                        
    def ROIpercell_data_generator(self): #TAB 2
        #TAB 2
        nuc_count = dyn_data.nucleuscount
        kpbundle = dyn_data.kpchan_kpnuc_xbundle #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]

        self.data2 =    []

        col1 = 'Nuclei #'

        col2 = 'F1C0 # ROIs'
        col3 = 'F1C1 # ROIs'
        col4 = 'F1C2 # ROIs'
        col5 = 'F1C3 # ROIs'
        col6 = 'F1C4 # ROIs'
        col7 = 'F1C5 # ROIs'

        col8 = 'F2C0 # ROIs'
        col9 = 'F2C1 # ROIs'
        col10 = 'F2C2 # ROIs'
        col11 = 'F2C3 # ROIs'
        col12 = 'F2C4 # ROIs'
        col13 = 'F2C5 # ROIs'

        
        coldict = {
            1: col1,
            2: col2,
            3: col3,
            4: col4,
            5: col5,
            6: col6,
            7: col7,
            8: col8,
            9: col9,
            10: col10,
            11: col11,
            12: col12,
            13: col13
        }
                
        temprow = [col1,col2,col3,col4,col5,col6,col7,col8,col9,col10,col11,col12,col13]
        self.data2.append(temprow)

        for i in range(1,nuc_count+1):
            temprow = [i,0,0,0,0,0,0,0,0,0,0,0,0]
            
            chancounter = 2
            for kpchan_bundle in kpbundle:
                roicounter = 0
                for row in kpchan_bundle:
                    kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID = row
                    if nucInd == i:
                        roicounter += 1
                print(f"chancounter {chancounter} - roi counter {roicounter}")
                temprow[chancounter-1] = roicounter
                chancounter += 1
            self.data2.append(temprow)

       # self.data2.append(col1)
       # self.data2.append(col2)
       # self.data2.append(col3)
       # self.data2.append(col4)
       # self.data2.append(col5)
       # self.data2.append(col6)
       # self.data2.append(col7)
       # self.data2.append(col8)
       # self.data2.append(col9)

       # self.data2 = [[str(cell) for cell in row] for row in self.data2]

    def allrois_generator(self): #TAB 3
        #TAB 3
        #kpbundle = dyn_data.kpchan_kpnuc_xbundle #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID]
        kpxyzbundle = dyn_data.kpchan_kpnucxyz_xbundle #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID,z_slice,max_intensity_pixel,channel]
        self.data3 =    []

        col1 = 'ROI #'
        col2 = 'Nuclei #'
        col3 = 'Channel'
        col4 = 'X_pos'
        col5 = 'Y_pos'
        col6 = 'Z_slice'
        col7 = 'Max Intensity Pixel Value (0-65536)'

        coldict = {
            1: col1,
            2: col2,
            3: col3,
            4: col4,
            5: col5,
            6: col6,
            7: col7
        }

        channeldict = {
            0: 'F1C0B',
            1: 'F1C1G',
            2: 'F1C2R',
            3: 'F1C3M',
            4: 'F2C0B',
            5: 'F2C1G',
            6: 'F2C2R',
            7: 'F2C3M'
        }

        temprow = [col1,col2,col3,col4,col5,col6,col7]
        self.data3.append(temprow)
        for kpchan_bundle in kpxyzbundle:
            roicounter = 0
            for row in kpchan_bundle:
                kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID,kp_z,max_intens,channel = row
                #coldict[1].append(kpID)  
               # coldict[2].append(nucInd)
               ## coldict[3].append(channel)
               # coldict[4].append(kp_x)
               # coldict[5].append(kp_y)
               # coldict[6].append(kp_z)
               # coldict[7].append(max_intens)

                temprow=[kpID,nucInd,channel,kp_x,kp_y,kp_z,max_intens]
                self.data3.append(temprow)

       # self.data3.append(col1)
      #  self.data3.append(col2)
       # self.data3.append(col3)
       # self.data3.append(col4)
       # self.data3.append(col5)
       # self.data3.append(col6)
       # self.data3.append(col7)



    def noncolocalized_generator(self): #TAB 4
        #TAB 4
        self.data4 = []
        kpkpcolocbundle = dyn_data.kpchan_noncolocalized_xbundle #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        
        col1 = 'Nuc_#'
        col2 = 'Distance(um)'
        col3 = 'Channel_A'
        col4 = 'A_ROI_#'
        col5 = 'A_X'
        col6 = 'A_Y'
        col7 = 'A_Zslice'
        col8 = 'A_max_intensity_pixel_value(0-65536)'
        col9 = 'Channel_B'
        col10 = 'B_ROI_#'
        col11 = 'B_X'
        col12 = 'B_Y'
        col13 = 'B_Zslice'
        col14 = 'B_max_intensity_pixel_value(0-65536)'

        coldict = {
            1: col1,
            2: col2,
            3: col3,
            4: col4,
            5: col5,
            6: col6,
            7: col7,
            8: col8,
            9: col9,
            10: col10,
            11: col11,
            12: col12,
            13: col13,
            14: col14
        }

        channeldict = {
            0: 'F1C0B',
            1: 'F1C1G',
            2: 'F1C2R',
            3: 'F1C3M',
            4: 'F2C0B',
            5: 'F2C1G',
            6: 'F2C2R',
            7: 'F2C3M'
        }
        temprow = [col1,col2,col3,col4,col5,col6,col7,col8,col9,col10,col11,col12,col13,col14]
        self.data4.append(temprow)
        for row in kpkpcolocbundle:
            nucInd, distance, chanA, chanA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity = row
            #coldict[1].append(nucInd)  
            #coldict[2].append(distance)
            #coldict[3].append(chanA)
            #coldict[4].append(chanA_IND)
            #coldict[5].append(A_kp_x)
            #coldict[6].append(A_kp_y)
            #coldict[7].append(A_kp_z)
            #coldict[8].append(A_max_intensity)  
            #coldict[9].append(chanB)
            #coldict[10].append(chanB_IND)
            #coldict[11].append(B_kp_x)
            #coldict[12].append(B_kp_y)
            #coldict[13].append(B_kp_z)
            #coldict[14].append(B_max_intensity)
            temprow = [nucInd, distance, chanA, chanA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity,chanB,chanB_IND,B_kp_x,B_kp_y,B_kp_z,B_max_intensity]
            self.data4.append(temprow)

      #  self.data4.append(col1)
       # self.data4.append(col2)
        #self.data4.append(col3)
        #self.data4.append(col4)
        #self.data4.append(col5)
        #self.data4.append(col6)
        #self.data4.append(col7)
        #self.data4.append(col8)
        #self.data4.append(col9)
        #self.data4.append(col10)
        #self.data4.append(col11)
        #self.data4.append(col12)
        #self.data4.append(col13)
        #self.data4.append(col14)

    def colocalized_generator(self): #TAB 5
        #TAB 5
        self.data5 = []
        kpkpcolocbundle = dyn_data.kpchan_colocalized_xbundle #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        
        col1 = 'Nuc_#'
        col2 = 'Distance(um)'
        col3 = 'Channel_A'
        col4 = 'A_ROI_#'
        col5 = 'A_X'
        col6 = 'A_Y'
        col7 = 'A_Zslice'
        col8 = 'A_max_intensity_pixel_value(0-65536)'
        col9 = 'Channel_B'
        col10 = 'B_ROI_#'
        col11 = 'B_X'
        col12 = 'B_Y'
        col13 = 'B_Zslice'
        col14 = 'B_max_intensity_pixel_value(0-65536)'

        coldict = {
            1: col1,
            2: col2,
            3: col3,
            4: col4,
            5: col5,
            6: col6,
            7: col7,
            8: col8,
            9: col9,
            10: col10,
            11: col11,
            12: col12,
            13: col13,
            14: col14
        }

        channeldict = {
            0: 'F1C0B',
            1: 'F1C1G',
            2: 'F1C2R',
            3: 'F1C3M',
            4: 'F2C0B',
            5: 'F2C1G',
            6: 'F2C2R',
            7: 'F2C3M'
        }
        temprow = [col1,col2,col3,col4,col5,col6,col7,col8,col9,col10,col11,col12,col13,col14]
        self.data5.append(temprow)
        for row in kpkpcolocbundle:
            nucInd, distance, chanA, chanA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity = row
          #  coldict[1].append(nucInd)  
           # coldict[2].append(distance)
          #  coldict[3].append(chanA)
          #  coldict[4].append(chanA_IND)
          #  coldict[5].append(A_kp_x)
          #  coldict[6].append(A_kp_y)
           # coldict[7].append(A_kp_z)
         #   coldict[8].append(A_max_intensity)  
           # coldict[9].append(chanB)
          #  coldict[10].append(chanB_IND)
          #  coldict[11].append(B_kp_x)
          #  coldict[12].append(B_kp_y)
          #  coldict[13].append(B_kp_z)
           # coldict[14].append(B_max_intensity)
            temprow = [nucInd, distance, chanA, chanA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity,chanB,chanB_IND,B_kp_x,B_kp_y,B_kp_z,B_max_intensity]
            self.data5.append(temprow)
       # self.data5.append(col1)
      #  self.data5.append(col2)
     #   self.data5.append(col3)
      #  self.data5.append(col4)
      #  self.data5.append(col5)
      #  self.data5.append(col6)
      #  self.data5.append(col7)
      #  self.data5.append(col8)
      #  self.data5.append(col9)
       # self.data5.append(col10)
       # self.data5.append(col11)
       # self.data5.append(col12)
       # self.data5.append(col13)
       # self.data5.append(col14)

    def arrows_generator(self): #TAB 6
        #TAB 6
        self.data6 = []
        arrowsbundle = dyn_data.arrows_xbundle  #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
        
        col1 = 'Arrow Index'
        col2 = 'X1'
        col3 = 'Y1'
        col4 = 'X2'
        col5 = 'Y2'
   
        temprow = [col1,col2,col3,col4,col5]
        self.data6.append(temprow)
        for row in arrowsbundle:
            arrowind, _arrowid,x1, y1, x2, y2,_atextid,_atextx,_atexty = row
            temprow = [arrowind, x1, y1, x2, y2]
            self.data6.append(temprow)



    def writesheetpandas(self, data1, data2, data3, data4, data5):
        with pd.ExcelWriter('Outputs/AIPoutput.xlsx') as writer:
            # Write each DataFrame to a separate sheet
            df1 = pd.DataFrame(data1)
            df1.to_excel(writer, sheet_name='MetaData', index=False)

            df2 = pd.DataFrame(data2)
            df2.to_excel(writer, sheet_name='Channel ROI per Nucleus', index=False)

            df3 = pd.DataFrame(data3)
            df3.to_excel(writer, sheet_name='ALL ROIs', index=False)

            df4 = pd.DataFrame(data4)
            df4.to_excel(writer, sheet_name='NonColocalized ROIs', index=False)
            
            df5 = pd.DataFrame(data5)
            df5.to_excel(writer, sheet_name='Colocalized ROIs', index=False)
        
        sys.exit()

    def writesheetwb(self, data1, data2,data3, data4, data5,data6):
        wb = Workbook()
        # Create a new Workbook

        # Write each data to a separate sheet
        ws1 = wb.active
        ws1.title = "MetaData"
        for row in data1:
            ws1.append(row)

        ws2 = wb.create_sheet(title="Channel ROI per Nucleus")
        for row in data2:
            ws2.append(row)

        ws3 = wb.create_sheet(title="ALL ROIs")
        for row in data3:
            ws3.append(row)

        ws4 = wb.create_sheet(title="NonColocalized ROIs")
        for row in data4:
            ws4.append(row)

        ws5 = wb.create_sheet(title="Colocalized ROIs")
        for row in data5:
            ws5.append(row)

        ws6 = wb.create_sheet(title="Arrows")
        for row in data6:
            ws6.append(row)

        wb.save('Outputs/AIPoutput.xlsx')        
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class DynData():
     def __init__(self):

        #SOFTWARE VERSION NUMBER
        self.aip_version_number = '0.8 - 20240905'

        #FILE 1 and FILE 2 INPUT
        self.f1inputpath = ''
        self.f2inputpath = ''

        self.f1nametag = 'RRNA'
        self.f2nametag = 'DDNA'

        self.f1numslices = "?"
        self.f2numslices = "?"
        self.f1channels = "?"
        self.f2channels = "?"

        #Offset of File2 from File1
        self.f_offset = [0,0]
        
        #[index, i , x, y]      #Index = nucleus, i = contour line id
        self.arr_nucleus_contours = [0,0,0,0] 

        #[Index,polygon] from Nuc Picking
        self.nucpolygons =[0,0]

        #KEYPOINTS
        self.keypoints_array = [0,0]    #from Blob detection to input into selection GUI


        #KEYPOINT-X-Y-POLGON ARRAY
        self.assigned_kp_p_array = [0,0,0,0]    #assigned from Picking, sent to calculations


        #CALCULATIONS EXPORT ARRAY
        self.zsliceexportarray = [0,0,0,0,0,0,0,0]

        #OUTPUT METRICS
        self.nucleuscount = 0

        #CHANNEL KEYPOINTS EXPORT ARRAY from KP_PICKING
        self.kpchan_kpnuc_xbundle = []
        self.kpchan_ROIradius_xbundle = []
        self.kpchan_coloc_xbundle = []
        self.kpchan_analysistoggle_xbundle = []

        self.arrows_xbundle = []

        #MODIFIED KPCHAN with ZSLICE - FROM CALCULATIONS
        self.kpchan_kpnucxyz_xbundle = [] #[kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID,z_slice,max_intensity_pixel]
        self.kpchan_noncolocalized_xbundle = [] #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        self.kpchan_colocalized_xbundle = [] #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        
        self.f1zspacing = 0
        self.f2zspacing = 0

        #CROPPED Zslice Arrays from REGISTERXY CROPPING
        self.f1_Zstack_cropped = []
        self.f2_Zstack_cropped = []
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
if __name__ == "__main__":
    dyn_data = DynData() 

    #INJECTION FOR DEVELOPMENT
    dyn_data.f_offset = [32.497,]

    root = tk.Tk()
    mainapp = MainApplication(root)
    root.withdraw()
    root.mainloop()