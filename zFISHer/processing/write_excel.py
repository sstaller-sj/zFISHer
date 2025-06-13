#import zFISHer.config.config_manager as cfgmgr
import zFISHer.utils.config as cfg
import zFISHer.data.parameters as aparams

import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl import cell
from openpyxl.cell import _writer
from openpyxl import writer



#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
class WriteExcel():
    def __init__(self,logger):
        self.logger = logger

        self.filename = 'AIPoutput_.xlsx'

        self.data1 = {'Column1': [1, 2, 3], 'Column2': ['A', 'B', 'C']}
        self.data2 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data3 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data4 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data5 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}
        self.data6 = {'Column3': [4, 5, 6], 'Column4': ['D', 'E', 'F']}

        self.metadata_generator() #TAB 1 - METADATA
        self.ROIpercell_data_generator() #TAB 2 - ROI PER CELL
        self.allrois_generator() #TAB3 - ALL ROIS
        self.noncolocalized_generator() #TAB4 - NONCOLOCALIZED ROIS
        self.colocalized_generator() #TAB5 - COLOZALIZED ROIS
        self.arrows_generator() #TAB6 - COLOZALIZED ROIS
        self.channel_metadata_generator()

    def channel_metadata_generator(self):
        self.data7 = []

        # Combine all channels with unique keys like base:ch0 and moving:ch0
        all_channels = {}
        for role in ["fixed", "moving"]:
            for ch_name, metadata in cfg.CH_METADATA.get(role, {}).items():
                unique_key = f"{ch_name}:{role}:"
                all_channels[unique_key] = metadata

        # Collect all unique metadata keys
        all_keys = set()
        for ch_data in all_channels.values():
            all_keys.update(ch_data.keys())
        all_keys = sorted(all_keys)  # Optional: keep keys sorted

        # First row: header with unique channel names
        header = ["Metadata Key"] + list(all_channels.keys())
        self.data7.append(header)

        # Each row corresponds to one metadata key
        for key in all_keys:
            row = [key]
            for ch in all_channels.keys():
                val = all_channels[ch].get(key, "")
                row.append(val)
            self.data7.append(row)

        #TAB 1 - METADATA
        #max_length_data1 = len(self.data1)
        #self.data1_padded = {key: value + [None] * (max_length_data1 - len(value)) for key, value in enumerate(self.data1)}

        #TAB 2 - ROI PER CELL
       # max_length_data2 = len(self.data2)
       # self.data2_padded = {key: value + [None] * (max_length_data2 - len(value)) for key, value in enumerate(self.data2)}

        #TAB3 - ALL ROIS
       # max_length_data3 = len(self.data3)
       # self.data3_padded = {key: value + [None] * (max_length_data3 - len(value)) for key, value in enumerate(self.data3)}

        #TAB4 - NONCOLOCALIZED ROIS
       # max_length_data4 = len(self.data4)
       # self.data4_padded = {key: value + [None] * (max_length_data4 - len(value)) for key, value in enumerate(self.data4)}

        #TAB5 - COLOCALIZED ROIs
       # max_length_data5 = len(self.data5)
       # self.data5_padded = {key: value + [None] * (max_length_data5 - len(value)) for key, value in enumerate(self.data5)}

        self.writesheetwb(self.data1,self.data2,self.data3,self.data4,self.data5,self.data6,self.data7)

    def metadata_generator(self):   #TAB 1
        #TAB 1
        currentdate = str(datetime.now())
        versionnumber = 1.0
        file1_filename = "FILE 1" #TODO FILE 1 NAME
        file2_filename = "FILE 2" #TODO FILE 2 NAME
        nuclei_count = aparams.seg_nuc_count
        foffset = [cfg.OFFSET_X,cfg.OFFSET_Y]
        f1nametag = cfg.F1_NTAG
        f2nametag = cfg.F2_NTAG


 

        f1slicecount = cfg.F1_Z_NUM
        f2slicecount = cfg.F2_Z_NUM

       # atoggles = dyn_data.kpchan_analysistoggle_xbundle
        ROIradii = aparams.kpchan_ROIradius_xbundle   
        coloctols = aparams.kpchan_coloc_xbundle

        # Random colocalization percent
        rand_coloc_perc = cfg.RAND_COLOC_PERCENT

        self.data1 = [
            ('DATE', str(currentdate)),
            ('AIP_VERSION#', str(versionnumber)),
            ('File1', str(file1_filename)),
            ('File2', str(file2_filename)),
            ('F1_Nametag', str(f1nametag)),
            ('F2_Nametag', str(f2nametag)),
            ('#Nuclei', str(nuclei_count)),
            ('File2_Offset(pixels)', str(foffset)),
            ('File1_#Slices', str(f1slicecount)),
            ('File2_#Slices', str(f2slicecount)),
            ('Channel Radii(um)', str(ROIradii)),
            ('Coloc Cutoffs(um)', str(coloctols)),
            ('Random Colocalization %', str(rand_coloc_perc))

        ]
                        
    def ROIpercell_data_generator(self): #TAB 2
        #TAB 2
        nuc_count = aparams.seg_nuc_count
        in_bundle_ROI = aparams.kpchan_kpnucxyz_xbundle #file_id,channel,chanrad,chan_e,kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID, max_slice_index, max_intensity

        f1_cs = cfg.F1_C_LIST
        f2_cs = cfg.F2_C_LIST
        f1_ntag = cfg.F1_NTAG
        f2_ntag = cfg.F2_NTAG

        # Create dictionary that will have each nuclei polygon as a key.
        nuclei_collection ={}


        # Create header for excel sheet
        header =["Nuclei #"]
        # Step 1: Create an entry for each nucleus based on nuc_count
        for i in range(1, nuc_count + 1):  # Nucleus IDs: "nucleus_1", "nucleus_2", ..., "nucleus_n"
            nucInd = i
            
            # Initialize the dictionary for this nucleus, starting with all channels set to 0
            nuclei_collection[nucInd] = {}

            # Initialize channels for this nucleus from both f1_cs and f2_cs
            for channel in f1_cs:
                chankey = f"FILE_1_{channel}"
                nuclei_collection[nucInd][chankey] = 0
                header.append(chankey)
            for channel in f2_cs:
                chankey = f"FILE_2_{channel}"
                nuclei_collection[nucInd][chankey] = 0
                header.append(chankey)

        # Step 2: count each ROI per channel per nuclei
        for row in in_bundle_ROI:
            file = row[0]
            channel = row[1]
            nucInd = row[8]

            chankey = f"FILE_{file}_{channel}"

            # Check if the nucleus exists in the collection
            if nucInd in nuclei_collection:
                # Check if the channel exists in the corresponding nucleus
                if channel in nuclei_collection[nucInd]:
                    # Increment the count for this channel in the nucleus
                    nuclei_collection[nucInd][channel] += 1
                else:
                    # If the channel doesn't exist for this nucleus, initialize it to 0
                    nuclei_collection[nucInd][channel] = 0
            else:
                # If the nucleus doesn't exist, initialize it with the channel, starting from 0
                nuclei_collection[nucInd] = {channel: 0}

            # Now, increment the count of the channel regardless (set it to 1)
            nuclei_collection[nucInd][channel] += 1

            print("FINISHED NUCLEI COUNTING")
            print(nuclei_collection)


        # Populate array to send for writing to Excel
        self.data2 =    []
        self.data2.append(header)
        for i in range(1, nuc_count + 1):  # Nucleus IDs: "nucleus_1", "nucleus_2", ..., "nucleus_n"
            nucInd = i
            # Initialize the dictionary for this nucleus, starting with all channels set to 0
            newrow = [nucInd]
            # Initialize channels for this nucleus from both f1_cs and f2_cs
            for channel in f1_cs:
                chankey = f"FILE_1_{channel}"
                c_count = nuclei_collection[nucInd][chankey]
                newrow.append(c_count)
            for channel in f2_cs:
                chankey = f"FILE_2_{channel}"
                c_count = nuclei_collection[nucInd][chankey]
                newrow.append(c_count)
            self.data2.append(newrow)

    def allrois_generator(self): #TAB 3
        #TAB 3
        in_bundle_ROI = aparams.kpchan_kpnucxyz_xbundle #file_id,channel,chanrad,chan_e,kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID, max_slice_index, max_intensity
        self.data3 =    []

        col1 = 'ROI #'
        col2 = 'Nuclei #'
        col3 = 'Channel'
        col4 = 'X_pos'
        col5 = 'Y_pos'
        col6 = 'Z_slice'
        col7 = 'Max Intensity Pixel Value (0-65536)'

        temprow = [col1,col2,col3,col4,col5,col6,col7]
        self.data3.append(temprow)
        for row in in_bundle_ROI:
            file_id,channel,chanrad,chan_e,kpID,kpOvalID,kp_x,kp_y,nucInd,polyID,polyTextID, max_slice_index, max_intensity = row
            temprow=[kpID,nucInd,channel,kp_x,kp_y,max_slice_index,max_intensity]
            self.data3.append(temprow)

    def noncolocalized_generator(self): #TAB 4
        #TAB 4
        self.data4 = []
        kpkpcolocbundle = aparams.no_arr #
        
        col1 = 'Nuc_#'
        col2 = 'Distance(um)'
        col3 = 'Channel_A'
        col4 = 'A_ROI_#'
        col5 = 'A_X'
        col6 = 'A_Y'
        col7 = 'A_Zslice'
        col8 = 'A_max_intensity_pixel_value(0-65536)'
        col9 = 'Channel_B'
        col10 = 'B_ROI_#'
        col11 = 'B_X'
        col12 = 'B_Y'
        col13 = 'B_Zslice'
        col14 = 'B_max_intensity_pixel_value(0-65536)'


        temprow = [col1,col2,col3,col4,col5,col6,col7,col8,col9,col10,col11,col12,col13,col14]
        self.data4.append(temprow)
        for row in kpkpcolocbundle:
            nucInd, distance, c_A, c_A_roi_id, r1_x, r1_y, r1_zslice, r1_max, c_B, c_B_roi_id, r2_x, r2_y, r2_zslice, r2_max = row


            temprow = [nucInd, distance, c_A, c_A_roi_id, r1_x, r1_y, r1_zslice, r1_max, c_B, c_B_roi_id, r2_x, r2_y, r2_zslice, r2_max]
            self.data4.append(temprow)


    def colocalized_generator(self): #TAB 5
        #TAB 5
        self.data5 = []
        kpkpcolocbundle = aparams.yes_arr #[nucInd, distance, chanA, chandA_IND, A_kp_x,A_kp_y,A_kp_z, A_max_intensity, chanB, chanB_IND, B_kp_x,B_kp_y,B_kp_z, B_max_intensity]
        
        col1 = 'Nuc_#'
        col2 = 'Distance(um)'
        col3 = 'Channel_A'
        col4 = 'A_ROI_#'
        col5 = 'A_X'
        col6 = 'A_Y'
        col7 = 'A_Zslice'
        col8 = 'A_max_intensity_pixel_value(0-65536)'
        col9 = 'Channel_B'
        col10 = 'B_ROI_#'
        col11 = 'B_X'
        col12 = 'B_Y'
        col13 = 'B_Zslice'
        col14 = 'B_max_intensity_pixel_value(0-65536)'



        temprow = [col1,col2,col3,col4,col5,col6,col7,col8,col9,col10,col11,col12,col13,col14]
        self.data5.append(temprow)
        for row in kpkpcolocbundle:
            nucInd, distance, c_A, c_A_roi_id, r1_x, r1_y, r1_zslice, r1_max, c_B, c_B_roi_id, r2_x, r2_y, r2_zslice, r2_max = row


            temprow = [nucInd, distance, c_A, c_A_roi_id, r1_x, r1_y, r1_zslice, r1_max, c_B, c_B_roi_id, r2_x, r2_y, r2_zslice, r2_max]
            self.data5.append(temprow)

    def arrows_generator(self): #TAB 6
        #TAB 6
        self.data6 = []
        arrowsbundle = aparams.arrows_xbundle  #[arrowIndex,arrowID,x1,y1,x2,y2,arrowTextID,text_x,text_y]
        
        col1 = 'Arrow Index'
        col2 = 'X1'
        col3 = 'Y1'
        col4 = 'X2'
        col5 = 'Y2'
   
        temprow = [col1,col2,col3,col4,col5]
        self.data6.append(temprow)
        for row in arrowsbundle:
            arrowind, _arrowid,x1, y1, x2, y2,_atextid,_atextx,_atexty = row
            temprow = [arrowind, x1, y1, x2, y2]
            self.data6.append(temprow)


    def writesheetwb(self, data1, data2,data3, data4, data5,data6,data7):
        wb = Workbook()
        # Create a new Workbook

        # Write each data to a separate sheet
        ws1 = wb.active
        ws1.title = "MetaData"
        for row in data1:
            ws1.append(row)

        ws2 = wb.create_sheet(title="Channel Metadata")
        for row in data7:
            ws2.append(row)

        ws3 = wb.create_sheet(title="Channel ROI per Nucleus")
        for row in data2:
            ws3.append(row)

        ws4 = wb.create_sheet(title="ALL ROIs")
        for row in data3:
            ws4.append(row)

        ws5 = wb.create_sheet(title="NonColocalized ROIs")
        for row in data4:
            ws5.append(row)

        ws6 = wb.create_sheet(title="Colocalized ROIs")
        for row in data5:
            ws6.append(row)

        ws7 = wb.create_sheet(title="Arrows")
        for row in data6:
            ws7.append(row)

        output_dir = cfg.OUTPUT_DIR
        file_name = 'AIP_output.xlsx'
        wb.save(os.path.join(output_dir,file_name))        
#-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
